# -*- coding: utf-8 -*-
"""
backlog-xlsx / create_evidence.py
エビデンス.xlsx を生成するスクリプト (GF-327 ビジュアル互換版)

Usage:
    python create_evidence.py <folder> <issue_id>
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)


# ================= スタイルパレット =================
TITLE_FILL  = PatternFill("solid", fgColor="1F4E79")
SEC_FILL    = PatternFill("solid", fgColor="D6E4F0")
HDR_FILL    = PatternFill("solid", fgColor="2E75B6")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
STRIPE_FILL = PatternFill("solid", fgColor="F2F7FB")
PASTE_FILL  = PatternFill("solid", fgColor="FAFAFA")  # 貼付枠（超薄グレー）

TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
SEC_FONT   = Font(color="000000", bold=True, size=10)
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
CELL_FONT  = Font(size=10)
BOLD_FONT  = Font(bold=True, size=10)

TITLE_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
SEC_ALIGN   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
HDR_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
PASTE_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN  = Side(style="thin",   color="000000")
_MEDIUM = Side(style="medium", color="808080")
BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
PASTE_BORDER  = Border(left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM)


def _style(cell, fill, font, align, border=None):
    cell.fill = fill
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = border


def main():
    if len(sys.argv) < 3:
        print("Usage: python create_evidence.py <folder> <issue_id>")
        sys.exit(1)

    FOLDER   = sys.argv[1]
    ISSUE_ID = sys.argv[2]

    def title_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        _style(c, TITLE_FILL, TITLE_FONT, TITLE_ALIGN)
        return c

    def sec_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        _style(c, SEC_FILL, SEC_FONT, SEC_ALIGN)
        return c

    def hdr_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        _style(c, HDR_FILL, HDR_FONT, HDR_ALIGN, BORDER)
        return c

    def data_cell(ws, row, col, val="", stripe=False):
        c = ws.cell(row=row, column=col, value=val)
        fill = STRIPE_FILL if stripe else WHITE_FILL
        _style(c, fill, CELL_FONT, CELL_ALIGN, BORDER)
        return c

    def paste_cell(ws, row, col, val=""):
        c = ws.cell(row=row, column=col, value=val)
        _style(c, PASTE_FILL, CELL_FONT, PASTE_ALIGN, PASTE_BORDER)
        return c

    def note_cell(ws, row, col, val=""):
        """プレースホルダー説明文（背景なし・ボーダーなし）"""
        c = ws.cell(row=row, column=col, value=val)
        c.font = CELL_FONT
        c.alignment = CELL_ALIGN
        return c

    def merge(ws, r1, c1, r2, c2):
        ws.merge_cells(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")

    def set_col_widths(ws, widths):
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    def set_row_heights(ws, heights):
        for r, h in heights.items():
            ws.row_dimensions[r].height = h

    os.makedirs(FOLDER, exist_ok=True)
    wb = openpyxl.Workbook()

    # ==========================================================
    # Sheet1: テスト仕様
    # ==========================================================
    ws1 = wb.active
    ws1.title = "テスト仕様"
    set_col_widths(ws1, {"A": 6, "B": 40, "C": 15, "D": 50, "E": 40, "F": 35, "G": 20})
    set_row_heights(ws1, {1: 38, 2: 28, **{r: 45 for r in range(3, 12)}})

    title_cell(ws1, 1, 1, "テスト仕様"); merge(ws1, 1, 1, 1, 7)
    for i, h in enumerate(["No", "確認観点", "タイミング", "確認手順", "期待結果", "エビデンスの取り方", "貼付先シート"], 1):
        hdr_cell(ws1, 2, i, h)
    for r in range(3, 12):
        stripe = (r % 2 == 0)
        for c in range(1, 8):
            data_cell(ws1, r, c, "", stripe=stripe)

    # ==========================================================
    # Sheet2: 実装前エビデンス
    # ==========================================================
    ws2 = wb.create_sheet("実装前エビデンス")
    set_col_widths(ws2, {"A": 6, "B": 50, "C": 12, "D": 40})
    set_row_heights(ws2, {
        1: 38, 2: 45, 3: 8, 4: 28, 5: 28,
        **{r: 25 for r in range(6, 8)}, 8: 8, 9: 28, 10: 25,
        **{r: 30 for r in range(11, 21)}, 21: 8, 22: 25,
        **{r: 30 for r in range(23, 33)}
    })

    # r1: タイトル (A:D merged)
    title_cell(ws2, 1, 1, "実装前エビデンス"); merge(ws2, 1, 1, 1, 4)
    # r2: 説明 (A:D merged)
    note_cell(ws2, 2, 1,
              "テスト仕様シートの「実装前」「両方」の項目を実施し、エビデンスを貼り付けてください。")
    merge(ws2, 2, 1, 2, 4)

    # r3: blank / r4: チェックリスト見出し, r5: ヘッダー, r6-7: チェック行
    sec_cell(ws2, 4, 1, "■ 確認チェックリスト（実装前）"); merge(ws2, 4, 1, 4, 4)
    for i, h in enumerate(["□", "確認観点", "結果", "メモ（スクリーンショット貼付欄）"], 1):
        hdr_cell(ws2, 5, i, h)
    for r in range(6, 8):
        stripe = (r % 2 == 0)
        data_cell(ws2, r, 1, "□", stripe=stripe)
        for c in range(2, 5):
            data_cell(ws2, r, c, "", stripe=stripe)

    # r8: blank / r9: エビデンス貼付欄見出し
    sec_cell(ws2, 9, 1, "■ エビデンス貼付欄"); merge(ws2, 9, 1, 9, 4)
    # r10: エビデンス① テキスト (A:D merged, note)
    note_cell(ws2, 10, 1, "エビデンス①: （スクリーンショットを貼り付けてください）")
    merge(ws2, 10, 1, 10, 4)
    # r11-20: 貼付枠（A:D merged, 10行、薄グレー背景）
    paste_cell(ws2, 11, 1, "ここにスクリーンショットを貼り付けてください")
    merge(ws2, 11, 1, 20, 4)

    # r21: blank / r22: エビデンス② テキスト (A:D merged)
    note_cell(ws2, 22, 1, "エビデンス②: （スクリーンショットを貼り付けてください）")
    merge(ws2, 22, 1, 22, 4)
    # r23-32: 貼付枠 (A:D merged, 10行)
    paste_cell(ws2, 23, 1, "ここにスクリーンショットを貼り付けてください")
    merge(ws2, 23, 1, 32, 4)

    # ==========================================================
    # Sheet3: 実装後エビデンス
    # ==========================================================
    ws3 = wb.create_sheet("実装後エビデンス")
    set_col_widths(ws3, {"A": 6, "B": 50, "C": 12, "D": 40})
    set_row_heights(ws3, {
        1: 38, 2: 30, 3: 8, 4: 28, 5: 28,
        **{r: 25 for r in range(6, 13)}, 13: 8, 14: 28,
        **{r: 25 for r in range(15, 16)}, **{r: 30 for r in range(16, 26)},
        26: 8, 27: 25, **{r: 30 for r in range(28, 38)},
        38: 8, 39: 25, **{r: 30 for r in range(40, 50)},
        50: 8, 51: 25, **{r: 30 for r in range(52, 62)}
    })

    # r1: タイトル (A:D merged)
    title_cell(ws3, 1, 1, "実装後エビデンス"); merge(ws3, 1, 1, 1, 4)
    # r2: 説明
    note_cell(ws3, 2, 1, "テスト仕様シートの「実装後」「両方」の項目を実施し、エビデンスを貼り付けてください。")
    merge(ws3, 2, 1, 2, 4)

    # r3: blank / r4: チェックリスト見出し, r5: ヘッダー, r6-12: チェック行 (7行)
    sec_cell(ws3, 4, 1, "■ 確認チェックリスト（実装後）"); merge(ws3, 4, 1, 4, 4)
    for i, h in enumerate(["□", "確認観点", "結果", "メモ（スクリーンショット貼付欄）"], 1):
        hdr_cell(ws3, 5, i, h)
    for r in range(6, 13):
        stripe = (r % 2 == 0)
        data_cell(ws3, r, 1, "□", stripe=stripe)
        for c in range(2, 5):
            data_cell(ws3, r, c, "", stripe=stripe)

    # r13: blank / r14: エビデンス貼付欄見出し
    sec_cell(ws3, 14, 1, "■ エビデンス貼付欄"); merge(ws3, 14, 1, 14, 4)
    # r15: エビデンス① テキスト (A:D merged)
    note_cell(ws3, 15, 1, "エビデンス①: （スクリーンショットを貼り付けてください）")
    merge(ws3, 15, 1, 15, 4)
    # r16-25: 貼付枠 (A:D merged, 10行)
    paste_cell(ws3, 16, 1, "ここにスクリーンショットを貼り付けてください")
    merge(ws3, 16, 1, 25, 4)

    # r26: blank / r27: エビデンス② テキスト
    note_cell(ws3, 27, 1, "エビデンス②: （スクリーンショットを貼り付けてください）")
    merge(ws3, 27, 1, 27, 4)
    # r28-37: 貼付枠
    paste_cell(ws3, 28, 1, "ここにスクリーンショットを貼り付けてください")
    merge(ws3, 28, 1, 37, 4)

    # r38: blank / r39: エビデンス③ テキスト
    note_cell(ws3, 39, 1, "エビデンス③: （スクリーンショットを貼り付けてください）")
    merge(ws3, 39, 1, 39, 4)
    # r40-49: 貼付枠
    paste_cell(ws3, 40, 1, "ここにスクリーンショットを貼り付けてください")
    merge(ws3, 40, 1, 49, 4)

    # r50: blank / r51: エビデンス④ テキスト
    note_cell(ws3, 51, 1, "エビデンス④: （スクリーンショットを貼り付けてください）")
    merge(ws3, 51, 1, 51, 4)
    # r52-61: 貼付枠
    paste_cell(ws3, 52, 1, "ここにスクリーンショットを貼り付けてください")
    merge(ws3, 52, 1, 61, 4)

    path = os.path.join(FOLDER, f"{ISSUE_ID}_エビデンス.xlsx")
    wb.save(path)
    print(f"生成完了: {path}")


if __name__ == "__main__":
    main()
