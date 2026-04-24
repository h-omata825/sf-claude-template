# -*- coding: utf-8 -*-
"""
backlog-xlsx / update_records.py
対応記録.xlsx を更新するスクリプト

Usage (タイムライン行を追加):
    python update_records.py --folder FOLDER --issue-id ID timeline \
      --phase "調査" --source "Claude" --content "〇〇を調査: 原因は△△"

Usage (セルを直接更新):
    python update_records.py --folder FOLDER --issue-id ID cell \
      --sheet "対応方針" --row 10 --col 1 --value "採用理由の説明"
"""

import argparse
import datetime
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

WRAP = Alignment(wrap_text=True, vertical="top")
STRIPE_A = PatternFill("solid", fgColor="FFFFFF")  # 奇数行
STRIPE_B = PatternFill("solid", fgColor="F2F7FB")  # 偶数行（薄青）


def find_next_empty_row(ws, col=1, start_row=1):
    """指定列で最初の空行を返す（start_row から下方向に検索）"""
    r = start_row
    while ws.cell(row=r, column=col).value is not None:
        r += 1
    return r


def cmd_timeline(args, wb):
    """サマリー・経緯シートのタイムラインに1行追加する"""
    sheet_name = "サマリー・経緯"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] シート '{sheet_name}' が見つかりません。")
        sys.exit(1)
    ws = wb[sheet_name]

    # タイムラインヘッダー行を探す（"No" がある行）
    timeline_header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "No":
                timeline_header_row = cell.row
                break
        if timeline_header_row:
            break

    if not timeline_header_row:
        print("[ERROR] タイムラインのヘッダー行（'No' セル）が見つかりません。")
        sys.exit(1)

    data_start = timeline_header_row + 1
    next_row = find_next_empty_row(ws, col=1, start_row=data_start)

    # No 列は現在の行数から算出
    no = next_row - data_start + 1
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    fill = STRIPE_A if (no % 2 == 1) else STRIPE_B

    for col, value in enumerate([no, now, args.source, args.phase, args.content, args.reason or ""], start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = WRAP
        cell.fill = fill

    print(f"タイムライン追加: 行{next_row} / {now} / {args.phase} / {args.content[:30]}...")


def cmd_cell(args, wb):
    """指定したシート・行・列のセルを更新する"""
    if args.sheet not in wb.sheetnames:
        print(f"[ERROR] シート '{args.sheet}' が見つかりません。利用可能: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[args.sheet]
    ws.cell(row=args.row, column=args.col, value=args.value).alignment = WRAP
    print(f"セル更新: {args.sheet}!({args.row},{args.col}) = {args.value[:40]}...")


def main():
    parser = argparse.ArgumentParser(description="対応記録.xlsx を更新する")
    parser.add_argument("--folder",   required=True, help="保存先フォルダパス")
    parser.add_argument("--issue-id", required=True, dest="issue_id", help="課題ID (例: GF-327)")

    sub = parser.add_subparsers(dest="command", required=True)

    # タイムライン追加サブコマンド
    p_tl = sub.add_parser("timeline", help="タイムラインに行を追加する")
    p_tl.add_argument("--phase",   required=True, help="フェーズ名 (例: 調査, 実装, テスト)")
    p_tl.add_argument("--source",  default="Claude", help="発生元 (例: Claude, ユーザ)")
    p_tl.add_argument("--content", required=True, help="内容・決定事項")
    p_tl.add_argument("--reason",  default="", help="変更・判断の理由（任意）")

    # セル直接更新サブコマンド
    p_cell = sub.add_parser("cell", help="特定セルを直接更新する")
    p_cell.add_argument("--sheet", required=True, help="シート名")
    p_cell.add_argument("--row",   required=True, type=int, help="行番号")
    p_cell.add_argument("--col",   required=True, type=int, help="列番号")
    p_cell.add_argument("--value", required=True, help="書き込む値")

    args = parser.parse_args()

    xlsx_path = os.path.join(args.folder, f"{args.issue_id}_対応記録.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"[ERROR] ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)

    if args.command == "timeline":
        cmd_timeline(args, wb)
    elif args.command == "cell":
        cmd_cell(args, wb)

    wb.save(xlsx_path)
    print(f"保存完了: {xlsx_path}")


if __name__ == "__main__":
    main()
