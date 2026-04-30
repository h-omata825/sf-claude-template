# -*- coding: utf-8 -*-
"""backlog-xlsx / create_records.py
対応記録.xlsx を生成する（Phase 3 完了直後に MD3点を読んで全シート埋め）

Usage:
    python create_records.py \\
      --folder FOLDER --issue-id ID \\
      --investigation PATH \\
      --approach-plan PATH \\
      --implementation-plan PATH
"""

import argparse
import datetime
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill
except ImportError:
    print("[ERROR] openpyxl がインストールされていません。`pip install openpyxl` を実行してください。")
    sys.exit(1)

TEMPLATE = Path(__file__).parent / "対応記録テンプレート.xlsx"
WRAP = Alignment(wrap_text=True, vertical="top")
STRIPE_A = PatternFill("solid", fgColor="FFFFFF")
STRIPE_B = PatternFill("solid", fgColor="F2F7FB")


# ── MD パースユーティリティ ─────────────────────────────────────────────────

def read_md(path):
    if path and Path(path).exists():
        return Path(path).read_text(encoding="utf-8")
    return ""


def extract_section(md, *headings):
    """指定見出し（## または ###）のセクション本文を返す。
    複数見出しは先にマッチしたものを使用。
    末尾の括弧「（確定後に記入）」のような付記も許容する。  [M2]
    """
    for h in headings:
        pat = r"^#{1,3}\s+" + re.escape(h) + r"(?:\s*[（(][^)）]*[)）])?\s*$"
        m = re.search(pat, md, re.MULTILINE)
        if m:
            start = m.end()
            rest = md[start:]
            end_m = re.search(r"^#{1,3}\s", rest, re.MULTILINE)
            body = rest[: end_m.start()] if end_m else rest
            return body.strip()
    return ""


def extract_section_after_keyword(md, keyword):
    """## 見出しではなく本文中のキーワード行以降のセクションを返す補助関数。  [M8]"""
    idx = md.find(keyword)
    if idx == -1:
        return ""
    rest = md[idx:]
    end_m = re.search(r"^#{1,3}\s", rest[len(keyword):], re.MULTILINE)
    if end_m:
        return rest[len(keyword): len(keyword) + end_m.start()].strip()
    return rest[len(keyword):].strip()


def parse_md_table(section_text):
    """Markdown テーブルを [{col_name: value, ...}] のリストに変換する。"""
    rows = []
    headers = []
    for line in section_text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if all(re.match(r"^[-: ]+$", c) for c in cells):
            continue  # 区切り行
        if not headers:
            headers = cells
        else:
            rows.append(dict(zip(headers, cells)))
    return rows


def parse_checklist(section_text):
    """- [ ] または - [x] の行からチェックリスト文字列のリストを返す。"""
    items = []
    for line in section_text.splitlines():
        m = re.match(r"^\s*-\s+\[[ xX]\]\s+(.+)", line)
        if m:
            items.append(m.group(1).strip())
    return items


def parse_numbered_list(section_text):
    """1. 2. ... の番号付きリストを文字列リストで返す。"""
    items = []
    for line in section_text.splitlines():
        m = re.match(r"^\s*\d+[\.\)]\s+(.+)", line)
        if m:
            items.append(m.group(1).strip())
        elif line.strip() and not re.match(r"^#", line) and items:
            # 継続行
            items[-1] += " " + line.strip()
    return items


def extract_metadata(md, key):
    """key: value 形式の値を取る。
    以下の全形式に対応:  [M1]
      - key: value        (プレーン)
      - **key**: value    (太字)
      - - key: value      (リスト)
      - - **key**: value  (リスト+太字)
    半角・全角コロン両対応。
    """
    pat = rf"^\s*(?:[-*+]\s+)?(?:\*\*)?{re.escape(key)}(?:\*\*)?\s*[:|：]\s*(.+?)\s*$"
    m = re.search(pat, md, re.MULTILINE)
    if m:
        return re.sub(r"\*\*\s*$", "", m.group(1)).strip()
    return ""


def to_median_hours(text):
    """「2〜4h」「2~4 時間」「2-4h」を中央値「3h」へ変換。  [M5]
    範囲表記でなければそのまま返す。
    """
    m = re.match(r"^\s*(\d+(?:\.\d+)?)\s*[〜~\-－]\s*(\d+(?:\.\d+)?)\s*(h|時間|H)\s*$", text or "")
    if not m:
        return text
    lo, hi = float(m.group(1)), float(m.group(2))
    med = (lo + hi) / 2
    return f"{med:g}h"


def get_col(row, *candidates):
    """複数の列名候補から最初にヒットした値を返す。  [M11, M12]"""
    for c in candidates:
        if c in row and row[c]:
            return row[c]
    return ""


def parse_approach_options_h3(section_md):
    """### 案A: 方針名 配下の - **概要**: ... 形式を dict リストへ変換。  [M4]
    既存 parse_md_table が空の場合の fallback として使う。
    """
    options = []
    for m in re.finditer(
        r"^###\s+案([A-ZA-Z])[:：]\s*(.+?)(?:\s*【.+?】)?\s*$",
        section_md, re.MULTILINE
    ):
        no, name = m.group(1), m.group(2).strip()
        body_start = m.end()
        rest = section_md[body_start:]
        next_h = re.search(r"^#{2,3}\s", rest, re.MULTILINE)
        body = rest[: next_h.start()] if next_h else rest

        opt = {"案No": f"案{no}", "方針名": name}
        for key in ["概要", "メリット", "デメリット", "リスク", "前提", "見込み工数"]:
            sub_pat = rf"^\s*-\s+\*\*{re.escape(key)}\*\*\s*[:|：]\s*(.+?)(?=^\s*-\s+\*\*|\Z)"
            sm = re.search(sub_pat, body, re.MULTILINE | re.DOTALL)
            if sm:
                value = re.sub(r"\s+", " ", sm.group(1)).strip()
                opt["工数" if key == "見込み工数" else key] = value
        options.append(opt)
    return options


def find_header_row(ws, candidates):
    """A 列を走査して candidates のいずれかに一致する行番号を返す。見つからなければ None。"""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and any(str(c) in str(cell.value) for c in candidates):
            return cell.row
    return None


def copy_row_style(ws, src_row, dst_row, max_col=8):
    """src_row の書式を dst_row にコピーする（insert_rows 後のスタイル継承用）。"""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst._style = src._style
            dst.alignment = WRAP


def insert_rows_with_format(ws, insert_at, count, source_row, max_col):
    """insert_rows + 行高継承 + マージ補修を一括で行う (openpyxl の既知バグを回避)。

    openpyxl の insert_rows は:
    1. row_dimensions[height] のシフトが誤る場合がある
    2. マージセルをシフトせず元位置に残し、シフト後位置にも追加して重複を作る
    本関数はこれら両方を「挿入前スナップショット → クリア → 挿入 → 完全再構築」で補正する。
    """
    # 挿入前にマージ全件と行高全件をスナップショット
    all_merges = [(m.min_row, m.max_row, m.min_col, m.max_col)
                  for m in list(ws.merged_cells.ranges)]
    row_heights = {r: ws.row_dimensions[r].height
                   for r in ws.row_dimensions
                   if ws.row_dimensions[r].height is not None}
    src_h = row_heights.get(source_row)

    # マージを全クリア (insert_rows による重複マージ作成を防止)
    for mcr in list(ws.merged_cells.ranges):
        ws.merged_cells.ranges.discard(mcr)

    # 挿入実行
    ws.insert_rows(insert_at, amount=count)

    # 行高をスナップショットから完全再構築 (openpyxl のシフト誤りを上書き)
    # まず insert_at 以降のシフト元位置を None にリセット (stale コピーを除去)
    for r in row_heights:
        if r >= insert_at:
            ws.row_dimensions[r].height = None
    for r, h in row_heights.items():
        new_r = r + count if r >= insert_at else r
        ws.row_dimensions[new_r].height = h

    # 挿入行に source_row の行高 + セルスタイルをコピー
    for r in range(insert_at, insert_at + count):
        if src_h:
            ws.row_dimensions[r].height = src_h
        copy_row_style(ws, source_row, r, max_col=max_col)

    # マージをスナップショットから完全再構築 (insert_at 以降は count 行シフト)
    for (min_r, max_r, min_c, max_c) in all_merges:
        if min_r >= insert_at:
            ws.merge_cells(start_row=min_r + count, end_row=max_r + count,
                           start_column=min_c, end_column=max_c)
        elif max_r >= insert_at:
            ws.merge_cells(start_row=min_r, end_row=max_r + count,
                           start_column=min_c, end_column=max_c)
        else:
            ws.merge_cells(start_row=min_r, end_row=max_r,
                           start_column=min_c, end_column=max_c)


# ── セル書き込みユーティリティ ──────────────────────────────────────────────

def wset(ws, row, col, value, stripe=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = WRAP
    if stripe:
        cell.fill = stripe
    return cell


# ── サマリー・経緯シート ────────────────────────────────────────────────────

def fill_summary(ws, args, inv_md, approach_md, impl_md):
    # 課題情報 [M1, M3]
    issue_id   = args.issue_id
    title      = extract_metadata(inv_md, "件名") or extract_metadata(inv_md, "タイトル") or ""
    priority   = extract_metadata(inv_md, "優先度") or ""
    deadline   = extract_metadata(inv_md, "期限") or ""
    issue_type = extract_metadata(inv_md, "種別") or extract_metadata(inv_md, "課題種別") or ""

    # 背景: 見出し別名を追加 [M3]
    summary_bg = extract_section(
        inv_md,
        "概要", "背景", "背景・要件", "課題概要",
        "課題サマリー", "要件理解", "一言要約",
    )
    if not summary_bg:
        paras = [p.strip() for p in inv_md.split("\n\n") if p.strip() and not p.strip().startswith("#")]
        summary_bg = paras[0][:200] if paras else ""

    wset(ws, 3, 2, issue_id)
    wset(ws, 4, 2, title)
    wset(ws, 5, 2, f"優先度: {priority} / 期限: {deadline}")
    wset(ws, 6, 2, issue_type)
    wset(ws, 7, 2, "対応中")
    wset(ws, 8, 2, summary_bg)

    # タイムライン 3 行（Phase 1〜3）
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    inv_result = extract_section(
        inv_md,
        "根本原因", "調査結果", "原因", "調査・まとめ",
        "根本原因 / 要件の本質", "要件の本質", "調査まとめ",
    )
    inv_result_oneliner = inv_result.replace("\n", " ")[:80] if inv_result else "調査完了"
    approach_adopted = extract_section(approach_md, "採用方針", "推奨案", "推奨案と根拠")
    approach_oneliner = approach_adopted.replace("\n", " ")[:80] if approach_adopted else "対応方針確定"
    impl_summary = extract_section(impl_md, "実装方針まとめ", "概要", "方針まとめ")
    if not impl_summary:
        change_files = extract_section(impl_md, "変更ファイル一覧", "変更ファイル")
        impl_summary = change_files[:80] if change_files else "実装方針確定"
    impl_oneliner = impl_summary.replace("\n", " ")[:80]

    tl_rows = [
        (1, now, "Claude", "調査", f"調査完了: {inv_result_oneliner}", ""),
        (2, now, "ユーザ", "方針策定", f"対応方針確定: {approach_oneliner}", ""),
        (3, now, "ユーザ", "実装方針確定", f"全判断ポイント確定: {impl_oneliner}", ""),
    ]
    for i, row in enumerate(tl_rows):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        for j, val in enumerate(row, start=1):
            wset(ws, 13 + i, j, val, fill)


# ── 対応方針シート ──────────────────────────────────────────────────────────

def fill_approach(ws, approach_md):
    # 方針比較テーブル（r4〜、テンプレ標準 3 件枠）[M3, M4, M5, M7]
    table_text = extract_section(
        approach_md,
        "方針比較", "方針比較テーブル", "対応方針比較",
        "対応方針の各案", "案一覧",
    )
    rows = parse_md_table(table_text)
    if not rows and table_text:
        # H3 + 箇条書き形式 (### 案A: ... / - **概要**: ...) を fallback でパース [M4]
        rows = parse_approach_options_h3(table_text)
    if not rows and approach_md:
        # セクション名が section_text に入らず approach_md 全体に対して走らせる
        rows = parse_approach_options_h3(approach_md)

    col_order = ["案No", "方針名", "概要", "メリット", "デメリット", "リスク", "工数"]
    APPROACH_START = 4
    APPROACH_LIMIT = 2  # テンプレ r4-r5
    extra_approach = max(0, len(rows) - APPROACH_LIMIT)
    if extra_approach > 0:
        insert_rows_with_format(
            ws,
            APPROACH_START + APPROACH_LIMIT,
            extra_approach,
            source_row=APPROACH_START + APPROACH_LIMIT - 1,
            max_col=7,
        )

    for i, row in enumerate(rows):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        for j, col in enumerate(col_order, start=1):
            val = row.get(col, "")
            if col == "工数":
                val = to_median_hours(val)  # [M5]
            wset(ws, APPROACH_START + i, j, val, fill)

    # 採用方針（r8 or 行数シフト後の位置） [M2, M3]
    adopted_row = APPROACH_START + len(rows) + 3  # ■採用方針ヘッダの次
    # テンプレ r7 = ■採用方針、r8 以降に内容。insert_rows でシフト済みならその位置
    # ヘッダ行を A 列走査で特定する
    adopted_header_row = find_header_row(ws, ("■ 採用方針",))
    adopted_write_row = (adopted_header_row + 1) if adopted_header_row else 8

    adopted = extract_section(approach_md, "採用方針", "推奨案", "推奨案と根拠")
    if adopted:
        # insert_rows でマージが欠落することがあるため明示的に再付与  [L3]
        has_merge = any(
            m.min_row == adopted_write_row and m.max_row == adopted_write_row
            and m.min_col == 1 and m.max_col == 7
            for m in ws.merged_cells.ranges
        )
        if not has_merge:
            ws.merge_cells(start_row=adopted_write_row, end_row=adopted_write_row,
                           start_column=1, end_column=7)
        wset(ws, adopted_write_row, 1, adopted)

    # 実施前確認事項（テンプレ標準 5 件枠 r22〜r26）[M3, M7]
    checks_text = extract_section(
        approach_md,
        "実施前確認事項", "確認事項", "事前確認",
        "業務要件の確認事項", "前提確認",
    )
    checks = parse_checklist(checks_text)
    if not checks:
        # チェックリスト形式でなければ numbered list を試みる
        checks = parse_numbered_list(checks_text)

    confirm_header_row = find_header_row(ws, ("■ 実施前確認事項",))
    confirm_data_start = (confirm_header_row + 2) if confirm_header_row else 22
    CONFIRM_LIMIT = 4  # テンプレ r22-r25

    extra_confirm = max(0, len(checks) - CONFIRM_LIMIT)
    if extra_confirm > 0:
        insert_rows_with_format(
            ws,
            confirm_data_start + CONFIRM_LIMIT,
            extra_confirm,
            source_row=confirm_data_start + CONFIRM_LIMIT - 1,
            max_col=2,
        )

    for i, item in enumerate(checks):
        wset(ws, confirm_data_start + i, 1, "□")
        wset(ws, confirm_data_start + i, 2, item)

    # 懸念事項（テンプレ r27「■ 懸念事項」の直下から書き込み）[M6]
    concerns_text = extract_section(approach_md, "懸念事項", "リスク・懸念事項", "懸念点")
    concerns = parse_numbered_list(concerns_text)
    if not concerns:
        concerns = [l.strip().lstrip("0123456789.。 ").lstrip("- ") for l in concerns_text.splitlines() if l.strip()]

    concern_header_row = find_header_row(ws, ("■ 懸念事項", "懸念事項"))
    concern_data_start = (concern_header_row + 1) if concern_header_row else 28
    CONCERN_LIMIT = 3  # テンプレ r28-r30 (各行 A:G マージ済)

    extra_concerns = max(0, len(concerns) - CONCERN_LIMIT)
    if extra_concerns > 0:
        insert_rows_with_format(
            ws,
            concern_data_start + CONCERN_LIMIT,
            extra_concerns,
            source_row=concern_data_start + CONCERN_LIMIT - 1,
            max_col=7,
        )
    for i, item in enumerate(concerns):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        target_row = concern_data_start + i
        # テンプレ枠外の行はマージなし → 明示付与  [L4]
        has_merge = any(
            m.min_row == target_row and m.max_row == target_row
            and m.min_col == 1 and m.max_col == 7
            for m in ws.merged_cells.ranges
        )
        if not has_merge:
            ws.merge_cells(start_row=target_row, end_row=target_row,
                           start_column=1, end_column=7)
        wset(ws, target_row, 1, f"{i + 1}. {item}", fill)

    # 方針変更履歴（r28-30 の位置は懸念事項書き込みにより移動している可能性がある）
    # 別名候補で探してヒットしなければスキップ（LINK-139 のように改版履歴しかない MD でも安全）
    log_text = extract_section(approach_md, "方針変更履歴", "変更履歴", "ログ履歴")
    if log_text:
        log_rows = parse_md_table(log_text)
        log_header_row = find_header_row(ws, ("■ 方針変更履歴", "■ ログ履歴"))
        log_data_start = (log_header_row + 1) if log_header_row else None
        if log_data_start:
            col_order_log = ["変更日", "旧方針", "新方針", "理由"]
            for i, row in enumerate(log_rows[:3]):
                fill = STRIPE_A if i % 2 == 0 else STRIPE_B
                if any(col in row for col in col_order_log):
                    line = " / ".join(
                        f"{col}: {row.get(col, '')}" for col in col_order_log if row.get(col, "")
                    )
                else:
                    line = " / ".join(f"{k}: {v}" for k, v in row.items() if v)
                wset(ws, log_data_start + i, 1, line, fill)


# ── 調査・影響範囲シート ────────────────────────────────────────────────────

def fill_investigation(ws, inv_md):
    # 仮説検証（r4-9）[M3, M8]
    hypo_text = extract_section(
        inv_md,
        "仮説検証", "仮説・検証",
        "代替アプローチ", "代替経路", "業務要件の不確実点",
    )
    rows = parse_md_table(hypo_text)
    if not rows:
        # フリーテキストで「ただし以下の代替経路の検討は…」の後のテーブルを探す [M8]
        for kw in ("ただし以下の代替経路", "代替経路の検討", "再検討の余地"):
            alt_text = extract_section_after_keyword(inv_md, kw)
            rows = parse_md_table(alt_text)
            if rows:
                break

    hypo_col_map = [
        ("No", ["No", "#"]),
        ("仮説内容", ["仮説内容", "代替アプローチ", "仮説", "アプローチ"]),
        ("検証方法", ["検証方法", "実現可能性", "方法"]),
        ("検証結果", ["検証結果", "備考", "補足"]),
        ("判定", ["判定"]),
    ]
    for i, row in enumerate(rows[:6]):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        for j, (_, candidates) in enumerate(hypo_col_map, start=1):
            wset(ws, 4 + i, j, get_col(row, *candidates), fill)

    # コード根拠（r12-17）[M3, M9]
    code_text = extract_section(
        inv_md,
        "コード根拠", "コード根拠テーブル",
        "使用中のフィールドAPI名", "参照コード", "フィールドAPI名",
    )
    code_rows = parse_md_table(code_text)
    if not code_rows:
        # ## 使用中のフィールドAPI名 直下の ### サブセクション2本を fallback で結合 [M9]
        sub1 = extract_section(inv_md, "標準 Prospect オブジェクト", "標準Prospectオブジェクト")
        sub2 = extract_section(inv_md, "Pardot 連携カスタム項目", "Pardot連携カスタム項目")
        code_rows = parse_md_table(sub1) + parse_md_table(sub2)

    code_col_map = [
        ("ファイル名", ["ファイル名", "フィールド概念", "ファイル", "コンポーネント"]),
        ("行番号", ["行番号", "確認済み API名", "API名", "行"]),
        ("コード内容", ["コード内容", "確認元", "コード", "参照元"]),
        ("説明", ["説明", "補足", "備考"]),
    ]
    for i, row in enumerate(code_rows[:6]):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        for j, (_, candidates) in enumerate(code_col_map, start=1):
            wset(ws, 12 + i, j, get_col(row, *candidates), fill)

    # 影響範囲（r20-25）[M3, M10]
    impact_text = extract_section(
        inv_md,
        "影響範囲", "影響範囲テーブル",
        "影響する処理・データ", "業務文脈",
    )
    impact_rows = parse_md_table(impact_text)
    if not impact_rows:
        # 業務文脈配下の関連フロー表を fallback
        ctx_text = extract_section(inv_md, "業務文脈", "業務文脈（docs/ から）")
        flow_text = extract_section_after_keyword(ctx_text or inv_md, "関連フロー")
        impact_rows = parse_md_table(flow_text)

    impact_col_map = [
        ("種別", ["種別"]),
        ("対象", ["対象", "フロー名", "コンポーネント名"]),
        ("内容", ["内容", "役割", "影響内容"]),
        ("根拠", ["根拠", "補足", "備考"]),
    ]
    for i, row in enumerate(impact_rows[:6]):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        for j, (_, candidates) in enumerate(impact_col_map, start=1):
            wset(ws, 20 + i, j, get_col(row, *candidates), fill)

    # 関連コンポーネント（r29-33）[M3, M11]
    comp_text = extract_section(inv_md, "関連コンポーネント", "関連コンポーネント一覧")
    comp_rows = parse_md_table(comp_text)
    for i, row in enumerate(comp_rows[:5]):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        # 列名 alias: 「ファイルパス」→「名前」相当として扱う [M11]
        name = get_col(row, "名前", "ファイルパス", "コンポーネント名")
        role = row.get("役割", "")
        finding = get_col(row, "調査結果", "補足", "備考")
        wset(ws, 29 + i, 1, row.get("種別", ""), fill)
        wset(ws, 29 + i, 2, name, fill)
        wset(ws, 29 + i, 3, role, fill)
        wset(ws, 29 + i, 4, finding, fill)


# ── 対応内容シート ──────────────────────────────────────────────────────────

def fill_content(ws, impl_md):
    # 変更ファイル一覧（r9-11、テンプレ標準 3 件枠）
    rows = parse_md_table(extract_section(impl_md, "変更ファイル一覧", "変更ファイル"))
    CHANGE_FILES_START = 9
    CHANGE_FILES_LIMIT = 3  # テンプレ r9-r11
    extra_chg = max(0, len(rows) - CHANGE_FILES_LIMIT)
    if extra_chg > 0:
        insert_rows_with_format(
            ws,
            CHANGE_FILES_START + CHANGE_FILES_LIMIT,
            extra_chg,
            source_row=CHANGE_FILES_START + CHANGE_FILES_LIMIT - 1,
            max_col=5,
        )
    for i, row in enumerate(rows):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        for j, col in enumerate(["No", "ファイルパス", "変更種別", "変更概要"], start=1):
            wset(ws, CHANGE_FILES_START + i, j, row.get(col, ""), fill)

    # Before/After セクションは動的に位置を特定
    ba_header_row = find_header_row(ws, ("■ Before / After",)) or (CHANGE_FILES_START + CHANGE_FILES_LIMIT + 1 + extra_chg)
    wset(ws, ba_header_row + 1, 1, "実装完了後、各ファイルの変更前後を記載する")

    # 影響確認チェックリスト（r21〜）[M3]
    checks = parse_checklist(extract_section(
        impl_md,
        "影響確認チェックリスト", "影響確認",
    ))
    impact_header_row = find_header_row(ws, ("■ 影響確認チェックリスト",))
    IMPACT_CHECK_START = (impact_header_row + 2) if impact_header_row else 21
    IMPACT_CHECK_LIMIT = 6  # テンプレ r21-r26
    extra_impact = max(0, len(checks) - IMPACT_CHECK_LIMIT)
    if extra_impact > 0:
        insert_rows_with_format(
            ws,
            IMPACT_CHECK_START + IMPACT_CHECK_LIMIT,
            extra_impact,
            source_row=IMPACT_CHECK_START + IMPACT_CHECK_LIMIT - 1,
            max_col=2,
        )
    for i, item in enumerate(checks):
        wset(ws, IMPACT_CHECK_START + i, 1, "□")
        wset(ws, IMPACT_CHECK_START + i, 2, item)


# ── テスト・検証記録シート ──────────────────────────────────────────────────

def fill_test(ws, impl_md):
    # テスト方針（r3-4）[M3]
    policy = extract_section(
        impl_md,
        "テスト方針", "テスト概要",
        "テスト方針・概要", "テストシナリオ",
    )
    if not policy:
        policy = "実装前後での動作確認を行う。実装前は現状把握、実装後は修正確認。"
    wset(ws, 3, 1, policy)

    # テストテーブル（r7〜、テンプレ標準 8 件枠）[M3]
    rows = parse_md_table(extract_section(
        impl_md,
        "テスト仕様", "テストケース", "テスト仕様テーブル",
        "テストシナリオ",
    ))
    TEST_START = 7
    TEST_LIMIT = 8  # テンプレ r7-r14
    extra_test = max(0, len(rows) - TEST_LIMIT)
    if extra_test > 0:
        insert_rows_with_format(
            ws,
            TEST_START + TEST_LIMIT,
            extra_test,
            source_row=TEST_START + TEST_LIMIT - 1,
            max_col=8,
        )
    for i, row in enumerate(rows):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        vals = [
            row.get("No", str(i + 1)),
            row.get("タイミング", row.get("区分", "")),
            row.get("確認観点", row.get("テスト項目", "")),
            row.get("確認手順", row.get("確認方法", "")),
            row.get("期待結果", ""),
            "",  # 実際の結果（実装後記入）
            "",  # 判定（実装後記入）
            row.get("エビデンス取得方法", row.get("根拠", "")),
        ]
        for j, val in enumerate(vals, start=1):
            wset(ws, TEST_START + i, j, val, fill)


# ── リリース・ロールバックシート ────────────────────────────────────────────

def fill_release(ws, impl_md, approach_md=""):
    # リリース対象（r4-6、テンプレ標準 3 件枠）[M3, M7, M12]
    rows = parse_md_table(extract_section(
        impl_md,
        "リリース対象", "リリース対象一覧",
        "デプロイ対象", "変更対象一覧",
    ))
    RELEASE_START = 4
    RELEASE_LIMIT = 2  # テンプレ r4-r5
    extra_release = max(0, len(rows) - RELEASE_LIMIT)
    if extra_release > 0:
        insert_rows_with_format(
            ws,
            RELEASE_START + RELEASE_LIMIT,
            extra_release,
            source_row=RELEASE_START + RELEASE_LIMIT - 1,
            max_col=6,
        )

    for i, row in enumerate(rows):
        fill = STRIPE_A if i % 2 == 0 else STRIPE_B
        # 「対象」列を「API名」の alias として扱う [M12]
        api_name = get_col(row, "API名", "対象", "ファイルパス")
        change_type = get_col(row, "変更種別", "種別変更")
        wset(ws, RELEASE_START + i, 1, row.get("No", str(i + 1)), fill)
        wset(ws, RELEASE_START + i, 2, row.get("種別", ""), fill)
        wset(ws, RELEASE_START + i, 3, api_name, fill)
        wset(ws, RELEASE_START + i, 4, change_type, fill)
        wset(ws, RELEASE_START + i, 5, row.get("デプロイ方法", ""), fill)
        wset(ws, RELEASE_START + i, 6, row.get("備考", ""), fill)

    # リリース前確認事項（r9-13）[M3, M13]
    pre_text = extract_section(
        impl_md,
        "リリース前確認事項", "リリース前確認", "デプロイ前確認",
        "事前準備", "実装前準備", "確認事項",
    )
    checks = parse_checklist(pre_text)
    if not checks:
        checks = parse_numbered_list(pre_text)
    # approach_md fallback: 実施前確認事項を転記 [M13]
    if not checks and approach_md:
        pre_fallback = extract_section(
            approach_md,
            "実施前確認事項", "確認事項", "事前確認",
            "業務要件の確認事項",
        )
        checks = parse_checklist(pre_fallback)
        if not checks:
            checks = parse_numbered_list(pre_fallback)

    pre_header_row = find_header_row(ws, ("■ リリース前確認事項",))
    pre_data_start = (pre_header_row + 2) if pre_header_row else 9
    PRE_CHECK_LIMIT = 4  # テンプレ r9-r12
    extra_pre = max(0, len(checks) - PRE_CHECK_LIMIT)
    if extra_pre > 0:
        insert_rows_with_format(
            ws,
            pre_data_start + PRE_CHECK_LIMIT,
            extra_pre,
            source_row=pre_data_start + PRE_CHECK_LIMIT - 1,
            max_col=2,
        )
    for i, item in enumerate(checks):
        wset(ws, pre_data_start + i, 1, "□")
        wset(ws, pre_data_start + i, 2, item)

    # デプロイ手順（r15-18、テンプレ標準 4 件枠、各行 A:F マージ済）
    steps = parse_numbered_list(extract_section(impl_md, "デプロイ手順", "リリース手順"))
    deploy_header_row = find_header_row(ws, ("■ デプロイ手順",))
    deploy_data_start = (deploy_header_row + 1) if deploy_header_row else 15
    DEPLOY_LIMIT = 4  # テンプレ r15-r18
    extra_deploy = max(0, len(steps) - DEPLOY_LIMIT)
    if extra_deploy > 0:
        insert_rows_with_format(
            ws,
            deploy_data_start + DEPLOY_LIMIT,
            extra_deploy,
            source_row=deploy_data_start + DEPLOY_LIMIT - 1,
            max_col=6,
        )
    for i, step in enumerate(steps):
        target_row = deploy_data_start + i
        has_merge = any(
            m.min_row == target_row and m.max_row == target_row
            and m.min_col == 1 and m.max_col == 6
            for m in ws.merged_cells.ranges
        )
        if not has_merge:
            ws.merge_cells(start_row=target_row, end_row=target_row,
                           start_column=1, end_column=6)
        wset(ws, target_row, 1, f"{i + 1}. {step}")

    # デプロイ後確認事項（r22-26）[M3]
    post_text = extract_section(
        impl_md,
        "デプロイ後確認事項", "リリース後確認", "デプロイ後確認",
        "実装後確認", "モニタリング", "影響確認チェックリスト",
    )
    post_checks = parse_checklist(post_text)
    if not post_checks:
        post_checks = parse_numbered_list(post_text)

    post_header_row = find_header_row(ws, ("■ デプロイ後確認事項",))
    post_data_start = (post_header_row + 2) if post_header_row else 22
    POST_CHECK_LIMIT = 4  # テンプレ r22-r25
    extra_post = max(0, len(post_checks) - POST_CHECK_LIMIT)
    if extra_post > 0:
        insert_rows_with_format(
            ws,
            post_data_start + POST_CHECK_LIMIT,
            extra_post,
            source_row=post_data_start + POST_CHECK_LIMIT - 1,
            max_col=2,
        )
    for i, item in enumerate(post_checks):
        wset(ws, post_data_start + i, 1, "□")
        wset(ws, post_data_start + i, 2, item)

    # 注意事項（r28-30）[M3, M13]
    notes_text = extract_section(
        impl_md,
        "注意事項", "リスク・注意事項", "注意点",
        "懸念事項", "リスク",
    )
    # approach_md fallback: impl_md に注意事項がなければ approach の懸念事項を転記 [M13]
    if not notes_text and approach_md:
        notes_text = extract_section(approach_md, "懸念事項", "リスク・懸念事項", "注意事項")

    notes = parse_numbered_list(notes_text)
    if not notes:
        notes = [l.strip().lstrip("- ") for l in notes_text.splitlines() if l.strip().startswith("-")]

    notes_header_row = find_header_row(ws, ("■ 注意事項・リスク", "■ 注意事項"))
    notes_data_start = (notes_header_row + 1) if notes_header_row else 28
    NOTES_LIMIT = 2  # テンプレ r28-r29 (各行 A:F マージ済)
    extra_notes = max(0, len(notes) - NOTES_LIMIT)
    if extra_notes > 0:
        insert_rows_with_format(
            ws,
            notes_data_start + NOTES_LIMIT,
            extra_notes,
            source_row=notes_data_start + NOTES_LIMIT - 1,
            max_col=6,
        )
    for i, item in enumerate(notes):
        target_row = notes_data_start + i
        has_merge = any(
            m.min_row == target_row and m.max_row == target_row
            and m.min_col == 1 and m.max_col == 6
            for m in ws.merged_cells.ranges
        )
        if not has_merge:
            ws.merge_cells(start_row=target_row, end_row=target_row,
                           start_column=1, end_column=6)
        wset(ws, target_row, 1, item)

    # ロールバック手順（r32-r35、テンプレ標準 4 件枠、各行 A:F マージ済）
    rb_steps = parse_numbered_list(extract_section(impl_md, "ロールバック手順"))
    rb_header_row = find_header_row(ws, ("■ ロールバック手順",))
    rb_data_start = (rb_header_row + 1) if rb_header_row else 32
    RB_LIMIT = 4  # テンプレ r32-r35
    extra_rb = max(0, len(rb_steps) - RB_LIMIT)
    if extra_rb > 0:
        insert_rows_with_format(
            ws,
            rb_data_start + RB_LIMIT,
            extra_rb,
            source_row=rb_data_start + RB_LIMIT - 1,
            max_col=6,
        )
    for i, step in enumerate(rb_steps):
        target_row = rb_data_start + i
        has_merge = any(
            m.min_row == target_row and m.max_row == target_row
            and m.min_col == 1 and m.max_col == 6
            for m in ws.merged_cells.ranges
        )
        if not has_merge:
            ws.merge_cells(start_row=target_row, end_row=target_row,
                           start_column=1, end_column=6)
        wset(ws, target_row, 1, f"{i + 1}. {step}")


# ── main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="対応記録.xlsx を生成する（Phase 3 完了後の全シート埋め版）")
    parser.add_argument("--folder",              required=True)
    parser.add_argument("--issue-id",            required=True, dest="issue_id")
    parser.add_argument("--investigation",       required=True, dest="investigation",
                        help="docs/logs/{issueID}/investigation.md のパス")
    parser.add_argument("--approach-plan",       required=True, dest="approach_plan",
                        help="docs/logs/{issueID}/approach-plan.md のパス")
    parser.add_argument("--implementation-plan", required=True, dest="implementation_plan",
                        help="docs/logs/{issueID}/implementation-plan.md のパス")
    args = parser.parse_args()

    if not TEMPLATE.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    inv_md  = read_md(args.investigation)
    app_md  = read_md(args.approach_plan)
    impl_md = read_md(args.implementation_plan)

    missing = []
    if not inv_md:
        missing.append(args.investigation)
    if not app_md:
        missing.append(args.approach_plan)
    if not impl_md:
        missing.append(args.implementation_plan)
    if missing:
        print(f"[ERROR] 以下の MD ファイルが見つかりません:\n" + "\n".join(f"  {p}" for p in missing))
        sys.exit(1)

    os.makedirs(args.folder, exist_ok=True)
    wb = load_workbook(TEMPLATE)

    fill_summary(wb["サマリー・経緯"], args, inv_md, app_md, impl_md)
    fill_approach(wb["対応方針"], app_md)
    fill_investigation(wb["調査・影響範囲"], inv_md)
    fill_content(wb["対応内容"], impl_md)
    fill_test(wb["テスト・検証記録"], impl_md)
    fill_release(wb["リリース・ロールバック"], impl_md, app_md)  # [M13] approach_md を渡す

    path = os.path.join(args.folder, f"{args.issue_id}_対応記録.xlsx")
    wb.save(path)
    print(f"生成完了: {path}")


if __name__ == "__main__":
    main()
