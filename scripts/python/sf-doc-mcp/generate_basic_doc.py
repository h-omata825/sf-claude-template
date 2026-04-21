"""
プロジェクト概要書.xlsx を生成する（直接生成方式 + 高品質図埋め込み）。

入力 (docs/ 配下):
  docs/overview/org-profile.md         — 組織・プロジェクト基本情報・用語集
  docs/requirements/requirements.md    — 導入背景・目的
  docs/architecture/system.json        — システム構成図データ
  docs/flow/swimlanes.json             — 業務フロー（As-Is/To-Be）
  docs/catalog/_index.md               — オブジェクト一覧
  docs/catalog/_data-model.md          — オブジェクト関連定義（ER図）

出力:
  プロジェクト概要書.xlsx（5シート: 表紙/システム概要/業務フロー図/ER図/用語集）
  ※ 図は diagram_gen.py (graphviz/drawsvg) で高品質PNG生成して埋め込み

Usage:
  python generate_basic_doc.py \\
    --docs-dir <path/to/project/docs> \\
    --output <output/プロジェクト概要書.xlsx> \\
    --author "作成者名" \\
    [--project-name "プロジェクト名"]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import diagram_gen as dg

# ── デザインシステム（build_basic_doc_template.py と統一）────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"
FONT_NAME    = "游ゴシック"
THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")
GRID_LEFT  = 2
GRID_RIGHT = 31


def _fill(c):   return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_med(): return Border(left=MED,  right=MED,  top=MED,  bottom=MED)

def _set_h(ws, row, h): ws.row_dimensions[row].height = h
def _setup_grid(ws):
    ws.column_dimensions["A"].width = 2.0
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        ws.column_dimensions[get_column_letter(i)].width = 4.2
    ws.sheet_view.showGridLines = False

def _MW(ws, row, cs, ce, value="", bold=False, fg=C_FONT_D, bg=None,
        h="left", v="center", wrap=True, border=None, size=10):
    if border:
        for c in range(cs, ce + 1): ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1): ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws.cell(row=row, column=cs, value=value)
    cell.font = _fnt(bold=bold, color=fg, size=size)
    cell.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:     cell.fill = _fill(bg)
    if border: cell.border = border
    return cell

def _title_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_W, bg=C_TITLE_DARK,
        h="center", size=14, border=B_med())
    _set_h(ws, row, 28)
    return row + 1

def _section_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, border=B_all())
    _set_h(ws, row, 18)
    return row + 1

def _sub_section_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_D, bg=C_LABEL_BG, border=B_all())
    _set_h(ws, row, 16)
    return row + 1

def _meta_row(ws, row, label, value="", col_label_end=8):
    _MW(ws, row, GRID_LEFT, col_label_end, label,
        bold=True, bg=C_LABEL_BG, border=B_all())
    _MW(ws, row, col_label_end + 1, GRID_RIGHT, value, border=B_all())
    _set_h(ws, row, 16)
    return row + 1

def _hdr_row(ws, row, cols: list[tuple[int, int, str]]):
    for cs, ce, label in cols:
        _MW(ws, row, cs, ce, label,
            bold=True, fg=C_FONT_W, bg=C_HDR_BLUE, h="center", border=B_all())
    _set_h(ws, row, 18)
    return row + 1

def _data_row(ws, row, cols_vals: list[tuple[int, int, str]], row_h=16):
    for cs, ce, val in cols_vals:
        _MW(ws, row, cs, ce, val or "", border=B_all())
    _set_h(ws, row, row_h)
    return row + 1

def _empty_rows(ws, row, count, col_groups: list[tuple[int, int]], row_h=16):
    for r in range(row, row + count):
        for cs, ce in col_groups:
            for c in range(cs, ce + 1):
                ws.cell(row=r, column=c).border = B_all()
            ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
        _set_h(ws, r, row_h)
    return row + count

def _text_area(ws, row, n_rows, value="", row_h=18):
    for r in range(row, row + n_rows):
        for c in range(GRID_LEFT, GRID_RIGHT + 1):
            ws.cell(row=r, column=c).border = B_all()
        _set_h(ws, r, row_h)
    ws.merge_cells(start_row=row, start_column=GRID_LEFT,
                   end_row=row + n_rows - 1, end_column=GRID_RIGHT)
    cell = ws.cell(row=row, column=GRID_LEFT, value=value)
    cell.font = _fnt()
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = B_all()
    return row + n_rows

def _margin(ws, row, h=6):
    _set_h(ws, row, h)
    return row + 1


# ── docs パーサー ─────────────────────────────────────────────────

def _table_val(text: str, key: str) -> str:
    m = re.search(rf'\|\s*{re.escape(key)}\s*\|\s*(.+?)\s*\|', text)
    return m.group(1).strip() if m else ""

def _section_text(text: str, heading: str) -> str:
    # ## / ### / #### 見出しにキーワードが含まれていれば一致（番号付き見出し対応）
    # 注: f-string では {n,m} が tuple に展開されるため string 連結で構築する
    pat = r'#{2,4}[^\n]*' + re.escape(heading) + r'[^\n]*\n(.*?)(?=\n#{2,4}[^\n]|\Z)'
    m = re.search(pat, text, re.DOTALL)
    return m.group(1).strip() if m else ""


def parse_org(path: Path) -> dict:
    if not path.exists(): return {}
    t = path.read_text(encoding="utf-8")
    def tv(k): return _table_val(t, k)

    # 用語集
    glossary = []
    sec = _section_text(t, "用語集") or _section_text(t, "Glossary")
    for line in sec.splitlines():
        if not line.strip().startswith("|"): continue
        cols = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cols) >= 2 and cols[0] and cols[0] not in ("業務用語", "---", "用語"):
            glossary.append({"biz": cols[0], "sf": cols[1] if len(cols) > 1 else "",
                             "desc": cols[2] if len(cols) > 2 else ""})

    # 体制（担当ベンダー表も対象）
    stakeholders = []
    for sec_name in ("担当ベンダー", "ステークホルダー", "体制", "関係者"):
        sec = _section_text(t, sec_name)
        for line in sec.splitlines():
            if not line.strip().startswith("|"): continue
            cols = [c.strip() for c in line.strip().strip("|").split("|")]
            if len(cols) >= 2 and cols[0] and cols[0] not in ("役割", "---", "会社/担当者"):
                stakeholders.append({"role": cols[0], "name": cols[1] if len(cols)>1 else "",
                                     "area": cols[2] if len(cols)>2 else "",
                                     "note": cols[3] if len(cols)>3 else ""})
        if stakeholders: break

    # 背景・目的（AS-IS課題 + TO-BE目的 を結合）
    bg_asis = _section_text(t, "導入背景") or _section_text(t, "AS-IS課題")
    bg_tobe = _section_text(t, "導入目的") or _section_text(t, "TO-BE")
    background = "\n\n".join(s for s in [bg_asis, bg_tobe] if s)

    system_name = tv("システム名") or tv("会社名")
    project_name = tv("プロジェクト名") or (f"{system_name} Salesforce導入プロジェクト" if system_name else "")

    return {
        "project_name": project_name,
        "system_name":  system_name,
        "sf_edition":   tv("Salesforce Edition") or tv("Edition"),
        "start_date":   tv("開始日") or tv("プロジェクト開始日") or tv("Salesforce利用開始"),
        "end_date":     tv("終了予定日") or tv("リリース予定日"),
        "go_live_date": tv("本番公開日"),
        "target_biz":   tv("対象業務"),
        "stakeholders": stakeholders[:6],
        "glossary":     glossary[:30],
        "background":   background,
    }


def parse_requirements(path: Path) -> dict:
    if not path.exists(): return {}
    t = path.read_text(encoding="utf-8")

    # 背景（複数のセクション名を試みる）
    bg = (_section_text(t, "背景・目的") or _section_text(t, "目的") or
          _section_text(t, "導入背景") or "")

    # スコープ（対象）: 1stステップ表から項目名を抽出
    scope_in_sec = _section_text(t, "1stステップ") or _section_text(t, "対象")
    scope_in_items = []
    for line in scope_in_sec.splitlines():
        if not line.strip().startswith("|"): continue
        cols = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cols) >= 2 and cols[0] and not re.fullmatch(r'[-:\s]+', cols[0]) and \
                cols[0] not in ("スコープ項目", "スコープ", "---"):
            scope_in_items.append(cols[0])
    scope_in = "\n".join(f"・{i}" for i in scope_in_items) if scope_in_items else scope_in_sec[:300]

    # スコープ（対象外）: 2ndステップ箇条書きを抽出
    scope_out_sec = _section_text(t, "2ndステップ") or _section_text(t, "対象外")
    scope_out_items = []
    for line in scope_out_sec.splitlines():
        stripped = line.strip()
        if re.match(r'^[-*・]\s+', stripped):
            scope_out_items.append(re.sub(r'^[-*・]\s+', '', stripped))
    scope_out = "\n".join(f"・{i}" for i in scope_out_items) if scope_out_items else scope_out_sec[:300]

    return {"background": bg[:600], "scope_in": scope_in[:400], "scope_out": scope_out[:300]}


def parse_system_json(path: Path) -> dict:
    if not path.exists(): return {}
    try: return json.loads(path.read_text(encoding="utf-8"))
    except Exception: return {}


def parse_swimlanes(path: Path) -> dict:
    if not path.exists(): return {}
    try: return json.loads(path.read_text(encoding="utf-8"))
    except Exception: return {}


def parse_catalog_index(path: Path) -> list[dict]:
    if not path.exists(): return []
    t = path.read_text(encoding="utf-8")
    objs = []
    for line in t.splitlines():
        if not line.strip().startswith("|"): continue
        cols = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cols) >= 2 and cols[0] and cols[0] not in ("API名", "---"):
            objs.append({"api": cols[0], "label": cols[1] if len(cols)>1 else "",
                         "type": cols[2] if len(cols)>2 else ""})
    return objs[:30]


def parse_data_model(path: Path) -> list[dict]:
    """
    _data-model.md から関連定義をパースする。
    Mermaid ERD 形式（||--o{ など）を優先し、なければ親/子テーブル形式を試みる。
    """
    if not path.exists(): return []
    t = path.read_text(encoding="utf-8")
    rels = []

    # ── Mermaid ERD 形式: "  ObjectA ||--o{ ObjectB : "label"" ──
    mermaid_pat = re.compile(
        r'^\s*([\w]+)\s+\|[|o]--[o|>]?\{?\s+([\w]+)\s*:\s*"([^"]*)"',
        re.MULTILINE,
    )
    for m in mermaid_pat.finditer(t):
        parent, child, label = m.group(1), m.group(2), m.group(3)
        rel_type = "master-detail" if ("MD" in label or "主従" in label) else "1-N"
        rels.append({"parent": parent, "rel": rel_type, "child": child, "field": label})
    if rels:
        return rels[:30]

    # ── フォールバック: 親オブジェクト/子オブジェクト列のテーブル ──
    in_rel_table = False
    col_parent = col_child = col_rel = col_field = -1
    for line in t.splitlines():
        stripped = line.strip()
        if not stripped.startswith("|"):
            in_rel_table = False
            continue
        cols = [c.strip() for c in stripped.strip("|").split("|")]
        if "親オブジェクト" in cols and "子オブジェクト" in cols:
            col_parent = cols.index("親オブジェクト")
            col_child  = cols.index("子オブジェクト")
            col_rel    = next((i for i, c in enumerate(cols)
                               if c in ("関係", "リレーション", "種別", "rel")), -1)
            col_field  = next((i for i, c in enumerate(cols)
                               if "項目" in c or "field" in c.lower()), -1)
            in_rel_table = True
            continue
        if not in_rel_table:
            continue
        if all(re.fullmatch(r'[-: ]+', c) for c in cols if c):
            continue
        parent = cols[col_parent] if col_parent < len(cols) else ""
        child  = cols[col_child]  if col_child  < len(cols) else ""
        if parent and child and not parent.isdigit() and not child.isdigit():
            rels.append({"parent": parent, "rel": "", "child": child, "field": ""})

    return rels[:20]


def _pick_flows(swimlanes: dict) -> tuple[dict | None, dict | None]:
    """swimlanes.json から As-Is / To-Be フローを抽出"""
    flows = swimlanes.get("flows", [])
    asis = tobe = overall = None
    for f in flows:
        title = (f.get("title") or "").lower()
        if "as-is" in title or "現状" in title or "asis" in title:
            asis = f
        elif "to-be" in title or "導入後" in title or "tobe" in title:
            tobe = f
        elif f.get("flow_type") == "overall" and overall is None:
            overall = f
    # As-Is が未設定なら overall を当てる
    if asis is None and overall is not None:
        asis = overall
    return asis, tobe


# ── シート 1: 表紙 ─────────────────────────────────────────────
def _build_cover(ws, org: dict, req: dict, author: str):
    _setup_grid(ws)
    r = 2  # row 1 は余白
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "プロジェクト概要書")
    r = _margin(ws, r)
    r = _section_row(ws, r, "プロジェクト基本情報")
    for label, src, key in [
        ("プロジェクト名",     org, "project_name"),
        ("システム名",        org, "system_name"),
        ("導入目的・背景",     req, "background"),
        ("スコープ（対象）",   req, "scope_in"),
        ("スコープ（対象外）", req, "scope_out"),
        ("開始日",          org, "start_date"),
        ("終了予定日",       org, "end_date"),
        ("本番公開日",       org, "go_live_date"),
    ]:
        r = _meta_row(ws, r, label, src.get(key, ""))

    r = _margin(ws, r)
    r = _section_row(ws, r, "体制")
    r = _hdr_row(ws, r, [(2,6,"役割"),(7,16,"氏名 / 組織"),(17,22,"担当領域"),(23,31,"備考")])
    for s in org.get("stakeholders", []):
        r = _data_row(ws, r, [(2,6,s["role"]),(7,16,s["name"]),(17,22,s.get("area","")),
                               (23,31,s.get("note",""))])
    if len(org.get("stakeholders", [])) < 6:
        r = _empty_rows(ws, r, 6 - len(org.get("stakeholders", [])), [(2,6),(7,16),(17,22),(23,31)])

    r = _margin(ws, r)
    r = _section_row(ws, r, "改版履歴")
    r = _hdr_row(ws, r, [(2,3,"版"),(4,7,"改版日"),(8,12,"改版者"),(13,31,"改版内容")])
    today = date.today().strftime("%Y-%m-%d")
    r = _data_row(ws, r, [(2,3,"1.0"),(4,7,today),(8,12,author),(13,31,"初版作成")])
    r = _empty_rows(ws, r, 9, [(2,3),(4,7),(8,12),(13,31)])


# ── シート 2: システム概要 ────────────────────────────────────────
def _build_system_overview(ws, req: dict, system: dict, sys_img_path: str | None):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "システム概要")
    r = _margin(ws, r)

    # 導入背景・課題
    r = _section_row(ws, r, "導入背景・解決する課題")
    bg_text = req.get("background", "")
    n_bg_rows = max(5, len(bg_text) // 60 + 2)
    r = _text_area(ws, r, n_bg_rows, bg_text)
    r = _margin(ws, r)

    # 外部連携先一覧（表を先に、図は後）
    r = _section_row(ws, r, "外部連携先一覧")
    r = _hdr_row(ws, r, [(2,8,"連携先システム"),(9,12,"方向"),(13,18,"方式"),
                          (19,22,"頻度"),(23,31,"目的・概要")])
    for ex in system.get("external_systems", [])[:10]:
        r = _data_row(ws, r, [
            (2,8,ex.get("name","")), (9,12,ex.get("direction","")),
            (13,18,ex.get("protocol","")), (19,22,ex.get("frequency","")),
            (23,31,ex.get("purpose","")),
        ])
    if len(system.get("external_systems", [])) < 5:
        r = _empty_rows(ws, r, 5 - len(system.get("external_systems", [])),
                        [(2,8),(9,12),(13,18),(19,22),(23,31)])
    r = _margin(ws, r)

    # システム全体構成図
    r = _section_row(ws, r, "システム全体構成")
    if sys_img_path:
        n_img_rows = dg.embed_image_in_sheet(ws, sys_img_path, anchor_row=r,
                                             anchor_col=GRID_LEFT, max_width_px=1800)
        r += n_img_rows
    else:
        r = _text_area(ws, r, 15, "（system.json が見つかりません）")


# ── シート 3: 業務フロー図 ────────────────────────────────────────
def _build_flow_sheet(ws, asis_flow: dict | None, tobe_flow: dict | None,
                      asis_img: str | None, tobe_img: str | None):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "業務フロー図")
    r = _margin(ws, r)

    for label, flow, img_path in [
        ("As-Is 業務フロー（現状）",             asis_flow, asis_img),
        ("To-Be 業務フロー（Salesforce導入後）", tobe_flow, tobe_img),
    ]:
        r = _section_row(ws, r, label)

        # 手順テーブルを先に（表が図に埋もれないよう）
        sub = label.split("（")[0]
        r = _sub_section_row(ws, r, f"{sub} 手順")
        r = _hdr_row(ws, r, [(2,3,"No"),(4,8,"担当"),(9,22,"操作・処理内容"),(23,31,"分岐条件")])
        steps = (flow or {}).get("steps", [])[:20]
        # ステップIDごとの遷移条件を収集（分岐が発生する場合に備考欄へ表示）
        step_conds: dict[str, list[str]] = {}
        for tr in (flow or {}).get("transitions", []):
            src = str(tr.get("from", ""))
            cond = tr.get("condition", "")
            if cond and src:
                step_conds.setdefault(src, []).append(cond)
        for i, s in enumerate(steps):
            sid = str(s.get("id",""))
            cond_text = " / ".join(step_conds.get(sid, []))
            r = _data_row(ws, r, [
                (2,3,sid),
                (4,8,str(s.get("lane",""))),
                (9,22,str(s.get("label","") or s.get("title",""))),
                (23,31,cond_text),
            ], row_h=18)
        if len(steps) < 5:
            r = _empty_rows(ws, r, 5 - len(steps), [(2,3),(4,8),(9,22),(23,31)], row_h=18)
        r = _margin(ws, r)

        # フロー図（表の後）
        r = _sub_section_row(ws, r, f"{sub} フロー図")
        if img_path:
            n_img_rows = dg.embed_image_in_sheet(ws, img_path, anchor_row=r,
                                                 anchor_col=GRID_LEFT, max_width_px=1100)
            r += n_img_rows
        else:
            r = _text_area(ws, r, 12, "（フローデータなし）")
        r = _margin(ws, r, 12)


# ── シート 4: ER図 ──────────────────────────────────────────────
def _build_er_sheet(ws, objects: list, relations: list, er_img: str | None):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "ER図（オブジェクト関連図）")
    r = _margin(ws, r)

    # 関連定義表を先に（表が図に埋もれないよう）
    r = _section_row(ws, r, "関連定義表")
    r = _hdr_row(ws, r, [(2,8,"親オブジェクト"),(9,11,"種別"),
                          (12,18,"子オブジェクト"),(19,31,"関係・用途")])
    for rel in relations[:25]:
        r = _data_row(ws, r, [
            (2,8,rel["parent"]), (9,11,rel["rel"]),
            (12,18,rel["child"]), (19,31,rel.get("field","")),
        ])
    if len(relations) < 5:
        r = _empty_rows(ws, r, 5 - len(relations),
                        [(2,8),(9,11),(12,18),(19,31)])
    r = _margin(ws, r)

    # ER図（表の後、max_width 拡大で可読性向上）
    r = _section_row(ws, r, "オブジェクト関連図")
    if er_img:
        n_img_rows = dg.embed_image_in_sheet(ws, er_img, anchor_row=r,
                                             anchor_col=GRID_LEFT, max_width_px=1800)
        r += n_img_rows
    else:
        r = _text_area(ws, r, 18, "（カタログデータなし）")


# ── シート 5: 用語集 ────────────────────────────────────────────
def _build_glossary_sheet(ws, glossary: list):
    _setup_grid(ws)
    r = 2
    _set_h(ws, 1, 8)
    r = _title_row(ws, r, "用語集")
    r = _margin(ws, r)
    r = _section_row(ws, r, "業務用語・Salesforce用語 対照表")
    r = _hdr_row(ws, r, [(2,3,"No"),(4,10,"業務用語"),
                          (11,18,"Salesforce用語 / オブジェクト名"),(19,31,"説明")])
    for i, t in enumerate(glossary):
        r = _data_row(ws, r, [(2,3,str(i+1)),(4,10,t["biz"]),
                               (11,18,t["sf"]),(19,31,t["desc"])], row_h=18)
    if len(glossary) < 10:
        r = _empty_rows(ws, r, 10 - len(glossary), [(2,3),(4,10),(11,18),(19,31)], row_h=18)


# ── メイン ──────────────────────────────────────────────────────
def generate(docs_dir: Path, output: Path, author: str, project_name: str):
    org      = parse_org(docs_dir / "overview" / "org-profile.md")
    req      = parse_requirements(docs_dir / "requirements" / "requirements.md")
    system   = parse_system_json(docs_dir / "architecture" / "system.json")
    swim     = parse_swimlanes(docs_dir / "flow" / "swimlanes.json")
    objects  = parse_catalog_index(docs_dir / "catalog" / "_index.md")
    relations = parse_data_model(docs_dir / "catalog" / "_data-model.md")

    if project_name:
        org["project_name"] = project_name

    # requirements.mdに背景がなければorg-profile.mdの背景を使う
    if not req.get("background") and org.get("background"):
        req["background"] = org["background"]

    asis_flow, tobe_flow = _pick_flows(swim)

    # 図を一時ディレクトリに生成
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        sys_img = er_img = asis_img = tobe_img = None

        if system:
            try:
                dg.render_system_diagram(system, str(tmp / "system.png"))
                sys_img = str(tmp / "system.png")
                print("  [OK] システム構成図")
            except Exception as e:
                print(f"  [WARN] システム構成図: {e}")

        if objects and relations:
            try:
                dg.render_er_diagram(objects, relations, str(tmp / "er.png"))
                er_img = str(tmp / "er.png")
                print("  [OK] ER図")
            except Exception as e:
                print(f"  [WARN] ER図: {e}")

        if asis_flow:
            try:
                dg.render_swimlane(asis_flow, str(tmp / "asis.png"))
                asis_img = str(tmp / "asis.png")
                print("  [OK] As-Is フロー")
            except Exception as e:
                print(f"  [WARN] As-Is フロー: {e}")

        if tobe_flow:
            try:
                dg.render_swimlane(tobe_flow, str(tmp / "tobe.png"))
                tobe_img = str(tmp / "tobe.png")
                print("  [OK] To-Be フロー")
            except Exception as e:
                print(f"  [WARN] To-Be フロー: {e}")

        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("表紙")
        ws2 = wb.create_sheet("システム概要")
        ws3 = wb.create_sheet("業務フロー図")
        ws4 = wb.create_sheet("ER図")
        ws5 = wb.create_sheet("用語集")

        _build_cover(ws1, org, req, author)
        _build_system_overview(ws2, req, system, sys_img)
        _build_flow_sheet(ws3, asis_flow, tobe_flow, asis_img, tobe_img)
        _build_er_sheet(ws4, objects, relations, er_img)
        _build_glossary_sheet(ws5, org.get("glossary", []))

        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output))

    print(f"saved: {output}")


def main():
    ap = argparse.ArgumentParser(description="プロジェクト概要書.xlsx 生成")
    ap.add_argument("--docs-dir",     required=True)
    ap.add_argument("--output",       required=True)
    ap.add_argument("--author",       default="")
    ap.add_argument("--project-name", default="")
    args = ap.parse_args()
    generate(Path(args.docs_dir), Path(args.output),
             args.author, args.project_name)


if __name__ == "__main__":
    main()
