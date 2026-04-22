# scripts/python/sf-doc-mcp/generate_detail_design.py
# -*- coding: utf-8 -*-
"""
詳細設計書.xlsx を1機能分生成する（テンプレート読込方式・新JSONスキーマ対応）。

  詳細設計書テンプレート.xlsx（build_detail_design_template.py で生成した「器」）を
  コピーしてセル値 + 図形PNGを流し込む。

6シート構成:
  1. 改版履歴           : メタ + 履歴テーブル
  2. 概要               : 機能名 / 機能概要 / 目的 / 利用者 / 起点画面 / 操作トリガー
  3. 業務フロー         : スイムレーン図PNG + フロー表(No/アクター/処理内容/分岐条件)
  4. 対象オブジェクト   : ER図PNG + 項目表(オブジェクト名/項目API名/項目ラベル/読み書き区分/備考)
  5. 処理概要           : フローチャートPNG + 処理表(No/処理内容/コンポーネント/分岐条件)
  6. 関連コンポーネント : コンポーネント図PNG + 一覧表(コンポーネント名/種別/役割/依存方向)

Usage:
  python generate_detail_design.py \\
    --input  detail_design.json \\
    --template "C:/.../詳細設計書テンプレート.xlsx" \\
    --output-dir "C:/.../出力先" \\
    [--version-increment minor]
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from datetime import date as _date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import design_revision as dr
from build_detail_design_template import (
    section_band, diagram_area, data_rows, setup_grid, set_h,
    GRID_LEFT, GRID_RIGHT,
    C_BAND_BLUE, C_TITLE_DARK,
)
from meta_store import read_meta, write_meta
from version_manager import increment_version

# ── 色定数 ─────────────────────────────────────────────────────────
C_HDR_BLUE  = "2E75B6"
C_BAND_BLUE = "0070C0"
C_LABEL_BG  = "D9E1F2"
C_FONT_D    = "000000"
C_FONT_W    = "FFFFFF"

THIN = Side(style="thin", color="8B9DC3")

# ── テンプレート行番号定数 ─────────────────────────────────────────
# build_detail_design_template.py の構造から算出
# 改版履歴
REV_META_ROW       = 3
REV_META_PROJECT_V = (6, 18)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6,  11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 31),
}

# 概要（row3〜row8: 機能名/機能概要/目的/利用者/起点画面/操作トリガー）
OV_LABEL_VAL_CS = 7
OV_LABEL_VAL_CE = 31
OV_ROWS = {
    "name_ja":        3,
    "summary":        4,
    "purpose":        5,
    "users":          6,
    "trigger_screen": 7,
    "trigger":        8,
}

# 業務フロー: テーブルヘッダは row4、データは row5 から動的追加
BF_DATA_ROW_START  = 5
BF_STEP_CS,  BF_STEP_CE  = 2,  3
BF_ACTOR_CS, BF_ACTOR_CE = 4,  8
BF_ACT_CS,   BF_ACT_CE   = 9,  31
# BF_COL_GROUPS: データ行のマージセル定義
BF_COL_GROUPS = [
    (BF_STEP_CS, BF_STEP_CE),
    (BF_ACTOR_CS, BF_ACTOR_CE),
    (BF_ACT_CS, BF_ACT_CE),
]

# 対象オブジェクト: 表上・図下レイアウト（動的行）
# ※ 項目ラベルが左（8-14）、API名が右（15-20）の順
OBJ_DATA_ROW_START = 5
OBJ_NAME_CS,  OBJ_NAME_CE  = 2,  7
OBJ_FLBL_CS,  OBJ_FLBL_CE  = 8,  14   # 項目ラベル（左）
OBJ_FAPI_CS,  OBJ_FAPI_CE  = 15, 20   # 項目API名（右）
OBJ_ACC_CS,   OBJ_ACC_CE   = 21, 23
OBJ_NOTE_CS,  OBJ_NOTE_CE  = 24, 31
OBJ_COL_GROUPS = [
    (OBJ_NAME_CS, OBJ_NAME_CE),
    (OBJ_FLBL_CS, OBJ_FLBL_CE),
    (OBJ_FAPI_CS, OBJ_FAPI_CE),
    (OBJ_ACC_CS,  OBJ_ACC_CE),
    (OBJ_NOTE_CS, OBJ_NOTE_CE),
]

# 処理概要: テーブルヘッダは row4、データは row5 から動的追加
PROC_DATA_ROW_START = 5
PROC_STEP_CS, PROC_STEP_CE = 2,  3
PROC_DESC_CS, PROC_DESC_CE = 4,  19
PROC_COMP_CS, PROC_COMP_CE = 20, 31
PROC_COL_GROUPS = [
    (PROC_STEP_CS, PROC_STEP_CE),
    (PROC_DESC_CS, PROC_DESC_CE),
    (PROC_COMP_CS, PROC_COMP_CE),
]

# 関連コンポーネント: テーブルヘッダは row4、データは row5 から動的追加
COMP_DATA_ROW_START = 5
COMP_NAME_CS, COMP_NAME_CE = 2,  9
COMP_TYPE_CS, COMP_TYPE_CE = 10, 13
COMP_ROLE_CS, COMP_ROLE_CE = 14, 31  # 依存方向列廃止のため役割を31まで拡張
COMP_COL_GROUPS = [
    (COMP_NAME_CS, COMP_NAME_CE),
    (COMP_TYPE_CS, COMP_TYPE_CE),
    (COMP_ROLE_CS, COMP_ROLE_CE),
]

GRID_RIGHT_C = 31

SCALAR_FIELDS  = ["summary", "purpose", "users", "trigger_screen", "trigger"]
SECTION_SHEETS = {
    "business_flow":    "業務フロー",
    "process_steps":    "処理概要",
    "related_objects":  "対象オブジェクト",
    "components":       "関連コンポーネント",
}

# 動的シートの図エリア高さ（行数）
DIAGRAM_AREA_ROWS = 30
# 動的シートの空行追加数（手動入力用）— 0=データ行のみ
DYNAMIC_EMPTY_ROWS = 0


# ── スタイルヘルパー ────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

# set_h is imported from build_detail_design_template


# ── PNG埋め込み ────────────────────────────────────────────────────
def _embed_image(ws, png_path: str, anchor: str,
                 img_w: int = 840, img_h: int | None = None,
                 max_h: int | None = None):
    """PNGをExcelシートに埋め込む。img_h=Noneなら縦横比維持。max_h で高さ上限を設定。"""
    try:
        if not Path(png_path).exists():
            return
        img = XLImage(png_path)
        img.anchor = anchor
        if img_h is None:
            ratio = img.height / img.width if img.width else 1.0
            w, h = img_w, int(img_w * ratio)
            if max_h and h > max_h:
                h = max_h
                w = int(h / ratio)
            img.width = w
            img.height = h
        else:
            img.width = img_w
            img.height = img_h
        ws.add_image(img)
    except Exception as e:
        print(f"  [WARN] 画像埋め込み失敗({anchor}): {e}")


# ── シート埋め込み ─────────────────────────────────────────────────
def fill_revision(ws, data: dict, history: list[dict]):
    vs, _ = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("project_name", ""))
    vs, _ = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("date", ""))
    dr.fill_revision_table(ws, history, REV_COLS, REV_DATA_ROW_START)


def fill_overview(ws, data: dict, changed_fields: set):
    """概要シートに値を書き込む。"""
    for key, row in OV_ROWS.items():
        val = data.get(key, "")
        if val:
            cell = ws.cell(row=row, column=OV_LABEL_VAL_CS, value=val)
            if key in changed_fields:
                dr.apply_red(cell, size=10)


def fill_business_flow(ws, data: dict, changed_step_nos: set,
                       png_path: str | None):
    """業務フローシート: テーブル(動的行) + スイムレーン図。"""
    flows = data.get("business_flow", [])
    n_data = len(flows)
    total_rows = n_data + DYNAMIC_EMPTY_ROWS

    # データ行 + 空行の枠を作成
    r = BF_DATA_ROW_START
    data_rows(ws, r, r + total_rows - 1, BF_COL_GROUPS, row_h=24)

    # データ書き込み
    for i, flow in enumerate(flows):
        step_no = flow.get("step", i + 1)
        is_changed = step_no in changed_step_nos
        set_h(ws, r, 24)

        nexts = flow.get("next", [])
        c1 = MW(ws, r, BF_STEP_CS,  BF_STEP_CE,  step_no,
                border=B_all(), h="center")
        c2 = MW(ws, r, BF_ACTOR_CS, BF_ACTOR_CE, flow.get("actor", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, BF_ACT_CS,   BF_ACT_CE,   flow.get("action", ""),
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # スペーサー + 図エリア
    spacer_row = BF_DATA_ROW_START + total_rows
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "業務フロー図（自動生成）", section_no=2)

    # 図埋め込み
    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        _embed_image(ws, png_path, img_anchor, img_w=880)


def _compute_obj_note(obj_api: str, field_access: str, data: dict) -> str:
    """object_access + process_steps から項目ごとの備考テキストを生成する。

    フィールドの読み書き区分に応じて関連するアクセスのみ抽出する。
    例（読み取り専用項目）: 「プリチェック判定時に参照。」
    例（書き込み項目）: 「コンサルテーション依頼時に更新・登録。見積作成時に新規作成。」
    """
    _OP_JA = {"R": "参照", "W": "更新・登録", "RW": "参照・更新", "INSERT": "新規作成"}
    _WRITE_OPS = {"W", "RW", "INSERT"}
    _READ_OPS  = {"R", "RW"}

    object_access = data.get("object_access", [])
    process_steps = data.get("process_steps", [])

    # comp_api_name（API名）でルックアップ。なければ component（タイプ名）でフォールバック
    comp_to_step: dict[str, str] = {}
    for ps in process_steps:
        key = ps.get("comp_api_name") or ps.get("component", "")
        if key and key not in comp_to_step:
            comp_to_step[key] = ps.get("title", "")

    accesses = [a for a in object_access if a.get("object") == obj_api]
    if not accesses:
        return ""

    # フィールドの読み書き区分に応じて関連するアクセスのみ抽出
    relevant = [
        a for a in accesses
        if (field_access == "R" and a.get("operation", "") in _READ_OPS)
        or (field_access in ("W", "INSERT") and a.get("operation", "") in _WRITE_OPS)
        or field_access == "RW"
    ]
    if not relevant:
        relevant = accesses  # フォールバック

    # 備考には「どの処理ステップで使われるか」のみ記載（操作種別は読み書き区分列と重複するため省く）
    step_titles = []
    for acc in relevant:
        step_title = comp_to_step.get(acc.get("component", ""), "")
        if step_title and step_title not in step_titles:
            step_titles.append(step_title)
    return "・".join(step_titles) if step_titles else ""


def fill_target_objects(ws, data: dict, changed_obj_keys: set,
                        png_path: str | None):
    """対象オブジェクトシート: 項目テーブル(動的行・縦結合) + 図。

    列順: オブジェクト名(縦結合) | 項目ラベル | 項目API名 | 読み書き区分(日本語) | 備考
    """
    _ACCESS_JA = {
        "R": "参照", "W": "更新", "RW": "参照・更新", "INSERT": "新規作成",
    }
    objects = data.get("related_objects", [])

    r = OBJ_DATA_ROW_START

    for obj in objects:
        obj_api   = obj.get("api_name", "")
        obj_label = f"{obj.get('label', '')} ({obj_api})"
        is_changed = obj_api in changed_obj_keys
        fields = obj.get("fields", [])
        if not fields:
            continue
        n_fields      = len(fields)
        obj_start_row = r

        # 読み書き区分・備考はオブジェクト単位で統一（フィールドごとに変わらないため縦結合）
        # obj['access_ja'] が設定されていればそれを優先（R+INSERT 等の複合操作を正しく表示）
        access    = fields[0].get("access", "") if fields else ""
        access_ja = obj.get("access_ja") or _ACCESS_JA.get(access, access)
        note      = fields[0].get("note", "") if fields else ""
        if not note:
            note = _compute_obj_note(obj_api, access, data)

        for fi, field in enumerate(fields):
            set_h(ws, r, 22)

            # オブジェクト名列: 全行書くが後で縦結合する
            MW(ws, r, OBJ_NAME_CS, OBJ_NAME_CE,
               obj_label if fi == 0 else "", border=B_all())

            c2 = MW(ws, r, OBJ_FLBL_CS, OBJ_FLBL_CE, field.get("label", ""),
                    border=B_all())
            c3 = MW(ws, r, OBJ_FAPI_CS, OBJ_FAPI_CE, field.get("api_name", ""),
                    border=B_all())
            # 読み書き区分・備考は初行のみ書き込み（後で縦結合）
            c4 = MW(ws, r, OBJ_ACC_CS,  OBJ_ACC_CE,  access_ja if fi == 0 else "",
                    border=B_all(), h="center")
            c5 = MW(ws, r, OBJ_NOTE_CS, OBJ_NOTE_CE, note if fi == 0 else "",
                    border=B_all(), wrap=True, v="top")
            if is_changed:
                for c in (c2, c3, c4, c5):
                    dr.apply_red(c)
            r += 1

        def _merge_col(ws, row_start, row_end, cs, ce, value, is_changed_flag, h="left", v="center"):
            for ri in range(row_start, row_end + 1):
                try:
                    ws.unmerge_cells(start_row=ri, start_column=cs,
                                     end_row=ri, end_column=ce)
                except Exception:
                    pass
            ws.merge_cells(start_row=row_start, start_column=cs,
                           end_row=row_end, end_column=ce)
            mc = ws.cell(row=row_start, column=cs)
            mc.value     = value
            mc.font      = _fnt()
            mc.alignment = _aln(h=h, v=v, wrap=True)
            mc.border    = B_all()
            if is_changed_flag:
                dr.apply_red(mc)

        # オブジェクト名・読み書き区分・備考を縦結合
        if n_fields > 1:
            _merge_col(ws, obj_start_row, r - 1, OBJ_NAME_CS, OBJ_NAME_CE,
                       obj_label, is_changed, v="center")
            _merge_col(ws, obj_start_row, r - 1, OBJ_ACC_CS, OBJ_ACC_CE,
                       access_ja, is_changed, h="center", v="center")
            _merge_col(ws, obj_start_row, r - 1, OBJ_NOTE_CS, OBJ_NOTE_CE,
                       note, is_changed, v="top")

    # スペーサー + 図エリア（動的位置）
    spacer_row = r
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "オブジェクト関連図（自動生成）", section_no=2)

    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        _embed_image(ws, png_path, img_anchor, img_w=880)


def _estimate_row_height(text: str, chars_per_line: int = 34,
                         line_pt: int = 14, min_h: int = 24, max_h: int = 300) -> int:
    """テキストの折り返しを考慮して行の高さ（ポイント）を推定する。"""
    import math as _math
    if not text:
        return min_h
    lines = text.split("\n")
    total = sum(_math.ceil(max(len(ln), 1) / chars_per_line) for ln in lines)
    return min(max_h, max(min_h, total * line_pt + 8))


def fill_process_overview(ws, data: dict, changed_step_nos: set,
                          png_path: str | None):
    """処理概要シート: テーブル(動的行・動的高) + フローチャート。"""
    steps = data.get("process_steps", [])
    n_data = len(steps)
    total_rows = n_data + DYNAMIC_EMPTY_ROWS

    # データ行の枠を作成（デフォルト高は後で上書き）
    r = PROC_DATA_ROW_START
    data_rows(ws, r, r + total_rows - 1, PROC_COL_GROUPS, row_h=30)

    for i, ps in enumerate(steps):
        step_no = ps.get("step", i + 1)
        is_changed = step_no in changed_step_nos

        desc_text = ps.get("description", "").strip()
        row_h = _estimate_row_height(desc_text)
        set_h(ws, r, row_h)

        component = ps.get("component") or ""
        branch = ps.get("branch") or ""

        c1 = MW(ws, r, PROC_STEP_CS, PROC_STEP_CE, step_no,
                border=B_all(), h="center")
        c2 = MW(ws, r, PROC_DESC_CS, PROC_DESC_CE, desc_text,
                border=B_all(), wrap=True, v="top")
        c3 = MW(ws, r, PROC_COMP_CS, PROC_COMP_CE, component,
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # スペーサー + 図エリア
    spacer_row = PROC_DATA_ROW_START + total_rows
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "処理フロー図（自動生成）", section_no=2)

    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        _embed_image(ws, png_path, img_anchor, img_w=880)


def fill_related_components(ws, data: dict, changed_comp_keys: set,
                            png_path: str | None):
    """関連コンポーネントシート: テーブル(動的行) + コンポーネント図。"""
    components = data.get("components", [])
    n_data = len(components)
    total_rows = n_data + DYNAMIC_EMPTY_ROWS

    r = COMP_DATA_ROW_START
    data_rows(ws, r, r + total_rows - 1, COMP_COL_GROUPS, row_h=24)

    for comp in components:
        api_name = comp.get("api_name", "")
        is_changed = api_name in changed_comp_keys
        set_h(ws, r, 24)

        c1 = MW(ws, r, COMP_NAME_CS, COMP_NAME_CE, api_name,
                border=B_all())
        c2 = MW(ws, r, COMP_TYPE_CS, COMP_TYPE_CE, comp.get("type", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, COMP_ROLE_CS, COMP_ROLE_CE, comp.get("role", ""),
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # スペーサー + 図エリア
    spacer_row = COMP_DATA_ROW_START + total_rows
    set_h(ws, spacer_row, 8)
    diagram_start = spacer_row + 1
    diagram_area(ws, diagram_start, "コンポーネント関連図（自動生成）", section_no=2)

    img_anchor = f"B{diagram_start + 1}"
    if png_path:
        _embed_image(ws, png_path, img_anchor, img_w=880)



# ── GFスキーマ正規化 ────────────────────────────────────────────────
import re as _re

# ── SFプロジェクト → メタデータパス マッピング ───────────────────────────
# flows/classes は greenfield、フィールドラベル翻訳は両方から取得（GF_UATが補完）
_SF_PROJECT_PATHS: dict[str, str] = {
    # キーは小文字で統一。data.get('project_name') は .lower() して照合する
    "greenfield": "C:/workspace/16_グリーンフィールド/greenfield",
}
_SF_EXTRA_LABEL_PATHS: list[str] = [
    "C:/workspace/16_グリーンフィールド/GF_UAT",
]
# メタデータから構築するフィールドラベルマップ {obj_api: {field_api: ja_label}}
_SF_FIELD_LABELS: dict[str, dict[str, str]] = {}
# オブジェクトラベルマップ {obj_api: ja_label}
_SF_OBJ_LABELS: dict[str, str] = {}
# コンポーネント別フィールドマップ {comp_api_name: {obj_api: {field_api}}}
_SF_COMP_FIELDS: dict[str, dict[str, set]] = {}


def _parse_flow_fields(flow_path: Path) -> dict[str, set]:
    """Flow XMLから {obj_api: {field_api}} を抽出する。"""
    try:
        content = flow_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return {}
    result: dict[str, set] = {}
    tags = ("recordCreates", "recordUpdates", "recordLookups", "recordDeletes")
    for tag in tags:
        for block in _re.findall(rf'<{tag}>(.*?)</{tag}>', content, _re.DOTALL):
            obj_m = _re.search(r'<object>([A-Za-z0-9_]+)</object>', block)
            if not obj_m:
                continue
            obj_api = obj_m.group(1)
            fields = set(_re.findall(r'<field>([A-Za-z0-9_]+)</field>', block))
            if fields:
                result.setdefault(obj_api, set()).update(fields)
    return result


def _parse_apex_fields(cls_path: Path) -> dict[str, set]:
    """Apexクラスから SOQL + DML + トリガーハンドラーのフィールドを抽出する。
    Returns {obj_api: {field_api}}
    """
    try:
        content = cls_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return {}
    result: dict[str, set] = {}

    # SOQL: SELECT f1, f2 FROM ObjectName
    for m in _re.finditer(
        r'SELECT\s+(.*?)\s+FROM\s+([A-Za-z0-9_]+)', content, _re.IGNORECASE | _re.DOTALL
    ):
        fields_str, obj_api = m.group(1), m.group(2)
        fields = {
            f.strip() for f in _re.split(r'[\s,]+', fields_str)
            if f.strip() and f.strip().lower() != "id" and _re.match(r'^[A-Za-z]\w*$', f.strip())
        }
        if fields:
            result.setdefault(obj_api, set()).update(fields)

    # DML: variable.FieldName__c = ...（オブジェクト特定不可のため __any__ に保持）
    for fapi in _re.findall(r'\.\s*([A-Za-z][A-Za-z0-9_]*__c)\s*=', content):
        result.setdefault("__any__", set()).add(fapi)

    # トリガーハンドラー: メソッド引数・ローカル変数の型宣言からプライマリオブジェクトを特定
    # 例: Map<Id, ViewAblePerson__c> や List<ViewAblePerson__c> → ViewAblePerson__c がプライマリ
    trigger_obj = None
    for pat in [r'Map<Id,\s*([A-Za-z][A-Za-z0-9]*__c)>',
                r'List<([A-Za-z][A-Za-z0-9]*__c)>',
                r'([A-Za-z][A-Za-z0-9]*__c)\s+\w+\s*[=;,)]']:
        m = _re.search(pat, content)
        if m:
            trigger_obj = m.group(1)
            break
    if trigger_obj:
        # そのオブジェクト型変数へのプロパティアクセス (.Field__c) を収集
        prop_fields = set(_re.findall(
            r'\b\w+\.\s*([A-Za-z][A-Za-z0-9]*__[cr])\b', content
        ))
        # Id系・リレーション(__r)は除外
        prop_fields = {f for f in prop_fields if not f.endswith('__r')}
        if prop_fields:
            result.setdefault(trigger_obj, set()).update(prop_fields)

    return result


def _load_sf_metadata(sf_project_path: str) -> None:
    """SFプロジェクトの objectTranslations/flows/classes からメタデータを構築してグローバルに格納する。"""
    global _SF_FIELD_LABELS, _SF_OBJ_LABELS, _SF_COMP_FIELDS
    base = Path(sf_project_path) / "force-app/main/default"

    # objectTranslations: フィールド・オブジェクトの日本語ラベル
    trans_dir = base / "objectTranslations"
    if trans_dir.exists():
        for obj_dir in trans_dir.iterdir():
            if not obj_dir.name.endswith("-ja"):
                continue
            obj_api = obj_dir.name[:-3]
            obj_trans = obj_dir / f"{obj_dir.name}.objectTranslation-meta.xml"
            if obj_trans.exists():
                content = obj_trans.read_text(encoding="utf-8")
                m = _re.search(r'<value>([^<]+)</value>', content)
                if m:
                    _SF_OBJ_LABELS[obj_api] = m.group(1).strip()
            for fxml in obj_dir.glob("*.fieldTranslation-meta.xml"):
                field_api = fxml.name.replace(".fieldTranslation-meta.xml", "")
                content = fxml.read_text(encoding="utf-8")
                m = _re.search(r'<label><!--\s*(.*?)\s*--></label>', content)
                if m and m.group(1):
                    _SF_FIELD_LABELS.setdefault(obj_api, {})[field_api] = m.group(1).strip()

    # flows: コンポーネント別のオブジェクト+フィールド
    flows_dir = base / "flows"
    if flows_dir.exists():
        for flow_file in flows_dir.glob("*.flow-meta.xml"):
            comp_api = flow_file.name.replace(".flow-meta.xml", "")
            obj_fields = _parse_flow_fields(flow_file)
            if obj_fields:
                _SF_COMP_FIELDS[comp_api] = {k: v for k, v in obj_fields.items()}

    # classes: コンポーネント別のオブジェクト+フィールド（テストクラスは除外）
    classes_dir = base / "classes"
    if classes_dir.exists():
        for cls_file in classes_dir.glob("*.cls"):
            if cls_file.name.endswith("Test.cls"):
                continue
            comp_api = cls_file.name.replace(".cls", "")
            obj_fields = _parse_apex_fields(cls_file)
            if obj_fields:
                _SF_COMP_FIELDS[comp_api] = {k: v for k, v in obj_fields.items()}

    # objects: field-meta.xml からラベルを補完（objectTranslations にないカスタムフィールド対応）
    objects_dir = base / "objects"
    if objects_dir.exists():
        for obj_dir in objects_dir.iterdir():
            if not obj_dir.is_dir():
                continue
            obj_api = obj_dir.name
            fields_dir = obj_dir / "fields"
            if not fields_dir.exists():
                continue
            for fxml in fields_dir.glob("*.field-meta.xml"):
                field_api = fxml.name.replace(".field-meta.xml", "")
                if field_api in _SF_FIELD_LABELS.get(obj_api, {}):
                    continue  # 翻訳ファイルで既に取得済み
                try:
                    content = fxml.read_text(encoding="utf-8")
                except Exception:
                    continue
                m = _re.search(r'<label>([^<]+)</label>', content)
                if m:
                    _SF_FIELD_LABELS.setdefault(obj_api, {})[field_api] = m.group(1).strip()

    # 追加翻訳パス（GF_UAT等）からobjectTranslationsのみ補完ロード
    for extra_path in _SF_EXTRA_LABEL_PATHS:
        extra_trans = Path(extra_path) / "force-app/main/default/objectTranslations"
        if not extra_trans.exists():
            continue
        for obj_dir in extra_trans.iterdir():
            if not obj_dir.name.endswith("-ja"):
                continue
            obj_api = obj_dir.name[:-3]
            for fxml in obj_dir.glob("*.fieldTranslation-meta.xml"):
                field_api = fxml.name.replace(".fieldTranslation-meta.xml", "")
                if field_api in _SF_FIELD_LABELS.get(obj_api, {}):
                    continue  # 既に取得済み
                try:
                    content = fxml.read_text(encoding="utf-8")
                except Exception:
                    continue
                m = _re.search(r'<label><!--\s*(.*?)\s*--></label>', content)
                if m and m.group(1):
                    _SF_FIELD_LABELS.setdefault(obj_api, {})[field_api] = m.group(1).strip()

    # 標準オブジェクトの主要フィールド日本語ラベル（翻訳ファイルが存在しない場合の補完）
    _STD_FIELD_LABELS_BUILTIN: dict[str, dict[str, str]] = {
        "User": {
            "Username": "ユーザー名", "Email": "メールアドレス", "IsActive": "有効",
            "FirstName": "名", "LastName": "姓", "Name": "フルネーム",
            "ContactId": "取引先責任者", "AccountId": "取引先", "ProfileId": "プロファイル",
            "UserRoleId": "ロール", "Title": "役職", "Department": "部門",
            "Phone": "電話", "MobilePhone": "携帯", "LanguageLocaleKey": "言語",
            "LocaleSidKey": "ロケール", "TimeZoneSidKey": "タイムゾーン",
        },
        "Account": {
            "Name": "取引先名", "Phone": "電話", "Website": "Webサイト",
            "BillingAddress": "請求先住所", "ShippingAddress": "配送先住所",
            "Industry": "業種", "AnnualRevenue": "年間売上", "NumberOfEmployees": "従業員数",
            "OwnerId": "所有者", "Type": "取引先種別", "ParentId": "親取引先",
        },
        "Contact": {
            "FirstName": "名", "LastName": "姓", "Name": "氏名",
            "Email": "メールアドレス", "Phone": "電話", "MobilePhone": "携帯",
            "AccountId": "取引先", "Title": "役職", "Department": "部門",
            "OwnerId": "所有者", "Birthdate": "生年月日",
        },
        "Lead": {
            "FirstName": "名", "LastName": "姓", "Name": "氏名",
            "Company": "会社名", "Email": "メールアドレス", "Phone": "電話",
            "MobilePhone": "携帯", "Title": "役職", "Industry": "業種",
            "LeadSource": "リードソース", "Status": "ステータス",
            "Street": "住所（番地）", "City": "市区町村", "State": "都道府県",
            "PostalCode": "郵便番号", "Country": "国", "OwnerId": "所有者",
            "HasOptedOutOfEmail": "メール配信停止", "IsConverted": "取引先に変換済み",
        },
        "Opportunity": {
            "Name": "商談名", "AccountId": "取引先", "CloseDate": "完了予定日",
            "StageName": "フェーズ", "Amount": "金額", "Probability": "確度",
            "OwnerId": "所有者", "LeadSource": "リードソース", "Type": "種別",
        },
    }
    for obj_api, fld_map in _STD_FIELD_LABELS_BUILTIN.items():
        for fapi, flabel in fld_map.items():
            _SF_FIELD_LABELS.setdefault(obj_api, {}).setdefault(fapi, flabel)


def _sf_field_label(obj_api: str, field_api: str) -> str:
    """フィールドAPI名を日本語ラベルに変換する（メタデータ優先、なければ加工済みAPI名）。"""
    label = _SF_FIELD_LABELS.get(obj_api, {}).get(field_api)
    if label:
        return label
    return field_api.replace("__c", "").replace("__", "_")


def _sf_obj_label(obj_api: str) -> str:
    """オブジェクトAPI名を日本語ラベルに変換する。"""
    if obj_api in _STD_OBJ_LABELS:
        return _STD_OBJ_LABELS[obj_api]
    return _SF_OBJ_LABELS.get(obj_api) or _obj_label_from_api(obj_api)


# ── Salesforce 標準オブジェクト API名 → 日本語ラベルマップ ───────────────
_STD_OBJ_LABELS = {
    "ContentDocumentLink": "コンテンツ紐付けレコード",
    "ContentDocument":     "コンテンツドキュメント",
    "ContentVersion":      "コンテンツファイル",
    "EmailMessage":        "メールメッセージ",
    "Attachment":          "添付ファイル",
    "Opportunity":         "商談",
    "Contact":             "取引先責任者",
    "Account":             "取引先",
    "Lead":                "リード",
    "Case":                "ケース",
    "Task":                "ToDo",
    "Event":               "行動",
    "User":                "ユーザー",
    "Quote":               "見積",
}

# 技術英語ジャーゴン → 日本語変換（SF オブジェクト名以外の技術用語）
# NOTE: re.ASCII により \b が ASCII 文字のみを word char として扱う（日本語隣接でも正しくマッチ）
# NOTE: "before insert" 等の複合パターンは単体 "insert" より前に置く（置換順依存を防ぐ）
_A = _re.ASCII  # shorthand
_AI = _re.ASCII | _re.IGNORECASE
_JARGON_JA: list[tuple] = [
    # Apex トリガー複合表現（単体 insert/delete/update より先に処理）
    (_re.compile(r'\bbefore\s+insert\b', _AI), '処理前（新規）'),
    (_re.compile(r'\bafter\s+insert\b',  _AI), '処理後（新規）'),
    (_re.compile(r'\bbefore\s+update\b', _AI), '処理前（更新）'),
    (_re.compile(r'\bafter\s+update\b',  _AI), '処理後（更新）'),
    (_re.compile(r'\bbefore\s+delete\b', _AI), '処理前（削除）'),
    (_re.compile(r'\bafter\s+delete\b',  _AI), '処理後（削除）'),
    # SF API / 外部サービス
    (_re.compile(r'\bOPROARTS\s+API\b', _A), '外部帳票サービス'),
    (_re.compile(r'\bOPROARTS\b', _A), '外部帳票サービス'),
    # コレクション型
    (_re.compile(r'\bList<Id>\b', _A), 'IDリスト'),
    (_re.compile(r'\bList<[A-Za-z]+>\b', _A), 'リスト'),
    # データ型
    (_re.compile(r'\bBlob\b', _A), 'バイナリデータ'),
    (_re.compile(r'\bvoid\b', _A), 'なし'),
    # DML 操作（単体）
    (_re.compile(r'\binsert\b', _AI), '新規作成'),
    (_re.compile(r'\bupsert\b', _AI), '登録・更新'),
    (_re.compile(r'\bdelete\b', _AI), '削除'),
    (_re.compile(r'\bupdate\b', _AI), '更新'),
    (_re.compile(r'\bquery\b',  _AI), '参照'),
    # Apex 固有クラス名・アノテーション
    (_re.compile(r'\bCustomerUser\b', _A), 'カスタマーユーザー'),
    (_re.compile(r'\b[A-Z][A-Za-z]*Tmp\b', _A), '一時データ'),
    (_re.compile(r'\bInvocableMethod\b', _A), 'フローアクション'),
    (_re.compile(r'\bAuraEnabled\b', _A), 'LWC公開メソッド'),
    # Apex 非同期アノテーション（後続の「で非同期実行する」も一緒に置換して重複を防ぐ）
    (_re.compile(r'@future(?:で非同期実行する|で実行される?)?'), '非同期で実行する'),
    (_re.compile(r'@\w+', _A), ''),
]

# 技術用語→日本語変換ルール（役割・説明文用）
_TECH_REPL = [
    # アノテーション → 日本語フレーズ
    # NOTE: _translate_jargon で InvocableMethod → フローアクション に変換済みのため
    # 変換後パターン "@フローアクション〜" も除去対象に含める
    (_re.compile(r'@InvocableMethod[としてで\s]*'), 'フローから呼び出され、'),
    (_re.compile(r'@フローアクション(?:として)?(?:Flowから呼ばれ)?[、\s]*'), 'フローから呼び出され、'),
    (_re.compile(r'@AuraEnabled[としてで\s]*'), 'LWCから呼び出され、'),
    (_re.compile(r'@LWC公開メソッド(?:として)?[、\s]*'), 'LWCから呼び出され、'),
    (_re.compile(r'@RemoteAction[としてで\s]*'), '非同期処理として呼び出され、'),
    (_re.compile(r'@\w+[としてで\s]*'), ''),  # 残った@アノテーションを除去
    # SOQL文を丸ごと除去（SELECT〜FROM〜を含む記述）
    (_re.compile(r'SELECT\s+.+?\s+FROM\s+\w+(?:\s+WHERE\s+[^。\n]+)?', _re.DOTALL | _re.IGNORECASE), ''),
    (_re.compile(r':\w+'), ''),  # SOQL bind変数
    # HTTPプロトコル技術記述（POST→302→GET等）を除去
    (_re.compile(r'[A-Z]{2,6}→\d{3}→[A-Z]{2,6}'), ''),
    # boolean変数が括弧で補足されているケース: （isXxx）（hasXxx）
    (_re.compile(r'[（(](?:is|has)[A-Z]\w*[）)]'), ''),
    # boolean条件文: "isXxxがtrueの場合は" → 変数名を除去し条件の結果のみ残す
    (_re.compile(r'(?:is|has)[A-Z]\w*が(?:false|true|null)(?:の場合[はに]?)?'), ''),
    # 単独のboolean変数（isXxx / hasXxx）を除去
    (_re.compile(r'(?<![A-Za-z])(?:is|has)[A-Z][A-Za-z]+(?![A-Za-z])'), ''),
    # 小文字始まりのcamelCase技術名（コンポーネント名等）が日本語助詞の前にある場合
    (_re.compile(r'(?<![A-Za-z])[a-z][a-zA-Z]{3,}(?=[をがはにでへのもと])'), ''),
    # 大文字始まりの長いCamelCaseクラス名が日本語に隣接する場合（VisaApplicationTypeMaster等）
    (_re.compile(r'[A-Z][a-zA-Z0-9]{6,}(?=[ぁ-ん一-龥ァ-ヶーをがはにでへのもと]|により|によって)'), ''),
    # クラス名.メソッド名 を除去
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+\([^)]*\)'), ''),
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+'), ''),
    # Apex トリガーイベント文脈（処理前/後）を除去
    (_re.compile(r'[^\s。]{2,20}(?:作成|更新|削除)時（処理前（(?:新規|更新|削除)）(?:/処理後（(?:新規|更新|削除)）)?）[にので]?'), ''),
    # 技術的なサーバーサイド表現を簡潔な日本語に
    (_re.compile(r'のサーバーサイドロジック'), ''),
    (_re.compile(r'のメインコンポーネント'), ''),
    (_re.compile(r'単一責務クラス'), 'クラス'),
    # "呼び出され、Flow〜から呼ばれ、" の重複を除去（ルックビハインド不可のため捕捉グループで実装）
    (_re.compile(r'(呼び出され)[、\s]{0,2}Flow[^\s、。]{0,20}から呼ばれ[、\s]*'), r'\1、'),
    # 同一表現の繰り返し（サービス名など）が括弧内外で重複: "外部帳票サービス（外部帳票サービス）"
    (_re.compile(r'([^\s（]{4,30})（\1）'), r'\1'),
    # メソッドが重複: "公開メソッドメソッド" → "公開メソッド"
    (_re.compile(r'(メソッド){2,}'), 'メソッド'),
    # 空括弧を除去: （）
    (_re.compile(r'（\s*）'), ''),
    (_re.compile(r'\(\s*\)'), ''),
    # 主語なし「・による〜」「/による〜」を除去（CamelCase除去後に残る）
    (_re.compile(r'[・、/]\s*による[^\s、。]{1,20}'), ''),
    # "/作成" を "・作成" に正規化し、重複した "・作成・作成" をまとめる
    (_re.compile(r'/作成'), '・作成'),
    (_re.compile(r'(?:・作成){2,}'), '・作成'),
    # 助詞・接続詞が文節先頭に孤立するケース（クラス名除去後）を修正
    (_re.compile(r'[・、]\s*を[ぁ-ん一-龥A-Za-z]{1,10}て(?=の)'), ''),  # を〜ての → 除去
    (_re.compile(r'[はがをにでへのも][。．]'), '。'),
    (_re.compile(r'^[、。\s]+'), ''),  # 文頭の余分な記号
    # 連続記号・空白の整理
    (_re.compile(r'[ \t]{2,}'), ' '),
    (_re.compile(r'[・、]{2,}'), '・'),
    (_re.compile(r'(、){2,}'), '、'),
    (_re.compile(r'(。){2,}'), '。'),
]

# 業務フロー・タイトル・概要用: SF標準オブジェクトは翻訳済み前提でAPIを除去
_TECH_REPL_BIZ = [
    (_re.compile(r'@InvocableMethod[としてで\s]*'), ''),
    (_re.compile(r'@AuraEnabled[としてで\s]*'), ''),
    (_re.compile(r'@RemoteAction[としてで\s]*'), ''),
    (_re.compile(r'@\w+'), ''),
    # SOQL・boolean変数・HTTPプロトコル除去（タイトル用）
    (_re.compile(r'SELECT\s+.+?\s+FROM\s+\w+(?:\s+WHERE\s+[^。\n]+)?', _re.DOTALL | _re.IGNORECASE), ''),
    (_re.compile(r':\w+'), ''),
    (_re.compile(r'[A-Z]{2,6}→\d{3}→[A-Z]{2,6}'), ''),
    (_re.compile(r'[（(](?:is|has)[A-Z]\w*[）)]'), ''),
    (_re.compile(r'(?:is|has)[A-Z]\w*が(?:false|true|null)(?:の場合[はに]?)?'), ''),
    (_re.compile(r'(?<![A-Za-z])(?:is|has)[A-Z][A-Za-z]+(?![A-Za-z])'), ''),
    (_re.compile(r'(?<![A-Za-z])[a-z][a-zA-Z]{3,}(?=[をがはにでへのもと])'), ''),
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+\([^)]*\)'), ''),
    (_re.compile(r'[A-Z][A-Za-z0-9]+\.[A-Za-z]\w+'), ''),
    (_re.compile(r'のサーバーサイドロジック'), ''),
    (_re.compile(r'のメインコンポーネント'), ''),
    (_re.compile(r'単一責務クラス'), 'クラス'),
    # List<CustomObj__c> → __c削除の前に除去（先に処理しないと List<> が残る）
    (_re.compile(r'List<[A-Z][A-Za-z0-9]*__[cepr]>'), 'レコードリスト'),
    (_re.compile(r'List<[A-Za-z]+>'), 'リスト'),
    # （trigger xxx） 等のApexトリガー技術的記述を除去
    (_re.compile(r'（trigger\s+\w+）'), ''),
    (_re.compile(r'\(trigger\s+\w+\)'), ''),
    # __c/__e/__r カスタムオブジェクト名（上記 List<> 除去後）
    (_re.compile(r'\b[A-Z][A-Za-z0-9]*__[cepr]\b'), ''),
    # Apex クラス名（Controller/Service/Handler/Manager/Batch/Trigger で終わるもの）
    (_re.compile(r'\b[A-Z][A-Za-z0-9]{2,}(?:Controller|Service|Handler|Manager|Batch|Trigger)\b'), ''),
    # 残ったCamelCase英語（日本語助詞に挟まれていない単独の英単語）
    (_re.compile(r'(?<![ぁ-ん一-龥ァ-ヶーa-z_])([A-Z][a-zA-Z]{3,})(?![ぁ-ん一-龥ァ-ヶーa-z_])'), ''),
    # 括弧内が英語・記号で始まる（技術情報）は除去
    (_re.compile(r'（[A-Z@#][^）]{0,60}）'), ''),
    (_re.compile(r'\([A-Z@#][^)]{0,60}\)'), ''),
    # 同一表現の繰り返し（括弧内外重複）: "外部帳票サービス（外部帳票サービス）"
    (_re.compile(r'([^\s（]{4,30})（\1）'), r'\1'),
    # メソッドが重複: "公開メソッドメソッド" → "公開メソッド"
    (_re.compile(r'(メソッド){2,}'), 'メソッド'),
    # 空括弧を除去
    (_re.compile(r'（\s*）'), ''),
    (_re.compile(r'\(\s*\)'), ''),
    # 主語なし「・による〜」を除去（CamelCase除去後に残る場合）
    (_re.compile(r'[・、/]\s*による[^\s、。]{1,20}'), ''),
    # "/作成" 正規化・重複まとめ
    (_re.compile(r'/作成'), '・作成'),
    (_re.compile(r'(?:・作成){2,}'), '・作成'),
    # Salesforce フィールド API名（__c/__r 等）を除去（\bはUnicode文字前で効かないため (?![A-Za-z0-9_]) を使用）
    (_re.compile(r'[A-Za-z][A-Za-z0-9_]*(?:__c|__r|__C|__R)(?![A-Za-z0-9_])'), ''),
    # フィールド名除去後の「の等」「の、」 → 整理
    (_re.compile(r'の(?=[等や・、。\s])'), ''),
    # 整理
    (_re.compile(r'[ \t]{2,}'), ' '),
    (_re.compile(r'[・、]{2,}'), '・'),
    (_re.compile(r'(、){2,}'), '、'),
    (_re.compile(r'^[、。・/\s]+|[、。・/\s]+$'), ''),
]

# 業務フロー step の先頭 preamble（呼び出し起点の技術的説明）を除去
_PREAMBLE_RE = _re.compile(
    r'^(?:[^、。]{0,25}として(?:呼ばれ|呼び出され)、|'
    r'Flowから呼ばれ、|'
    r'[^、。]{0,15}[がは]Flowから[^、。]{0,15}、)'
)


def _translate_sf_objects(text: str) -> str:
    """Salesforce 標準オブジェクト + ロード済みカスタムオブジェクト API名を日本語ラベルに置換する。

    NOTE: Python3 の re は Unicode モードで日本語文字も \\w 扱いするため、
    日本語に隣接する英語 API 名が word boundary にマッチしない。
    ASCII 専用の lookahead/lookbehind に変更。
    """
    for api, ja in _STD_OBJ_LABELS.items():
        text = _re.sub(rf'(?<![A-Za-z0-9_]){_re.escape(api)}(?![A-Za-z0-9_])', ja, text)
    # カスタムオブジェクト（メタデータ読み込み済みの場合のみ）
    for api, ja in _SF_OBJ_LABELS.items():
        text = _re.sub(rf'(?<![A-Za-z0-9_]){_re.escape(api)}(?![A-Za-z0-9_])', ja, text)
    return text


def _translate_jargon(text: str) -> str:
    """技術英語ジャーゴンを日本語に変換する。"""
    for pat, repl in _JARGON_JA:
        text = pat.sub(repl, text)
    return text


def _clean_tech(text: str) -> str:
    """役割・説明文用: アノテーション・クラス名.メソッド名を除去して日本語説明にする。"""
    for pattern, repl in _TECH_REPL:
        text = pattern.sub(repl, text)
    return text.strip()


_EC_PLACEHOLDER = "\x01EC\x01"

def _clean_tech_business(text: str) -> str:
    """業務フロー・タイトル・概要用: SF標準オブジェクトを日本語化→技術用語を全除去する。

    "Experience Cloud" はブランド名のため、CamelCase 除去パターンに巻き込まれないよう
    プレースホルダーで保護してから処理し、最後に復元する。
    """
    # Protect "Experience Cloud" before CamelCase removal
    text = text.replace("Experience Cloud", _EC_PLACEHOLDER)
    text = _translate_sf_objects(text)
    text = _translate_jargon(text)
    for pattern, repl in _TECH_REPL_BIZ:
        text = pattern.sub(repl, text)
    text = text.replace(_EC_PLACEHOLDER, "Experience Cloud")
    return text.strip()


def _translate_sf_fields(text: str) -> str:
    """ロード済みメタデータの全フィールドAPI名（__c等）を日本語ラベルに置換する。
    メタデータ未ロードの場合はそのまま返す（__c除去は _TECH_REPL_BIZ で後処理）。
    """
    for _obj_api, fields in _SF_FIELD_LABELS.items():
        for field_api, ja_label in fields.items():
            text = _re.sub(
                rf'(?<![A-Za-z0-9_]){_re.escape(field_api)}(?![A-Za-z0-9_])',
                ja_label, text,
            )
    return text


def _clean_io_text(text: str) -> str:
    """inputs/outputs テキストの技術用語を日本語化する（処理概要の説明文用）。"""
    text = text.replace("Experience Cloud", _EC_PLACEHOLDER)
    text = _translate_sf_objects(text)
    text = _translate_sf_fields(text)
    text = _translate_jargon(text)
    for pattern, repl in _TECH_REPL_BIZ:
        text = pattern.sub(repl, text)
    text = text.replace(_EC_PLACEHOLDER, "Experience Cloud")
    return text.strip()


def _short_title(responsibility: str, max_len: int = 18) -> str:
    """責務テキストから短い日本語アクションタイトルを生成する（フローチャートノード用）。

    方針:
      - 「〜を起点に、」「〜で起動される」「〜を受けて、」等のプリアンブルを除去
      - 語の途中で切らない（粒度: 日本語連続ラン / 助詞境界）
      - パターン1: `〜する画面フロー` → 「画面フローを〜する」
      - パターン2: `〜を[キーワード]` → 「〜を[キーワード]する」（最後の出現を優先）
      - パターン3: `[キーワード]` 単独 → 直前の日本語名詞句を補完
      - 18 文字以内に収める
    """
    clean = _clean_tech_business(responsibility)
    # 「Flowのアクションとして〜から」「〜クラスとして〜から」等のプリアンブルを除去
    clean = _re.sub(r'^[^。・\n]*?(?:アクション|クラス|ハンドラ|ハンドラー)として[^。・\n]*?から[、]?', '', clean).strip()
    clean = _PREAMBLE_RE.sub('', clean).strip()
    # 追加プリアンブル: トリガー記述に限定（読点付きのみ／Phase 0で拾える"起動する画面フロー"は残す）
    extra_preambles = [
        r'^[^。、]{0,40}?を起点に[、,]\s*',
        r'^[^。、]{0,40}?(?:で|から)起動される[、,]\s*',
        r'^[^。、]{0,40}?により[、,]\s*',
        r'^[^。、]{0,40}?を契機に[、,]\s*',
    ]
    for pat in extra_preambles:
        clean = _re.sub(pat, '', clean)
    # 末尾の「〜の責務を持つ」「〜責務を担う」を除去
    clean = _re.sub(r'(?:の)?責務を(?:持つ|担う|持ち)?。?$', '', clean).strip()

    _KW_PAIRS = [
        ("判定",   "を判定する"),
        ("検証",   "を検証する"),
        ("確認",   "を確認する"),
        ("チェック", "をチェックする"),
        ("作成",   "を作成する"),
        ("登録",   "を登録する"),
        ("更新",   "を更新する"),
        ("削除",   "を削除する"),
        ("同期",   "を同期する"),
        ("集約",   "を集約する"),
        ("送信",   "を送信する"),
        ("通知",   "を通知する"),
        ("制御",   "を制御する"),
        ("管理",   "を管理する"),
        ("実行",   "を実行する"),
        ("検知",   "を検知する"),
        ("委譲",   "を委譲する"),
        ("表示",   "を表示する"),
    ]

    # --- ヘルパー ---
    _BOUNDARIES = set("、。・「」『』（）() \t\n→")

    def _script_class(ch: str) -> str:
        if _re.match(r'[ぁ-んァ-ヶ一-龯々ー]', ch):
            return 'jp'
        if _re.match(r'[a-zA-Z0-9_]', ch):
            return 'en'
        return 'other'

    def _strip_leading_connectives(text: str) -> str:
        """先頭の接続語・助詞・動詞テ形・数字カウンタ等を除去する。"""
        prev = None
        # ループで除去 pass を繰り返す（複合プレフィックスに対応）
        while text != prev:
            prev = text
            text = _re.sub(r'^ー+', '', text)  # orphan 長音
            text = _re.sub(r'^[0-9０-９]+(?:[つ個件本回件枚人]の?)?', '', text)
            text = _re.sub(r'^(?:して|ために|ため|そして|また|により|ので|かつ|または|および)', '', text)
            # 動詞テ形（連用形＋て）: 返して/受けて/分岐して/連動して etc.
            text = _re.sub(r'^[ぁ-ん一-龯]{1,5}?て(?=[ぁ-んァ-ヶ一-龯])', '', text)
            text = _re.sub(r'^(?:て|に|で|と|の|は|が|を|も|や|ず)+', '', text)
            text = text.strip()
        return text

    def _trim_trailing_noise(text: str) -> str:
        """末尾の非日本語ノイズ（"/", スペース等）を除去する。"""
        return _re.sub(r'[^ぁ-んァ-ヶ一-龯々ー]+$', '', text)

    def _refine_prefix(text: str, target_len: int = 10) -> str:
        """プリフィクスを粒度よく整形する。助詞境界で分割し、最右の意味のある名詞句を返す。"""
        text = _strip_leading_connectives(text)
        text = _trim_trailing_noise(text)
        # 助詞境界で分割し、最右のセグメントを優先
        segs = _re.split(r'(?:を|は|が|に|で|と|から|まで|より)', text)
        segs = [_trim_trailing_noise(_strip_leading_connectives(s)) for s in segs if s.strip()]
        for seg in reversed(segs):
            if 2 <= len(seg) <= target_len + 4 and _re.search(r'[一-龯ぁ-んァ-ヶ]', seg):
                return seg
        # 末尾の漢字クラスター
        kanji_clusters = _re.findall(r'[一-龯]{2,}', text)
        if kanji_clusters:
            return kanji_clusters[-1]
        # カタカナクラスター
        kana_clusters = _re.findall(r'[ァ-ヶー]{2,}', text)
        if kana_clusters:
            return kana_clusters[-1]
        return text[:target_len]

    def _extract_jp_noun_backward(text_before: str) -> str:
        """キーワード直前から遡り、スクリプト遷移／区切り文字で停止して名詞句を抽出する。"""
        pos = len(text_before)
        prev_class = None
        stop_pos = 0
        while pos > 0:
            ch = text_before[pos - 1]
            if ch in _BOUNDARIES:
                stop_pos = pos
                break
            cls = _script_class(ch)
            if prev_class == 'jp' and cls != 'jp':
                stop_pos = pos
                break
            prev_class = cls
            pos -= 1
        text = text_before[stop_pos:].strip()
        return _refine_prefix(text)

    def _build(prefix: str, suffix: str, kw: str, clean: str = "") -> str:
        # 日本語を含まないプリフィクスは棄却（"/" のようなノイズ混入を防ぐ）
        if prefix and not _re.search(r'[ぁ-んァ-ヶ一-龯]', prefix):
            prefix = ""
        # prefix が空または貧弱なら、より広い範囲から拾う
        if (not prefix or len(prefix) < 2) and clean:
            m = _re.search(r'([ぁ-んァ-ヶ一-龯々ー]{2,10})を[^。\n]{0,40}?' + kw, clean)
            if m:
                candidate = _strip_leading_connectives(m.group(1))
                if candidate and len(candidate) >= 2:
                    prefix = candidate
        if prefix and prefix.endswith('を') and suffix.startswith('を'):
            prefix = prefix[:-1]
        if prefix:
            t = prefix + suffix
        else:
            t = kw + 'する'
        return _re.sub(r'(を|の){2,}', r'\1', t)

    # --- Phase 0: `(起動|実行|表示|遷移)する{名詞}` 複合パターン ---
    m = _re.search(
        r'(起動|実行|表示|遷移)する([ぁ-んァ-ヶ一-龯々ー]{2,10})(?=[。、・「」（）\s]|$)',
        clean
    )
    if m:
        verb = m.group(1)
        noun = _strip_leading_connectives(m.group(2))
        if noun:
            return (noun + 'を' + verb + 'する')[:max_len]

    # --- Phase 1: 厳格な `を[キーワード]` パターン（最後の出現を優先） ---
    strict_results: list[str] = []
    for kw, suffix in _KW_PAIRS:
        marker = 'を' + kw
        idx = clean.rfind(marker)
        if idx == -1:
            continue
        prefix = _extract_jp_noun_backward(clean[:idx])
        title = _build(prefix, suffix, kw, clean)
        strict_results.append(title)

    if strict_results:
        valid = [t for t in strict_results if _re.match(r'^[ぁ-んァ-ヶ一-龯]', t)]
        prefixed = [t for t in valid if len(t) >= 6]  # 名詞プリフィクス付きを優先
        pool = prefixed or valid or strict_results
        return min(pool, key=len)[:max_len]

    # --- Phase 2: 緩い `[キーワード]` 単独パターン ---
    lax_results: list[str] = []
    for kw, suffix in _KW_PAIRS:
        idx = clean.rfind(kw)
        if idx == -1:
            continue
        # 列挙文脈（"/", "・", "、"）直後のキーワードはスキップ（意味的に動詞でない可能性が高い）
        if idx > 0 and clean[idx - 1] in '/／・、,':
            continue
        prefix = _extract_jp_noun_backward(clean[:idx])
        title = _build(prefix, suffix, kw, clean)
        lax_results.append(title)

    if lax_results:
        valid = [t for t in lax_results if _re.match(r'^[ぁ-んァ-ヶ一-龯]', t)]
        prefixed = [t for t in valid if len(t) >= 6]
        pool = prefixed or valid or lax_results
        return min(pool, key=len)[:max_len]

    # --- Phase 3: フォールバック ---
    # 「〜を担当する」形式 → 先頭項目のみ
    m_tantou = _re.match(r'^(.+?)を担当する', clean)
    if m_tantou:
        items = [x.strip() for x in m_tantou.group(1).split('・') if x.strip()]
        if items:
            return (items[0] + 'を行う')[:max_len]

    # 先頭文の日本語名詞句を抽出
    first_sent = _re.split(r'[。\n]', clean)[0].strip()
    jp_tokens = [t for t in _re.findall(r'[ぁ-んァ-ヶ一-龯々ー]+', first_sent) if len(t) >= 2]
    if jp_tokens:
        base = jp_tokens[0]
        if len(base) > max_len - 3:
            base = base[:max_len - 3]
        return (base + 'を行う')[:max_len]

    # 最後の保険: 18 文字で丸める
    return clean[:max_len]


def _extract_actor(token: str) -> str:
    """フロートークンから日本語アクター名を推定する。"""
    t = _re.sub(r'（[^）]*）|\([^)]*\)', '', token).strip()
    if _re.search(r'お客様|顧客|申請者|依頼者', t):
        return "お客様"
    if _re.search(r'管理者|事務|担当者|スタッフ|GF社', t):
        return "GF社担当者"
    if _re.search(r'Flow|フロー|承認', t):
        return "自動フロー"
    if _re.search(r'画面|フォーム|入力|ページ', t):
        return "お客様"
    if _re.search(r'[A-Z][a-zA-Z]', t):
        return "システム"
    return t[:20] if t else "システム"


def _infer_trigger_screen(data: dict) -> str:
    """起点画面を推定する: screens[] → LWC → テキストキーワード の順に判定。"""
    # 1. screens[] の screen_name を使用（最優先）
    for s in data.get("screens", []):
        name = s.get("screen_name", "") or s.get("component", "")
        if name:
            return name

    # 2. LWC コンポーネントが存在する場合
    lwcs = [c.get("api_name", "") for c in data.get("components", []) if c.get("type") == "LWC"]
    if lwcs:
        return " / ".join(lwcs) + "（Lightningコンポーネント画面）"

    # 3. テキストキーワードから推定
    combined = " ".join([data.get("processing_purpose", ""), data.get("data_flow_overview", "")])
    first_token = data.get("data_flow_overview", "").split("→")[0].strip()

    if _re.search(r'Visualforce|VFページ|\bVF\b', combined):
        return "Visualforceページ（フォーム画面）"
    if _re.search(r'Experience Cloud|Experienceポータル|ポータル画面', combined):
        return "Experience Cloudポータル画面"
    if _re.search(r'Flow|フロー', first_token) and _re.search(r'管理者', first_token):
        return "Salesforce管理画面（またはFlowアクション）"
    if _re.search(r'管理者', first_token):
        return "Salesforce管理画面"
    if _re.search(r'Flow|フロー', first_token):
        return "Salesforce Flow（ボタンアクション）"
    if _re.search(r'お客様|顧客', first_token):
        return "Experience Cloudポータル画面"

    return "Salesforce管理画面"


def _infer_callees(data: dict) -> None:
    """data_flow_overview の「→」連鎖からコンポーネント間呼び出し関係を推論して callees に設定する。"""
    flow_text = data.get("data_flow_overview", "")
    if not flow_text:
        return
    comp_names = {c.get("api_name", "") for c in data.get("components", [])}
    if not comp_names:
        return

    # 全文を処理（。で区切らず全体から → を抽出）
    tokens = _re.split(r'→', flow_text.replace("\n", " "))

    def find_comp(token: str) -> str | None:
        for name in comp_names:
            if name and name in token:
                return name
        return None

    callees_map: dict[str, list[str]] = {n: [] for n in comp_names}
    prev = None
    for token in tokens:
        curr = find_comp(token)
        if curr and prev and prev != curr and curr not in callees_map[prev]:
            callees_map[prev].append(curr)
        if curr:
            prev = curr

    for comp in data.get("components", []):
        name = comp.get("api_name", "")
        if callees_map.get(name):
            comp["callees"] = callees_map[name]


def _build_business_flow(data: dict) -> list[dict]:
    """processing_purpose/screens/data_flow_overview から業務レベルフローを生成する。

    ルール:
    - 「誰が・何を起点に → 何が起きる → 誰が何を受け取る」を3〜4ステップで表現
    - クラス名・API名・アノテーションは一切使わない
    - 処理概要シートとの重複を避けるため、システム内部処理は書かない
    """
    purpose  = data.get("processing_purpose", "")
    flow_ov  = data.get("data_flow_overview", "")
    notes    = data.get("notes", "")
    screens  = data.get("screens", [])
    combined = purpose + " " + flow_ov + " " + notes

    steps    = []
    step_no  = 1

    # ─── Step 1: 起点アクター＋起動アクション ─────────────────────────────
    first_tok = (flow_ov.split("→")[0] if flow_ov else "").strip()

    # 技術ワードのみのトークン（例「画面フロー（Create_CustomerUser）」「Apexクラス」等）は
    # 業務アクションとして読めないため、適切なデフォルトに差し替える。
    first_tok_core = _re.sub(r'[（(].*?[）)]', '', first_tok).strip()
    _TECH_ONLY = ("画面フロー", "フロー", "Apex", "Apexクラス", "Batch",
                  "Flow", "Trigger", "LWC", "Aura", "Visualforce",
                  "VF", "Integration", "Controller", "Service", "Handler")
    is_tech_only = first_tok_core in _TECH_ONLY
    is_screen_flow_start = "画面フロー" in first_tok

    if screens:
        # 画面あり → お客様が画面から入力・送信
        scr = screens[0].get("screen_name") or screens[0].get("component", "画面")
        scr_clean = _clean_tech_business(scr)
        action1 = f"{scr_clean}から必要情報を入力し、送信する"
        steps.append({"step": step_no, "actor": "お客様", "action": action1, "next": []})
        step_no += 1
    elif is_screen_flow_start:
        # data_flow_overview が「画面フロー（...）→ ...」で始まる場合は画面起点。
        # screens が空でも Flow タイプの画面フローが存在する想定。
        # 管理系業務（ユーザー発行・管理者操作）は GF社担当者、それ以外はお客様。
        admin_hint = _re.search(r'管理者|発行|管理.*ユーザー|GF社担当|事務', purpose + notes)
        actor = "GF社担当者" if admin_hint else "お客様"
        steps.append({"step": step_no, "actor": actor,
                      "action": "画面から必要情報を入力し、処理を起動する", "next": []})
        step_no += 1
    elif is_tech_only:
        # 技術ワードのみ（画面フロー以外）→ 自動起点とみなす
        steps.append({"step": step_no, "actor": "自動フロー",
                      "action": "処理を起動する", "next": []})
        step_no += 1
    elif _re.search(r'管理者|GF社|担当者|事務', first_tok):
        # GF社担当者が直接操作して起動
        clean = _clean_tech_business(first_tok)
        steps.append({"step": step_no, "actor": "GF社担当者",
                      "action": clean or "処理を起動する", "next": []})
        step_no += 1
    elif _re.search(r'承認.*[Ff]low|承認.*フロー|承認フロー', first_tok):
        # 承認フロー起点 → 「誰かが何かを承認した後に動く」→ GF社担当者が承認
        m_ctx = _re.search(r'(.{2,25}(?:確定|承認|依頼|完了))後', purpose)
        ctx = _clean_tech_business(m_ctx.group(1)) if m_ctx else "業務処理"
        steps.append({"step": step_no, "actor": "GF社担当者",
                      "action": f"{ctx}後、処理を承認・起動する", "next": []})
        step_no += 1
    elif _re.search(r'お客様|顧客[^向]|Experience Cloud.*ポータル|ポータル.*画面', first_tok):
        clean = _clean_tech_business(first_tok)
        steps.append({"step": step_no, "actor": "お客様",
                      "action": clean or "操作を行う", "next": []})
        step_no += 1
    else:
        clean = _clean_tech_business(first_tok)
        steps.append({"step": step_no, "actor": "GF社担当者",
                      "action": clean or "処理を起動する", "next": []})
        step_no += 1

    # ─── Step 2: メイン業務処理（processing_purpose の主要部分） ──────────────
    sents = [s.strip() for s in purpose.split("。") if s.strip()]
    if sents:
        first = sents[0]
        # 「〜後の」前置き除去（「コンサルティング本確定後の」等）
        core = _re.sub(r'^.{0,35}後の', '', first).strip()
        main = _clean_tech_business(core or first)
        main = _PREAMBLE_RE.sub('', main).strip()
        # "〜を担う" 等の文末表現を除去（行動として読めるように）
        main = _re.sub(r'を担う$|を行う$|を実施する$', '', main).strip()
        main = _re.sub(r'^[、。・/\s]+|[、。・/\s]+$', '', main).strip()
        if main and len(main) > 5:
            steps.append({"step": step_no, "actor": "自動フロー",
                          "action": main, "next": []})
            step_no += 1

    # ─── Step 3: 顧客向け結果（キーワードで判定） ─────────────────────────────
    outcome = ""
    if _re.search(r'初期パスワード|パスワード.{0,5}メール', combined):
        outcome = "初期パスワードメールを受信し、Experience Cloudポータルにアクセスする"
    elif _re.search(r'thankPage|thanksPage|thank.*page', combined, _re.IGNORECASE):
        outcome = "フォーム送信後、受付完了ページで確認する"
    elif _re.search(r'顧客向け通知|通知.*メール|ダウンロード.*通知|完了通知', combined):
        outcome = "完了通知メールを受信し、書類を確認する"
    elif _re.search(r'ダウンロード', combined) and _re.search(r'Experience Cloud|ポータル', combined):
        outcome = "Experience Cloudポータルで書類をダウンロードする"
    elif _re.search(r'メール.*送信|自動メール|顧客宛', combined) and not screens:
        outcome = "自動送信されたメールを受信する"

    if outcome:
        steps.append({"step": step_no, "actor": "お客様",
                      "action": outcome, "next": []})
        step_no += 1

    # ─── next リンク設定 ───────────────────────────────────────────────
    for i in range(len(steps) - 1):
        steps[i]["next"] = [{"to": steps[i + 1]["step"]}]

    return steps


def _comp_type_label(comp: dict) -> str:
    """コンポーネントの種別ラベルを日本語で返す（クラス名は出さない）。"""
    api   = comp.get("api_name", "")
    ctype = comp.get("type", "Apex")
    if ctype == "Flow":
        return "フロー"
    if ctype == "LWC":
        return "LWCコンポーネント"
    if ctype == "Integration":
        return "外部サービス連携"
    if ctype == "Trigger" or ("Trigger" in api and "Handler" in api):
        return "Apexトリガーハンドラー"
    if "Trigger" in api:
        return "Apexトリガー"
    return "Apexクラス"


def _build_process_steps(data: dict) -> list[dict]:
    """components の responsibility から日本語の処理概要 steps を生成する。
    API名・クラス名・英語技術用語はすべて日本語に変換する。
    コンポーネント列は種別ラベル（Apexクラス／フロー等）を表示し、クラス名は出さない。
    """
    steps = []
    n_comps = len(data.get("components", []))
    for i, comp in enumerate(data.get("components", []), 1):
        responsibility = comp.get("responsibility", "")
        comp_type = comp.get("type", "Apex")

        # 責務を日本語化（SF オブジェクト名・フィールド名・ジャーゴン→日本語 → Apex用クリーン）
        resp_j = _translate_sf_fields(_translate_jargon(_translate_sf_objects(responsibility)))
        desc_main = _clean_tech(resp_j)
        # トリガータイミング注釈（処理前（新規）/処理後（新規））等を除去
        desc_main = _re.sub(
            r'[（(](?:処理前|処理後)（[^）]+）(?:/(?:処理前|処理後)（[^）]+）)*[）)]', '', desc_main
        ).strip()
        desc_main = _re.sub(r'\s{2,}', ' ', desc_main)

        title  = _short_title(responsibility) if responsibility else ""
        branch = None  # 実際の条件内容が取れない場合は空にする

        display_desc = desc_main or title

        _raw_api = comp.get("api_name", "")
        comp_name = _re.sub(r'（[^）]+）$', '', _raw_api).strip() or _raw_api

        steps.append({
            "step": i,
            "title": title,
            "description": display_desc,
            "component": comp_name,
            "comp_api_name": _raw_api,
            "branch": branch,
            "next": [{"to": i + 1}] if i < n_comps else [],
        })

    return steps


def _obj_label_from_api(api: str) -> str:
    """オブジェクトAPIから日本語ラベルを推定する（メタデータ優先）。"""
    if api in _STD_OBJ_LABELS:
        return _STD_OBJ_LABELS[api]
    if api in _SF_OBJ_LABELS:
        return _SF_OBJ_LABELS[api]
    raw = api.replace("__c", "").replace("__", "")
    base = _re.sub(r'([A-Z])', r' \1', raw).strip()
    return base


def _build_related_objects_and_access(data: dict) -> tuple[list[dict], list[dict]]:
    """components の inputs/outputs から related_objects と object_access を構築する。

    Returns: (related_objects, object_access)
    - related_objects: [{api_name, label, fields, relations}]
    - object_access:   [{component, object, operation}]
    """
    obj_comp_ops: dict[str, dict[str, str]] = {}   # obj_api → {comp_name → operation}
    obj_fields: dict[str, list[dict]] = {}          # obj_api → [field_dicts]

    def _register(obj_api: str, comp_name: str, op: str):
        obj_comp_ops.setdefault(obj_api, {})
        existing = obj_comp_ops[obj_api].get(comp_name, "")
        if not existing:
            obj_comp_ops[obj_api][comp_name] = op
        elif existing == "R" and op in ("W", "INSERT"):
            obj_comp_ops[obj_api][comp_name] = "RW"

    for comp in data.get("components", []):
        name = comp.get("api_name", "")

        # inputs → R（参照）。ただし trigger newList / trigger.new 文脈では INSERT
        for text in [comp.get("inputs", "")]:
            if not text:
                continue
            is_trigger_new = bool(_re.search(r'trigger\s+new|trigger\.new|newList', text, _re.IGNORECASE))
            for m in _re.finditer(r'(?<![A-Za-z0-9_])([A-Z][A-Za-z0-9]*__c)(?![A-Za-z0-9_])', text):
                _register(m.group(1), name, "INSERT" if is_trigger_new else "R")
            for std_api in _STD_OBJ_LABELS:
                if _re.search(rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])', text):
                    _register(std_api, name, "R")

        # outputs → W または INSERT
        for text in [comp.get("outputs", "")]:
            if not text:
                continue

            # "OBJ更新（FIELD1・FIELD2）" → W + フィールド名抽出
            for m in _re.finditer(r'([A-Z][A-Za-z0-9]*__c)(?:[^\n（]*?)更新[^\n（]*?(?:（([^）]+)）)?', text):
                obj_api = m.group(1)
                _register(obj_api, name, "W")
                if m.group(2):
                    fnames = _re.findall(r'([A-Za-z][A-Za-z0-9]*__c)', m.group(2))
                    obj_fields.setdefault(obj_api, [])
                    for fn in fnames:
                        if not any(f["api_name"] == fn for f in obj_fields[obj_api]):
                            obj_fields[obj_api].append({
                                "api_name": fn,
                                "label": _sf_field_label(obj_api, fn),
                                "access": "W", "note": "",
                            })

            # "OBJ（insert）" や "insert" を含む → INSERT
            if _re.search(r'(?i)\binsert\b|新規作成|新規登録', text):
                for m in _re.finditer(r'(?<![A-Za-z0-9_])([A-Z][A-Za-z0-9]*__c)(?![A-Za-z0-9_])', text):
                    _register(m.group(1), name, "INSERT")
                for std_api in _STD_OBJ_LABELS:
                    if _re.search(rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])', text):
                        _register(std_api, name, "INSERT")
            else:
                for m in _re.finditer(r'(?<![A-Za-z0-9_])([A-Z][A-Za-z0-9]*__c)(?![A-Za-z0-9_])', text):
                    _register(m.group(1), name, "W")
                for std_api in _STD_OBJ_LABELS:
                    if _re.search(rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])', text):
                        _register(std_api, name, "W")

    # data_flow_overview から「ObjName（Field__c等フラグ更新）」形式のフィールド情報を抽出
    dfo = data.get("data_flow_overview", "")
    if dfo:
        # 標準オブジェクト: "Contact（IsConsignee__c等フラグ更新）" → Contact.IsConsignee__c = W
        for std_api in _STD_OBJ_LABELS:
            for m in _re.finditer(rf'{std_api}（([^）]{{1,120}})）', dfo):
                parens = m.group(1)
                fnames = _re.findall(r'([A-Za-z][A-Za-z0-9]*__c)', parens)
                is_w = bool(_re.search(r'更新|フラグ|設定|保存', parens))
                op = "W" if is_w else "R"
                for fn in fnames:
                    label = _sf_field_label(std_api, fn)
                    obj_fields.setdefault(std_api, [])
                    if not any(f["api_name"] == fn for f in obj_fields[std_api]):
                        obj_fields[std_api].append(
                            {"api_name": fn, "label": label, "access": op, "note": ""})
        # カスタムオブジェクト: "Quote__c更新（QuoteLinkId__c・...）" → 既存パターンと同様
        for m in _re.finditer(r'([A-Z][A-Za-z0-9]*__c)（([^）]+)）', dfo):
            obj_api = m.group(1)
            parens = m.group(2)
            fnames = _re.findall(r'([A-Za-z][A-Za-z0-9]*__c)', parens)
            is_w = bool(_re.search(r'更新|フラグ|設定|保存', parens))
            op = "W" if is_w else "R"
            for fn in fnames:
                label = _sf_field_label(obj_api, fn)
                obj_fields.setdefault(obj_api, [])
                if not any(f["api_name"] == fn for f in obj_fields[obj_api]):
                    obj_fields[obj_api].append(
                        {"api_name": fn, "label": label, "access": op, "note": ""})

    # 他テキストから追加オブジェクトを収集（標準オブジェクトのみ）
    # ※ カスタムオブジェクトは inputs/outputs から取得済み。__c のスキャンは
    #   フィールド名（IsConsignee__c等）を誤ってオブジェクトとして追加する恐れがあるため除外する
    purpose_text = data.get("processing_purpose", "")
    combined_extra = " ".join([dfo, purpose_text])
    comp_names_real = [c.get("api_name", "") for c in data.get("components", []) if c.get("api_name")]

    for std_api in _STD_OBJ_LABELS:
        pat = rf'(?<![A-Za-z0-9_]){_re.escape(std_api)}(?![A-Za-z0-9_])'
        if not _re.search(pat, combined_extra):
            continue
        if std_api in obj_comp_ops:
            continue

        # 操作種別を推定（INSERT > W > R の優先順位）
        _insert_ctx = rf'(?:{_re.escape(std_api)}.{{0,20}}(?:作成|登録|新規)|(?:作成|登録|新規).{{0,20}}{_re.escape(std_api)})'
        _update_ctx = rf'(?:{_re.escape(std_api)}.{{0,20}}(?:更新|変更|保存)|(?:更新|変更|保存).{{0,20}}{_re.escape(std_api)})'
        if _re.search(_insert_ctx, combined_extra):
            inferred_op = "INSERT"
        elif _re.search(_update_ctx, combined_extra):
            inferred_op = "W"
        else:
            inferred_op = "R"

        # 前後200文字以内に実コンポーネント名があれば紐付ける
        linked_comp = None
        for m in _re.finditer(pat, combined_extra):
            window = combined_extra[max(0, m.start() - 200): m.end() + 200]
            for cname in comp_names_real:
                if cname in window:
                    linked_comp = cname
                    break
            if linked_comp:
                break

        # 紐付け先が見つかればそのコンポーネントへ。なければテキスト検出として仮登録
        assigned_comp = linked_comp if linked_comp else "（テキスト検出）"
        obj_comp_ops[std_api] = {assigned_comp: inferred_op}

    # related_objects 構築
    related_objects = []
    for obj_api, comp_ops in obj_comp_ops.items():
        # オブジェクトレベルの合算アクセス種別（マトリクスと一致させる）
        # R/W/INSERT を全て独立に保持し、表示はそれらを「・」連結する
        all_ops = list(comp_ops.values())
        has_read   = any(o in ("R", "RW") for o in all_ops)
        has_write  = any(o in ("W", "RW") for o in all_ops)
        has_insert = any(o == "INSERT" for o in all_ops)
        parts: list[str] = []
        if has_read:   parts.append("参照")
        if has_write:  parts.append("更新")
        if has_insert: parts.append("新規作成")
        obj_combined_ja = "・".join(parts) if parts else ""
        # 後方互換用の単一キー（最も強い操作を代表値として残す）
        if has_insert:
            obj_combined_op = "INSERT"
        elif has_write and has_read:
            obj_combined_op = "RW"
        elif has_write:
            obj_combined_op = "W"
        elif has_read:
            obj_combined_op = "R"
        else:
            obj_combined_op = all_ops[0] if all_ops else ""

        fields = obj_fields.get(obj_api, [])

        if not fields:
            # Flow/Apexメタデータからフィールドを補完
            meta = _SF_FIELD_LABELS.get(obj_api, {})
            seen_apis: set[str] = set()
            for comp in data.get("components", []):
                # JSONのapi_nameから型接尾辞（（Flow）等）を除去してコンポーネントキーに変換
                raw_api = comp.get("api_name", "")
                comp_key = _re.sub(r'（[^）]+）$', '', raw_api).strip()
                comp_field_map = _SF_COMP_FIELDS.get(comp_key, {})
                # 対象オブジェクトに直接マッチするフィールド
                for fapi in comp_field_map.get(obj_api, set()):
                    if fapi in seen_apis:
                        continue
                    seen_apis.add(fapi)
                    label = meta.get(fapi) or _sf_field_label(obj_api, fapi)
                    fields.append({"api_name": fapi, "label": label, "access": obj_combined_op, "note": ""})

        if not fields:
            # 最終フォールバック
            if has_insert and not has_read:
                label = "（レコード新規登録）"
            else:
                label = "（対象項目は別途設計書を参照）"
            fields = [{"api_name": "—", "label": label, "access": obj_combined_op, "note": ""}]
        else:
            # 既存フィールドの access もオブジェクト合算値に統一
            for f in fields:
                f["access"] = obj_combined_op

        related_objects.append({
            "api_name": obj_api,
            "label": _obj_label_from_api(obj_api),
            "fields": fields,
            "relations": [],
            "access_ja": obj_combined_ja,
        })

    # object_access 構築（テキスト検出の仮コンポーネントは除外）
    object_access = []
    for obj_api, comp_ops in obj_comp_ops.items():
        for comp_name, op in comp_ops.items():
            if comp_name and op and comp_name != "（テキスト検出）":
                object_access.append({"component": comp_name, "object": obj_api, "operation": op})

    return related_objects, object_access


def _enrich_related_object_fields(data: dict) -> None:
    """既存の related_objects + object_access を保持したまま、fields のみ補完する。

    キャッシュされた related_objects がオブジェクト構成は正しいが fields が空のとき、
    object_access と SF メタデータ（_SF_COMP_FIELDS / _SF_FIELD_LABELS）から
    コンポーネント単位で関連フィールドを抽出して埋める。
    """
    oa = data.get("object_access") or []
    ro = data.get("related_objects") or []
    if not oa or not ro:
        return

    # obj_api → 合算アクセス種別（R/W/INSERT/RW）を再計算
    obj_ops: dict[str, dict[str, bool]] = {}
    obj_comps: dict[str, set[str]] = {}
    for entry in oa:
        obj_api = entry.get("object", "")
        comp = entry.get("component", "")
        op = (entry.get("operation", "") or "").upper()
        if not obj_api or not op:
            continue
        flags = obj_ops.setdefault(obj_api, {"R": False, "W": False, "INSERT": False})
        if op in ("R", "RW"):
            flags["R"] = True
        if op in ("W", "RW"):
            flags["W"] = True
        if op == "INSERT":
            flags["INSERT"] = True
        if comp:
            obj_comps.setdefault(obj_api, set()).add(comp)

    def _combined(flags: dict) -> tuple[str, str]:
        parts: list[str] = []
        if flags.get("R"):      parts.append("参照")
        if flags.get("W"):      parts.append("更新")
        if flags.get("INSERT"): parts.append("新規作成")
        ja = "・".join(parts) if parts else ""
        if flags.get("INSERT"):
            key = "INSERT"
        elif flags.get("W") and flags.get("R"):
            key = "RW"
        elif flags.get("W"):
            key = "W"
        elif flags.get("R"):
            key = "R"
        else:
            key = ""
        return key, ja

    for obj in ro:
        if obj.get("fields"):
            continue
        obj_api = obj.get("api_name", "")
        if not obj_api:
            continue
        flags = obj_ops.get(obj_api, {})
        op_key, op_ja = _combined(flags)

        # SF メタデータからフィールドを収集
        fields: list[dict] = []
        seen: set[str] = set()
        for comp_name in obj_comps.get(obj_api, set()):
            comp_field_map = _SF_COMP_FIELDS.get(comp_name, {})
            for fapi in comp_field_map.get(obj_api, set()):
                if fapi in seen:
                    continue
                seen.add(fapi)
                label = _sf_field_label(obj_api, fapi)
                fields.append({"api_name": fapi, "label": label, "access": op_key, "note": ""})

        if not fields:
            # メタデータからも取れなければ代表ラベルで1行入れる
            label = "（レコード新規登録）" if flags.get("INSERT") and not flags.get("R") \
                    else "（対象項目は別途設計書を参照）"
            fields = [{"api_name": "—", "label": label, "access": op_key, "note": ""}]

        obj["fields"] = fields
        if op_ja and not obj.get("access_ja"):
            obj["access_ja"] = op_ja


def _infer_users(data: dict) -> str:
    """processing_purpose / data_flow_overview / components からユーザー/利用部門を推定する。"""
    comp_resp = " ".join(
        c.get("responsibility", "") for c in data.get("components", [])
    )
    combined = " ".join([
        data.get("processing_purpose", ""),
        data.get("data_flow_overview", ""),
        comp_resp,
    ])
    parts = []
    if _re.search(r'お客様|顧客|申請者|Experience Cloud|見込み客|問い合わせ者|HP問い合わせ|WebTo', combined):
        if _re.search(r'Experience Cloud|申請者', combined):
            parts.append("お客様（Experience Cloudユーザー）")
        else:
            parts.append("見込み客・問い合わせ者")
    if _re.search(r'管理者|GF社|担当者|事務|社内|内部', combined):
        parts.append("GF社担当者")
    if _re.search(r'コンサル', combined):
        parts.append("GF社コンサル部")
    if _re.search(r'営業', combined):
        parts.append("GF社営業部")
    if not parts:
        parts.append("GF社担当者")
    return "・".join(parts)


def _normalize_schema(data: dict) -> dict:
    """GFプロジェクト固有スキーマを標準スキーマに変換する。全テキストを日本語化する。"""
    # feature_id
    if not data.get("feature_id") and data.get("group_id"):
        data["feature_id"] = data["group_id"]

    # SFプロジェクトメタデータの自動ロード（project_name から判定、大小無視）
    proj = (data.get("project_name", "") or "").strip()
    if proj and not _SF_FIELD_LABELS:
        sf_path = _SF_PROJECT_PATHS.get(proj.lower())
        if sf_path:
            _load_sf_metadata(sf_path)

    # ── 概要フィールド（すべてAPI名・クラス名なしの日本語で） ────────────
    purpose_raw = data.get("processing_purpose", "")

    if not data.get("summary") and purpose_raw:
        # 最初の文のみ使用（2文目以降は技術実装詳細が多い）
        first_sent = purpose_raw.split("。")[0]
        data["summary"] = _clean_tech_business(first_sent)

    if not data.get("purpose") and purpose_raw:
        # 全文をクリーンして目的とする（data_flow_overview は技術的すぎるので使わない）
        cleaned_purpose = _clean_tech_business(purpose_raw)
        if data.get("notes"):
            cleaned_notes = _clean_tech_business(data["notes"])
            if cleaned_notes:
                cleaned_purpose += f"\n【前提・補足】{cleaned_notes}"
        data["purpose"] = cleaned_purpose

    # 利用者推定
    if not data.get("users"):
        data["users"] = _infer_users(data)

    # 起点画面（screens[] → LWC → キーワード 推定）
    if not data.get("trigger_screen"):
        data["trigger_screen"] = _infer_trigger_screen(data)

    # 操作トリガー: prerequisites を日本語化してセット
    if not data.get("trigger") and data.get("prerequisites"):
        data["trigger"] = _clean_tech_business(data["prerequisites"])

    # components: responsibility → role（日本語化）& callees 初期化
    for comp in data.get("components", []):
        if not comp.get("role"):
            comp["role"] = _clean_tech(comp.get("responsibility", ""))
        if "callees" not in comp:
            comp["callees"] = []

    # callees を data_flow_overview から推論
    if not any(comp.get("callees") for comp in data.get("components", [])):
        _infer_callees(data)

    # process_steps: components から日本語説明で生成
    if not data.get("process_steps"):
        data["process_steps"] = _build_process_steps(data)

    # business_flow: 業務レベルフロー生成
    _BF_TECH_ONLY = ("画面フロー", "フロー", "Apex", "Apexクラス", "Batch",
                     "Flow", "Trigger", "LWC", "Aura", "Visualforce",
                     "VF", "Integration", "Controller", "Service", "Handler")
    _bf = data.get("business_flow") or []
    if not _bf:
        data["business_flow"] = _build_business_flow(data)
    else:
        # キャッシュされた業務フローに技術ワードのみの action（"画面フロー" 等）が含まれる場合は
        # 該当ステップの action だけを業務的な表現に差し替える（他のステップは保持する）。
        for s in _bf:
            act = str(s.get("action", "")).strip()
            if act in _BF_TECH_ONLY:
                actor = str(s.get("actor", "")).strip()
                if "画面フロー" in act or act == "画面フロー":
                    s["action"] = "画面から必要情報を入力し、処理を起動する"
                elif actor in ("自動フロー", "システム"):
                    s["action"] = "処理を起動する"
                else:
                    s["action"] = "画面から必要情報を入力し、処理を起動する"

    # related_objects + object_access: components の inputs/outputs から構築
    _ro = data.get("related_objects") or []
    _oa = data.get("object_access") or []
    _ro_all_empty = bool(_ro) and all(not (o.get("fields") or []) for o in _ro)
    if not _ro or not _oa:
        rel_objs, obj_access = _build_related_objects_and_access(data)
        if not _ro:
            data["related_objects"] = rel_objs
        if not _oa:
            data["object_access"] = obj_access
    elif _ro_all_empty:
        # 既存 related_objects と object_access は保持し、fields のみ SF メタデータから補完する。
        _enrich_related_object_fields(data)

    return data


# ── 差分計算 ────────────────────────────────────────────────────────
def _compute_diffs(prev_data: dict | None, new_data: dict) -> dict:
    if prev_data is None:
        return {"scalars": [], "lists": {}}
    return {
        "scalars": dr.diff_scalars(prev_data, new_data, SCALAR_FIELDS),
        "lists": {
            "business_flow": dr.diff_list(
                prev_data.get("business_flow", []),
                new_data.get("business_flow", []), "step"),
            "related_objects": dr.diff_list(
                prev_data.get("related_objects", []),
                new_data.get("related_objects", []), "api_name"),
            "process_steps": dr.diff_list(
                prev_data.get("process_steps", []),
                new_data.get("process_steps", []), "step"),
            "components": dr.diff_list(
                prev_data.get("components", []),
                new_data.get("components", []), "api_name"),
        },
    }


# ── PNG生成 ────────────────────────────────────────────────────────
def _generate_diagrams(data: dict, tmp_dir: str) -> dict[str, str | None]:
    """4種の図形PNGを生成し、パスを返す。失敗したらNone。"""
    import diagram_gen as dg

    paths: dict[str, str | None] = {
        "swimlane":  None,
        "er":        None,
        "flowchart": None,
        "component": None,
    }

    # 1. スイムレーン図（業務フロー）
    flows = data.get("business_flow", [])
    if flows:
        try:
            sl_path = str(Path(tmp_dir) / "swimlane.png")
            dg.render_swimlane(_business_flow_to_swimlane(flows), sl_path)
            paths["swimlane"] = sl_path
            print("  [OK] スイムレーン図")
        except Exception as e:
            print(f"  [WARN] スイムレーン図: {e}")

    # 2. オブジェクト参照マトリクス（対象オブジェクト）
    objects = data.get("related_objects", [])
    object_access = data.get("object_access", [])
    components = data.get("components", [])
    if objects:
        try:
            er_path = str(Path(tmp_dir) / "er.png")
            if object_access:
                from diagram_utils import generate_object_component_matrix
                # テキスト検出のみのオブジェクトはマトリクスから除外
                oa_obj_apis = {oa["object"] for oa in object_access}
                matrix_objects = [o for o in objects if o["api_name"] in oa_obj_apis]
                generate_object_component_matrix(
                    object_access, components, matrix_objects, er_path)
            else:
                # object_access がない場合は従来のER図にフォールバック
                from er_utils import generate_er_image
                boxes, arrows = _related_objects_to_er_boxes(objects)
                n_obj = len(boxes)
                er_w = min(14, max(8, n_obj * 3.0))
                er_h = min(10, max(6, n_obj * 2.0))
                generate_er_image(boxes, arrows, er_path,
                                  title="オブジェクト関連図",
                                  slide_w=er_w, slide_h=er_h)
            paths["er"] = er_path
            print("  [OK] オブジェクト参照マトリクス")
        except Exception as e:
            print(f"  [WARN] オブジェクト参照マトリクス: {e}")

    # 3. フローチャート（処理概要）
    steps = data.get("process_steps", [])
    if steps:
        try:
            fc_path = str(Path(tmp_dir) / "flowchart.png")
            dg.render_flowchart(steps, fc_path)
            paths["flowchart"] = fc_path
            print("  [OK] フローチャート")
        except Exception as e:
            print(f"  [WARN] フローチャート: {e}")

    # 4. コンポーネント図（関連コンポーネント）
    components = data.get("components", [])
    if components:
        try:
            cm_path = str(Path(tmp_dir) / "component.png")
            dg.render_component_diagram(components, cm_path, steps=steps)
            paths["component"] = cm_path
            print("  [OK] コンポーネント図")
        except Exception as e:
            print(f"  [WARN] コンポーネント図: {e}")

    return paths


def _business_flow_to_swimlane(flows: list[dict]) -> dict:
    """business_flow リストを render_swimlane 用のフロー dict に変換する。"""
    lane_names = list(dict.fromkeys(f.get("actor", "不明") for f in flows))
    steps = [
        {"id": str(f.get("step", i + 1)), "lane": f.get("actor", "不明"),
         "label": f.get("action", "")}
        for i, f in enumerate(flows)
    ]
    transitions = []
    for i, f in enumerate(flows):
        nexts = f.get("next", [])
        src = str(f.get("step", i + 1))
        if nexts:
            for n in nexts:
                dst_step = n.get("to") or (flows[i + 1].get("step", i + 2) if i + 1 < len(flows) else None)
                if dst_step:
                    transitions.append({"from": src, "to": str(dst_step),
                                        "condition": n.get("condition", "")})
        elif i + 1 < len(flows):
            transitions.append({"from": src, "to": str(flows[i + 1].get("step", i + 2))})
    return {
        "title": "業務フロー",
        "lanes": [{"name": n} for n in lane_names],
        "steps": steps,
        "transitions": transitions,
    }


def _related_objects_to_er(objects: list[dict]) -> tuple[list, list]:
    """related_objects を render_er_diagram 用の (objects, relations) に変換する。"""
    obj_list = [
        {"api": o.get("api_name", ""), "label": o.get("label", ""),
         "type": "カスタム" if "__c" in o.get("api_name", "") else "標準"}
        for o in objects
    ]
    rels = []
    for o in objects:
        for r in o.get("relations", []):
            rels.append({
                "parent": o.get("api_name", ""),
                "rel":    r.get("type", "lookup"),
                "child":  r.get("to", ""),
                "field":  r.get("field", ""),
            })
    return obj_list, rels


def _related_objects_to_er_boxes(objects: list[dict]) -> tuple[list, list]:
    """related_objects を er_utils.generate_er_image 用の (boxes, arrows) に変換する。

    自動グリッドレイアウト: オブジェクトを横に並べ、収まらなければ折り返す。
    """
    import math

    n = len(objects)
    cols = min(n, 3)
    rows = math.ceil(n / cols)

    # レイアウト定数
    box_w = 3.0
    box_h_base = 1.2
    field_h = 0.32
    x_gap = 1.0
    y_gap = 1.0
    x_start = 1.0
    y_start = 1.5  # タイトルバー分

    boxes = []
    obj_id_map: dict[str, str] = {}  # api_name -> box id

    for i, obj in enumerate(objects):
        api = obj.get("api_name", "")
        label = obj.get("label", "")
        fk_fields = [f for f in obj.get("fields", []) if f.get("access") in ("RW", "W")]
        n_fields = min(len(fk_fields), 4)  # 表示上限4フィールド
        h = box_h_base + n_fields * field_h

        col_idx = i % cols
        row_idx = i // cols
        x = x_start + col_idx * (box_w + x_gap)
        y = y_start + row_idx * (box_h_base + 4 * field_h + y_gap)

        style = "primary" if "__c" in api else "secondary"
        box_id = api.replace("__c", "_c").replace("__", "_")
        obj_id_map[api] = box_id

        box_fields = []
        for fi, f in enumerate(fk_fields[:4]):
            box_fields.append({
                "name": f.get("api_name", ""),
                "label": f.get("label", ""),
                "is_fk": any(
                    r.get("to") == f.get("api_name", "")
                    or r.get("field", "") == f.get("api_name", "")
                    for r in obj.get("relations", [])
                ),
            })

        boxes.append({
            "id": box_id,
            "api": api,
            "label": label,
            "x": x, "y": y, "w": box_w, "h": h,
            "style": style,
            "fields": box_fields,
        })

    arrows = []
    for obj in objects:
        src_api = obj.get("api_name", "")
        src_id = obj_id_map.get(src_api, "")
        for rel in obj.get("relations", []):
            to_api = rel.get("to", "")
            dst_id = obj_id_map.get(to_api, "")
            if not dst_id:
                continue
            rel_type = rel.get("type", "lookup").lower()
            arrows.append({
                "from": src_id,
                "to": dst_id,
                "rel": rel_type,
                "field": rel.get("field", rel.get("label", "")),
            })

    return boxes, arrows


# ── メイン ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="詳細設計書 Excel 生成（新スキーマ対応）")
    parser.add_argument("--input",      required=True, help="詳細設計 JSON ファイルパス")
    parser.add_argument("--template",   required=True, help="詳細設計書テンプレート.xlsx パス")
    parser.add_argument("--output-dir", required=True, help="出力先ディレクトリ")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"])
    parser.add_argument("--source-hash", default="",
                        help="グループ内ソースのSHA256。_meta に保存して次回差分判定に使う")
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    data = _normalize_schema(data)   # GFスキーマ → 標準スキーマ変換
    today  = _date.today().strftime("%Y-%m-%d")
    author = data.get("author", "")

    feature_id = data.get("feature_id", "")
    name_ja    = data.get("name_ja", "機能")
    safe_name  = re.sub(r'[\\/:*?"<>|]', "_", name_ja)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    # 他の設計書（プログラム設計・オブジェクト定義書）と統一: 【{feature_id}】{safe_name}_詳細設計.xlsx
    if feature_id:
        out_path = out_dir / f"【{feature_id}】{safe_name}_詳細設計.xlsx"
    else:
        out_path = out_dir / f"{safe_name}_詳細設計.xlsx"

    # ── バージョン判定 ────────────────────────────────────────────
    # feature_id ベースで既存ファイルを検出（機能名変更に追随できる）
    source_file = ""
    if feature_id:
        existing = sorted(out_dir.glob(f"【{feature_id}】*.xlsx"),
                          key=lambda f: f.stat().st_mtime, reverse=True)
        if existing:
            source_file = str(existing[0])
            print(f"  [AUTO] 既存ファイルを検出: {existing[0].name}")
    if not source_file:
        # 後方互換: 旧命名（FG-prefixなし）も探す
        existing = sorted(out_dir.glob(f"*{safe_name}*詳細設計*.xlsx"),
                          key=lambda f: f.stat().st_mtime, reverse=True)
        if existing:
            source_file = str(existing[0])
            print(f"  [AUTO] 既存ファイル（旧命名）を検出: {existing[0].name}")

    prev_meta   = read_meta(source_file) if source_file else None
    prev_data   = prev_meta.get("data") if prev_meta else None

    # 第1ゲート: ソースハッシュ一致ならLLM呼び出しもExcel再生成もスキップ（プログラム設計と同じ方式）
    if (prev_meta and args.source_hash
            and args.version_increment == "minor"
            and prev_meta.get("source_hash") == args.source_hash):
        print("差分なし: ソースハッシュが既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    if prev_meta:
        current_version = increment_version(
            prev_meta.get("version", "1.0"), args.version_increment)
        history    = prev_meta.get("history", [])
        is_initial = False
        print(f"更新モード: {prev_meta.get('version', '?')} -> {current_version}")
    else:
        current_version = data.get("version") or "1.0"
        history    = []
        is_initial = True
        print(f"新規作成モード: v{current_version}")

    data["version"] = current_version
    if not data.get("date"):
        data["date"] = today

    # 第2ゲート: JSONフィールド差分ゼロならExcel再生成スキップ
    diffs = _compute_diffs(prev_data, data)
    if prev_meta and args.version_increment == "minor" and not dr.has_any_diff(diffs):
        print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    last_no = max((h["項番"] for h in history
                   if isinstance(h.get("項番"), int)), default=0)
    new_entries = dr.build_entries(
        current_version, diffs, author, today,
        start_no=last_no + 1,
        is_major=(args.version_increment == "major"),
        is_initial=is_initial,
        section_sheet_map=SECTION_SHEETS,
        scalar_sheet="概要",
    )
    history = history + new_entries

    changed_scalars    = dr.changed_scalar_fields(diffs)
    changed_flows      = dr.changed_ids(diffs, "business_flow")
    changed_objs       = dr.changed_ids(diffs, "related_objects")
    changed_proc_steps = dr.changed_ids(diffs, "process_steps")
    changed_comps      = dr.changed_ids(diffs, "components")
    is_major           = (args.version_increment == "major")

    # ── 図形PNG生成 ──────────────────────────────────────────────
    with tempfile.TemporaryDirectory() as tmp_dir:
        png_paths = _generate_diagrams(data, tmp_dir)

        # ── テンプレ読込 -> セル流し込み ──────────────────────────
        wb = load_workbook(args.template)

        fill_revision(wb["改版履歴"], data, history)

        fill_overview(
            wb["概要"], data,
            changed_fields=set() if is_major else changed_scalars)

        fill_business_flow(
            wb["業務フロー"], data,
            changed_step_nos=set() if is_major else changed_flows,
            png_path=png_paths.get("swimlane"))

        fill_process_overview(
            wb["処理概要"], data,
            changed_step_nos=set() if is_major else changed_proc_steps,
            png_path=png_paths.get("flowchart"))

        fill_target_objects(
            wb["対象オブジェクト"], data,
            changed_obj_keys=set() if is_major else changed_objs,
            png_path=png_paths.get("er"))

        fill_related_components(
            wb["関連コンポーネント"], data,
            changed_comp_keys=set() if is_major else changed_comps,
            png_path=png_paths.get("component"))

        meta_payload = {
            "version": current_version,
            "date":    today,
            "author":  author,
            "data":    data,
            "history": history,
        }
        if args.source_hash:
            meta_payload["source_hash"] = args.source_hash
        write_meta(wb, meta_payload)

        wb.save(str(out_path))

    sys.stdout.buffer.write(
        f"[OK] 詳細設計書を生成しました: v{current_version} -> {out_path}\n".encode("utf-8"))

    # 同一 feature_id の旧ファイルのみクリーンアップ（別名・旧命名ファイルは触らない）
    if feature_id:
        for old_f in out_dir.glob(f"【{feature_id}】*.xlsx"):
            if old_f.resolve() != out_path.resolve():
                old_f.unlink()
                sys.stdout.buffer.write(
                    f"  [CLEANUP] 旧ファイルを削除: {old_f.name}\n".encode("utf-8"))


if __name__ == "__main__":
    main()
