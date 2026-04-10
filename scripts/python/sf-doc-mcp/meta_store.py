# -*- coding: utf-8 -*-
"""Excel _meta シートへのメタデータ読み書き"""

import json

from openpyxl import load_workbook

META_SHEET = "_meta"

# _meta に保存しない大きなキー（describe は生の Salesforce レスポンスで巨大、field_usage は毎回再取得）
_EXCLUDE_KEYS = {"describe", "field_usage"}


# Excel セルの文字数上限は 32,767 文字なので分割して保存する
_CHUNK = 30000


def read_meta(source_file: str) -> dict | None:
    """既存 Excel の _meta シートからメタデータを読み込む。
    見つからない・読み込み失敗時は None を返す。"""
    try:
        wb = load_workbook(source_file, read_only=True, data_only=True)
        if META_SHEET not in wb.sheetnames:
            print(f"  [INFO] _meta シートなし: {source_file}")
            wb.close()
            return None
        ws = wb[META_SHEET]
        chunks = []
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] is not None:
                chunks.append(str(row[0]))
        wb.close()
        if not chunks:
            return None
        return json.loads("".join(chunks))
    except FileNotFoundError:
        print(f"  [ERROR] ファイルが見つかりません: {source_file}")
        return None
    except Exception as e:
        print(f"  [WARN] _meta 読み込み失敗: {e}")
        return None


def write_meta(wb, data: dict):
    """Workbook に _meta シートを追加（非表示）。既存シートは上書き。
    JSON が 30,000 字を超える場合は複数行に分割して保存する。"""
    if META_SHEET in wb.sheetnames:
        del wb[META_SHEET]
    ws = wb.create_sheet(META_SHEET)
    ws.sheet_state = "hidden"
    json_str = json.dumps(data, ensure_ascii=False, default=str)
    for i, start in enumerate(range(0, len(json_str), _CHUNK)):
        ws.cell(row=i + 1, column=1, value=json_str[start:start + _CHUNK])


def strip_meta(obj: dict) -> dict:
    """保存不要な大きなキーを除去した軽量メタデータを返す"""
    return {k: v for k, v in obj.items() if k not in _EXCLUDE_KEYS}
