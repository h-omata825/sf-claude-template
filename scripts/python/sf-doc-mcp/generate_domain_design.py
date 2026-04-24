# scripts/python/sf-doc-mcp/generate_domain_design.py
# -*- coding: utf-8 -*-
"""
ドメイン設計書.xlsx を1ドメイン分生成する（テンプレート読込方式）。

  ドメイン設計書テンプレート.xlsx（build_domain_design_template.py で生成した「器」）を
  コピーしてセル値 + 図形PNGを流し込む。

5シート構成:
  1. 改版履歴
  2. ドメイン概要         : メタ + 業務目的 / 対象ユーザー / 業務概要 / 前提条件
  3. 業務フロー           : フロー図PNG + フロー説明テーブル
  4. 画面構成             : 画面一覧 + 画面遷移図PNG
  5. コンポーネント構成   : コンポーネント一覧 + 関連図PNG + 使用オブジェクト + 外部連携

Usage:
  python generate_domain_design.py \\
    --input  domain_design.json \\
    --template "C:/.../ドメイン設計書テンプレート.xlsx" \\
    --output-dir "C:/.../出力先" \\
    [--source-hash ""] \\
    [--version-increment minor]
"""
from __future__ import annotations

import argparse
import json
import math
import re
import sys
import tempfile
from datetime import date as _date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import design_revision as dr
from meta_store import read_meta, write_meta
from tmp_utils import get_project_tmp_dir, set_project_tmp_dir
from version_manager import increment_version

# ── 色定数 ─────────────────────────────────────────────────────────
C_HDR_BLUE  = "2E75B6"
C_BAND_BLUE = "0070C0"
C_LABEL_BG  = "D9E1F2"
C_FONT_D    = "000000"
C_FONT_W    = "FFFFFF"
C_FONT_GRAY = "595959"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

# ── テンプレートと一致させる定数 ────────────────────────────────────
# 改版履歴
REV_META_ROW       = 3
REV_META_PROJECT_V = (6, 18)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6,  11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

# ドメイン概要（各メタ/ラベル行の値セル）
DOM_META_ROW_1   = 3
DOM_META_ROW_2   = 4
DOM_META_1_V = {
    "project_name": (6,  16),
    "domain_id":    (21, 24),
    "author":       (28, 29),
    "version":      (31, 31),
}
DOM_META_2_V = {
    "name_ja": (6,  24),
    "date":    (28, 31),
}
DOM_LABEL_VAL_CS = 7
DOM_LABEL_VAL_CE = 31
DOM_SECTION_ROW = {
    "purpose":       7,
    "target_users":  8,
    "overview":      9,
    "prerequisites": 12,
    "notes":         13,
}

# 業務フロー
BF_FLOW_DESC_ROW   = 4   # 説明ラベル行（値セルは col 7-31）
BF_IMG_ANCHOR      = "B5"
BF_DESC_COL        = 19  # 図形エリアの説明テキスト列
BF_DESC_ROW        = 5   # 図形エリア開始行
BF_DATA_ROW_START  = 28
BF_STEP_CS,  BF_STEP_CE  = 2,  3
BF_ACTOR_CS, BF_ACTOR_CE = 4,  8
BF_ACT_CS,   BF_ACT_CE   = 9,  22
BF_SYS_CS,   BF_SYS_CE   = 23, 31

# 画面構成
SC_DATA_ROW_START  = 5
SC_NO_CS,    SC_NO_CE    = 2,  3
SC_NAME_CS,  SC_NAME_CE  = 4,  10
SC_COMP_CS,  SC_COMP_CE  = 11, 18
SC_DESC_CS,  SC_DESC_CE  = 19, 31
SC_IMG_ANCHOR      = "B17"
SC_DESC_COL        = 19
SC_DESC_ROW        = 17
SC_WF_DATA_ROW_START = 49   # ワイヤーフレーム開始行（テンプレート row48 がセクション帯）
SC_WF_ROWS_PER_IMG   = 22   # 1画面あたりの行数（画像エリア）

# コンポーネント構成
CM_DATA_ROW_START  = 5
CM_TYPE_CS, CM_TYPE_CE = 2,  5
CM_API_CS,  CM_API_CE  = 6,  14
CM_ROLE_CS, CM_ROLE_CE = 15, 31
CM_IMG_ANCHOR      = "B22"
CM_DESC_COL        = 19
CM_DESC_ROW        = 22

OB_DATA_ROW_START  = 49
OB_API_CS,   OB_API_CE   = 2,  7
OB_LABEL_CS, OB_LABEL_CE = 8,  14
OB_USE_CS,   OB_USE_CE   = 15, 31

EX_DATA_ROW_START  = 62
EX_NAME_CS,  EX_NAME_CE  = 2,  7
EX_SCHED_CS, EX_SCHED_CE = 8,  12
EX_DESC_CS,  EX_DESC_CE  = 13, 22
EX_OBJ_CS,   EX_OBJ_CE   = 23, 31

GRID_RIGHT = 31

SCALAR_FIELDS  = ["purpose", "target_users", "overview", "prerequisites"]
SECTION_SHEETS = {
    "business_flow":        "業務フロー",
    "screens":              "画面構成",
    "components":           "コンポーネント構成",
    "related_objects":      "コンポーネント構成",
    "external_integrations": "コンポーネント構成",
}


# ── スタイルヘルパー ────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h


# ── PNG埋め込み ────────────────────────────────────────────────────
def _embed_image(ws, png_path: str, anchor: str,
                 img_w: int = 440, img_h: int | None = None):
    """PNGをExcelシートに埋め込む。

    img_h が None の場合は縦横比を維持。指定された場合はそのサイズで設定。
    失敗時は警告のみ。
    """
    try:
        if not Path(png_path).exists():
            return
        img = XLImage(png_path)
        img.anchor = anchor
        if img_h is None:
            # 縦横比を維持
            ratio = img.height / img.width if img.width else 1.0
            img.width = img_w
            img.height = int(img_w * ratio)
        else:
            img.width = img_w
            img.height = img_h
        ws.add_image(img)
    except Exception as e:
        print(f"  [WARN] 画像埋め込み失敗({anchor}): {e}")


# ── シート埋め込み ─────────────────────────────────────────────────
def fill_revision(ws, data: dict, history: list[dict]):
    vs, _ = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("project_name", ""))
    vs, _ = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("date", ""))
    dr.fill_revision_table(ws, history, REV_COLS, REV_DATA_ROW_START)


def fill_domain_overview(ws, data: dict, changed_fields: set):
    # メタ行1
    for key, (cs, _) in DOM_META_1_V.items():
        ws.cell(row=DOM_META_ROW_1, column=cs, value=data.get(key, ""))
    # メタ行2
    for key, (cs, _) in DOM_META_2_V.items():
        ws.cell(row=DOM_META_ROW_2, column=cs, value=data.get(key, ""))
    # セクション本文
    for key, row in DOM_SECTION_ROW.items():
        val = data.get(key, "")
        if val:
            cell = ws.cell(row=row, column=DOM_LABEL_VAL_CS, value=val)
            if key in changed_fields:
                dr.apply_red(cell, size=10)


def fill_business_flow(ws, data: dict, changed_step_nos: set,
                       bf_png_path: str | None):
    # PNG埋め込み（全幅）
    if bf_png_path:
        _embed_image(ws, bf_png_path, BF_IMG_ANCHOR, img_w=840)

    # フロー説明テーブル
    flows = data.get("business_flow", [])
    r = BF_DATA_ROW_START
    for i, flow in enumerate(flows):
        step_no = flow.get("step", str(i + 1))
        is_changed = step_no in changed_step_nos
        set_h(ws, r, 24)
        c1 = MW(ws, r, BF_STEP_CS,  BF_STEP_CE,  step_no,
                border=B_all(), h="center")
        c2 = MW(ws, r, BF_ACTOR_CS, BF_ACTOR_CE, flow.get("actor", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, BF_ACT_CS,   BF_ACT_CE,   flow.get("action", ""),
                border=B_all(), wrap=True, v="top")
        # description を「関連コンポーネント」列に流用（system + description）
        sys_text = flow.get("system", "")
        desc = flow.get("description", "")
        comp_text = f"{sys_text}\n{desc}".strip() if desc else sys_text
        c4 = MW(ws, r, BF_SYS_CS,   BF_SYS_CE,   comp_text,
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3, c4):
                dr.apply_red(c)
        r += 1


def fill_screens(ws, data: dict, changed_screen_keys: set,
                 sc_png_path: str | None,
                 wireframe_paths: list[tuple[str, str | None, int]] | None = None):
    # 画面一覧テーブル
    screens = data.get("screens", [])
    r = SC_DATA_ROW_START
    for i, scr in enumerate(screens):
        name = scr.get("name", "")
        is_changed = name in changed_screen_keys
        set_h(ws, r, 22)
        c1 = MW(ws, r, SC_NO_CS,   SC_NO_CE,   str(i + 1),
                border=B_all(), h="center")
        c2 = MW(ws, r, SC_NAME_CS, SC_NAME_CE, name,
                border=B_all())
        c3 = MW(ws, r, SC_COMP_CS, SC_COMP_CE, scr.get("component", ""),
                border=B_all())
        c4 = MW(ws, r, SC_DESC_CS, SC_DESC_CE, scr.get("description", ""),
                border=B_all(), wrap=True)
        if is_changed:
            for c in (c1, c2, c3, c4):
                dr.apply_red(c)
        r += 1

    # PNG埋め込み（画面遷移図・全幅）
    if sc_png_path:
        _embed_image(ws, sc_png_path, SC_IMG_ANCHOR, img_w=840)

    # ── 画面ワイヤーフレーム埋め込み ──────────────────────────────
    if wireframe_paths:
        r = SC_WF_DATA_ROW_START
        for screen_name, wf_path, img_h_px in wireframe_paths:
            # タイトル行（画面名）
            set_h(ws, r, 26)
            MW(ws, r, 2, 31, screen_name, bold=True,
               bg=C_LABEL_BG, border=B_all())
            r += 1
            # 画像の縦サイズから必要行数を計算（1行 ≒ 18px）
            if img_h_px and img_h_px > 0:
                rows_for_img = max(10, math.ceil(img_h_px / 18))
            else:
                rows_for_img = SC_WF_ROWS_PER_IMG
            # 画像エリア（rows_for_img 行分確保）
            for rr in range(r, r + rows_for_img):
                set_h(ws, rr, 20)
                for col in range(2, GRID_RIGHT + 1):
                    ws.cell(row=rr, column=col).fill = _fill("F2F2F2")
                    ws.cell(row=rr, column=col).border = B_all()
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r + rows_for_img - 1, end_column=GRID_RIGHT)
            if wf_path and Path(wf_path).exists():
                _embed_image(ws, wf_path, f"B{r}", img_w=840, img_h=None)
            r += rows_for_img + 1  # +1 スペーサー


def fill_components(ws, data: dict, changed_comp_keys: set,
                    changed_obj_keys: set, cm_png_path: str | None):
    # ── コンポーネント一覧 ──
    components = data.get("components", [])
    r = CM_DATA_ROW_START
    for comp in components:
        key = comp.get("api_name", "")
        is_changed = key in changed_comp_keys
        set_h(ws, r, 22)
        c1 = MW(ws, r, CM_TYPE_CS, CM_TYPE_CE, comp.get("type", ""),
                border=B_all(), h="center")
        c2 = MW(ws, r, CM_API_CS,  CM_API_CE,  key,
                border=B_all())
        c3 = MW(ws, r, CM_ROLE_CS, CM_ROLE_CE, comp.get("role", ""),
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # PNG埋め込み（全幅）
    if cm_png_path:
        _embed_image(ws, cm_png_path, CM_IMG_ANCHOR, img_w=840)

    # ── 使用オブジェクト ──
    objects = data.get("related_objects", [])
    r = OB_DATA_ROW_START
    for obj in objects:
        key = obj.get("api_name", "")
        is_changed = key in changed_obj_keys
        set_h(ws, r, 22)
        c1 = MW(ws, r, OB_API_CS,   OB_API_CE,   key,
                border=B_all())
        c2 = MW(ws, r, OB_LABEL_CS, OB_LABEL_CE, obj.get("label", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, OB_USE_CS,   OB_USE_CE,   obj.get("usage", ""),
                border=B_all(), wrap=True)
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # ── 外部連携・バッチ定義 ──
    integrations = data.get("external_integrations", [])
    r = EX_DATA_ROW_START
    for itg in integrations:
        set_h(ws, r, 22)
        MW(ws, r, EX_NAME_CS,  EX_NAME_CE,  itg.get("name", ""),
           border=B_all())
        MW(ws, r, EX_SCHED_CS, EX_SCHED_CE, itg.get("schedule", ""),
           border=B_all(), h="center")
        MW(ws, r, EX_DESC_CS,  EX_DESC_CE,  itg.get("description", ""),
           border=B_all(), wrap=True)
        MW(ws, r, EX_OBJ_CS,   EX_OBJ_CE,  itg.get("target_object", ""),
           border=B_all(), wrap=True)
        r += 1


# ── 差分計算 ────────────────────────────────────────────────────────
def _compute_diffs(prev_data: dict | None, new_data: dict) -> dict:
    if prev_data is None:
        return {"scalars": [], "lists": {}}
    return {
        "scalars": dr.diff_scalars(prev_data, new_data, SCALAR_FIELDS),
        "lists": {
            "business_flow": dr.diff_list(
                prev_data.get("business_flow", []),
                new_data.get("business_flow", []), "step"),
            "screens": dr.diff_list(
                prev_data.get("screens", []),
                new_data.get("screens", []), "name"),
            "components": dr.diff_list(
                prev_data.get("components", []),
                new_data.get("components", []), "api_name"),
            "related_objects": dr.diff_list(
                prev_data.get("related_objects", []),
                new_data.get("related_objects", []), "api_name"),
            "external_integrations": dr.diff_list(
                prev_data.get("external_integrations", []),
                new_data.get("external_integrations", []), "name"),
        },
    }


# ── PNG生成 ────────────────────────────────────────────────────────
def _generate_diagrams(data: dict, tmp_dir: str) -> dict[str, str | None]:
    """3種の図形PNGを生成し、パスを返す。失敗したらNone。"""
    paths: dict[str, str | None] = {
        "business_flow": None,
        "screen_transition": None,
        "component": None,
    }
    try:
        from diagram_utils import (
            generate_business_flow_diagram,
            generate_screen_transition_diagram,
            generate_component_diagram,
        )
    except ImportError:
        print("  [WARN] diagram_utils をインポートできません。図形生成をスキップします。")
        return paths

    # 業務フロー図
    flows = data.get("business_flow", [])
    if flows:
        bf_path = str(Path(tmp_dir) / "business_flow.png")
        try:
            if generate_business_flow_diagram(flows, bf_path):
                paths["business_flow"] = bf_path
                print(f"  [OK] 業務フロー図生成: {bf_path}")
        except Exception as e:
            print(f"  [WARN] 業務フロー図生成失敗: {e}")

    # 画面遷移図
    screens = data.get("screens", [])
    if screens:
        sc_path = str(Path(tmp_dir) / "screen_transition.png")
        try:
            trans = data.get("transitions", []) or None
            if generate_screen_transition_diagram(screens, sc_path, transitions=trans):
                paths["screen_transition"] = sc_path
                print(f"  [OK] 画面遷移図生成: {sc_path}")
        except Exception as e:
            print(f"  [WARN] 画面遷移図生成失敗: {e}")

    # コンポーネント関連図
    components = data.get("components", [])
    if components:
        cm_path = str(Path(tmp_dir) / "component.png")
        try:
            if generate_component_diagram(components, cm_path):
                paths["component"] = cm_path
                print(f"  [OK] コンポーネント関連図生成: {cm_path}")
        except Exception as e:
            print(f"  [WARN] コンポーネント関連図生成失敗: {e}")

    return paths


def _generate_wireframes(data: dict, tmp_dir: str,
                          project_dir: str | None) -> list[tuple[str, str | None, int]]:
    """各画面のワイヤーフレームPNGを生成。(screen_name, png_path|None, img_h_px) のリストを返す。"""
    if not project_dir:
        return []
    try:
        from diagram_utils import extract_lwc_ui_elements, generate_screen_wireframe
        has_mpl = True
    except ImportError:
        has_mpl = False

    try:
        from diagram_utils import generate_screen_wireframe_playwright
        has_playwright_fn = True
    except ImportError:
        has_playwright_fn = False

    if not has_mpl and not has_playwright_fn:
        print("  [WARN] diagram_utils をインポートできません。ワイヤーフレーム生成をスキップします。")
        return []

    results: list[tuple[str, str | None, int]] = []
    pd_root = Path(project_dir)

    for scr in data.get("screens", []):
        name = scr.get("name", "")
        comp = scr.get("component", "")
        if not comp:
            results.append((name, None, 0))
            continue

        html_path = (pd_root / "force-app" / "main" / "default"
                     / "lwc" / comp / f"{comp}.html")
        if not html_path.exists():
            print(f"  [WARN] LWC HTML が見つかりません: {html_path}")
            results.append((name, None, 0))
            continue

        try:
            html_content = html_path.read_text(encoding="utf-8")
            safe = re.sub(r'[\\/:*?"<>|]', "_", name or comp)
            wf_path = str(Path(tmp_dir) / f"wireframe_{safe}.png")
            wf_ok = False
            img_h_px = 0

            # Playwright 優先
            if has_playwright_fn:
                wf_ok, img_h_px = generate_screen_wireframe_playwright(name or comp, html_content, wf_path)

            # matplotlib fallback
            if not wf_ok and has_mpl:
                elements = extract_lwc_ui_elements(html_content)
                wf_ok = generate_screen_wireframe(name or comp, elements, wf_path)
                img_h_px = 0  # matplotlib fallback では画像サイズ不明

            if wf_ok:
                print(f"  [OK] ワイヤーフレーム生成: {name}")
                results.append((name, wf_path, img_h_px))
            else:
                results.append((name, None, 0))
        except Exception as e:
            print(f"  [WARN] ワイヤーフレーム生成失敗({name}): {e}")
            results.append((name, None, 0))

    return results


# ── メイン ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="ドメイン設計書 Excel 生成")
    parser.add_argument("--input",      required=True, help="ドメイン設計 JSON ファイルパス")
    parser.add_argument("--template",   required=True, help="ドメイン設計書テンプレート.xlsx パス")
    parser.add_argument("--output-dir", required=True, help="出力先ディレクトリ")
    parser.add_argument("--project-dir", default="",
                        help="Salesforce プロジェクトルートパス（LWC HTML 参照用）")
    parser.add_argument("--source-hash", default="",
                        help="ソースファイルの SHA256 ハッシュ")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"])
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    today  = _date.today().strftime("%Y-%m-%d")
    author = data.get("author", "")

    domain_id = data.get("domain_id", "DOM-000")
    name_ja   = data.get("name_ja", "ドメイン")
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", name_ja)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    set_project_tmp_dir(out_dir)
    out_path = out_dir / f"【{domain_id}】{safe_name}.xlsx"

    # ── バージョン判定 ────────────────────────────────────────────
    source_file = ""
    existing = sorted(out_dir.glob(f"【{domain_id}】*.xlsx"),
                      key=lambda f: f.stat().st_mtime, reverse=True)
    if existing:
        source_file = str(existing[0])
        print(f"  [AUTO] 既存ファイルを検出: {existing[0].name}")

    prev_meta   = read_meta(source_file) if source_file else None
    prev_data   = prev_meta.get("data") if prev_meta else None

    if prev_meta:
        prev_history_len = len(prev_meta.get("history", []))
        # 改版履歴 20 行制限: 既存履歴が 20 以上なら minor 指定でも major に強制昇格し履歴リセット
        forced_major = False
        if prev_history_len >= 20 and args.version_increment == "minor":
            print(f"  [WARN] 改版履歴が {prev_history_len} 件に達しているため minor → major に強制昇格し、履歴をリセットします")
            args.version_increment = "major"
            forced_major = True
        current_version = increment_version(
            prev_meta.get("version", "1.0"), args.version_increment)
        # major 時は履歴リセット（手動・強制問わず。メジャーUP 1行だけ残す）
        history    = [] if args.version_increment == "major" else prev_meta.get("history", [])
        is_initial = False
        if forced_major:
            print(f"メジャー昇格モード（履歴リセット）: {prev_meta.get('version', '?')} -> {current_version}")
        else:
            print(f"更新モード: {prev_meta.get('version', '?')} -> {current_version}")
    else:
        current_version = data.get("version") or "1.0"
        history    = []
        is_initial = True
        print(f"新規作成モード: v{current_version}")

    data["version"] = current_version
    if not data.get("date"):
        data["date"] = today

    diffs = _compute_diffs(prev_data, data)
    if prev_meta and args.version_increment == "minor" and not dr.has_any_diff(diffs):
        print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    last_no = max((h["項番"] for h in history
                   if isinstance(h.get("項番"), int)), default=0)
    new_entries = dr.build_entries(
        current_version, diffs, author, today,
        start_no=last_no + 1,
        is_major=(args.version_increment == "major"),
        is_initial=is_initial,
        section_sheet_map=SECTION_SHEETS,
        scalar_sheet="ドメイン概要",
    )
    history = history + new_entries

    changed_scalars = dr.changed_scalar_fields(diffs)
    changed_flows   = dr.changed_ids(diffs, "business_flow")
    changed_screens = dr.changed_ids(diffs, "screens")
    changed_comps   = dr.changed_ids(diffs, "components")
    changed_objs    = dr.changed_ids(diffs, "related_objects")
    is_major        = (args.version_increment == "major")

    # ── 図形PNG生成 ──────────────────────────────────────────────
    with tempfile.TemporaryDirectory(dir=get_project_tmp_dir()) as tmp_dir:
        png_paths = _generate_diagrams(data, tmp_dir)
        wireframe_paths = _generate_wireframes(
            data, tmp_dir, args.project_dir or None)

        # ── テンプレ読込 -> セル流し込み ──────────────────────────
        wb = load_workbook(args.template)

        fill_revision(wb["改版履歴"], data, history)

        fill_domain_overview(
            wb["ドメイン概要"], data,
            changed_fields=set() if is_major else changed_scalars)

        fill_business_flow(
            wb["業務フロー"], data,
            changed_step_nos=set() if is_major else changed_flows,
            bf_png_path=png_paths.get("business_flow"))

        fill_screens(
            wb["画面構成"], data,
            changed_screen_keys=set() if is_major else changed_screens,
            sc_png_path=png_paths.get("screen_transition"),
            wireframe_paths=wireframe_paths or None)

        fill_components(
            wb["コンポーネント構成"], data,
            changed_comp_keys=set() if is_major else changed_comps,
            changed_obj_keys=set() if is_major else changed_objs,
            cm_png_path=png_paths.get("component"))

        meta_payload = {
            "version": current_version,
            "date":    today,
            "author":  author,
            "data":    data,
            "history": history,
        }
        if args.source_hash:
            meta_payload["source_hash"] = args.source_hash
        write_meta(wb, meta_payload)

        wb.save(str(out_path))

    sys.stdout.buffer.write(
        f"[OK] ドメイン設計書を生成しました: v{current_version} -> {out_path}\n".encode("utf-8"))

    # 同一IDで別名の旧ファイルを削除
    for old_f in out_dir.glob(f"【{domain_id}】*.xlsx"):
        if old_f.resolve() != out_path.resolve():
            old_f.unlink()
            sys.stdout.buffer.write(
                f"  [CLEANUP] 旧ファイルを削除: {old_f.name}\n".encode("utf-8"))


if __name__ == "__main__":
    main()
