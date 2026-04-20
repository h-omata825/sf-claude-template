# scripts/python/sf-doc-mcp/build_domain_design_template.py
"""
ドメイン設計書テンプレート Excel を生成する。

5シート構成:
  1. 改版履歴               : 基本情報 + 改版履歴テーブル
  2. ドメイン概要           : メタ + 業務目的 / 対象ユーザー / 業務概要 / 前提条件
  3. 業務フロー             : フロー図 + フロー説明テーブル
  4. 画面構成               : 画面一覧 + 画面遷移図
  5. コンポーネント構成     : コンポーネント一覧 + 関連図 + 使用オブジェクト + 外部連携

方針:
  - 基本設計書テンプレートと同じデザインシステム（狭幅グリッド + セル結合 + 罫線統一）
  - グリッド: col A=2.0、col 2〜31=4.2（30列）
  - 色: 1F3864（タイトル濃紺）/ 2E75B6（ヘッダ青）/ 0070C0（セクション帯）/ D9E1F2（ラベル薄青）

Usage:
  python build_domain_design_template.py --output "C:/.../ドメイン設計書テンプレート.xlsx"
"""
from __future__ import annotations
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── カラー ──────────────────────────────────────────────────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"
C_FONT_GRAY  = "595959"
C_PLACEHOLDER_BG = "F2F2F2"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

# ── ヘルパー ───────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_med():
    return Border(left=MED, right=MED, top=MED, bottom=MED)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h): ws.row_dimensions[row].height = h
def set_w(ws, letter, w): ws.column_dimensions[letter].width = w


# ── 共通グリッド ───────────────────────────────────────────────
GRID_LEFT  = 2
GRID_RIGHT = 31

def setup_grid(ws):
    set_w(ws, "A", 2.0)
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        set_w(ws, get_column_letter(i), 4.2)
    ws.sheet_view.showGridLines = False

def title_band(ws, row, text, height=36):
    set_h(ws, row, height)
    MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_TITLE_DARK, h="center", size=14)

def section_band(ws, row, text, height=26, cs=None, ce=None):
    cs = cs if cs is not None else GRID_LEFT
    ce = ce if ce is not None else GRID_RIGHT
    set_h(ws, row, height)
    MW(ws, row, cs, ce, text,
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11)

def table_header(ws, row, cols: list[tuple], height=26):
    """cols: [(cs, ce, label), ...]"""
    set_h(ws, row, height)
    for cs, ce, label in cols:
        MW(ws, row, cs, ce, label,
           bold=True, fg=C_FONT_W, bg=C_HDR_BLUE, h="center", border=B_all())

def data_rows(ws, row_start, row_end, col_groups: list[tuple], row_h=22):
    """col_groups: [(cs, ce), ...] — 各行に空セルを罫線付きで配置"""
    for r in range(row_start, row_end + 1):
        set_h(ws, r, row_h)
        for cs, ce in col_groups:
            MW(ws, r, cs, ce, "", border=B_all())

def label_row(ws, row, label, label_cs=2, label_ce=6, val_cs=7, val_ce=31, height=22):
    set_h(ws, row, height)
    MW(ws, row, label_cs, label_ce, label,
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, row, val_cs, val_ce, "", border=B_all())

def diagram_area(ws, row_start, row_end, img_cs=2, img_ce=18,
                 desc_cs=19, desc_ce=31, row_h=18):
    """図形エリア: 左に画像プレースホルダー、右に説明テキストエリア。"""
    # 画像エリア（結合）
    for r in range(row_start, row_end + 1):
        set_h(ws, r, row_h)
        for c in range(img_cs, img_ce + 1):
            ws.cell(row=r, column=c).border = B_all()
            ws.cell(row=r, column=c).fill = _fill(C_PLACEHOLDER_BG)
        for c in range(desc_cs, desc_ce + 1):
            ws.cell(row=r, column=c).border = B_all()

    ws.merge_cells(start_row=row_start, start_column=img_cs,
                   end_row=row_end, end_column=img_ce)
    ws.merge_cells(start_row=row_start, start_column=desc_cs,
                   end_row=row_end, end_column=desc_ce)

    # プレースホルダーテキスト
    cell = ws.cell(row=row_start, column=img_cs)
    cell.value = "(図がここに挿入されます)"
    cell.font = _fnt(color=C_FONT_GRAY, size=9)
    cell.alignment = _aln(h="center", v="center")

    # 説明テキストエリア
    desc_cell = ws.cell(row=row_start, column=desc_cs)
    desc_cell.alignment = _aln(h="left", v="top", wrap=True)


# ─────────────────────────────────────────────────────────────
# 改版履歴テーブル定数
# ─────────────────────────────────────────────────────────────
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6,  11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}


# ─────────────────────────────────────────────────────────────
# Sheet 1: 改版履歴
# ─────────────────────────────────────────────────────────────
def build_revision(wb):
    ws = wb.active
    ws.title = "改版履歴"
    setup_grid(ws)

    title_band(ws, 1, "改版履歴")
    set_h(ws, 2, 6)

    # メタ行
    set_h(ws, 3, 22)
    MW(ws, 3, 2, 5,  "プロジェクト名", bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 6, 18, "", border=B_all())
    MW(ws, 3, 19, 22, "作成日",        bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 23, 31, "", border=B_all())

    set_h(ws, 4, 8)

    # テーブルヘッダー
    table_header(ws, 5, [(cs, ce, label) for label, (cs, ce) in REV_COLS.items()])

    # データ行 (20行)
    data_rows(ws, 6, 25, [(cs, ce) for cs, ce in REV_COLS.values()])


# ─────────────────────────────────────────────────────────────
# Sheet 2: ドメイン概要
# ─────────────────────────────────────────────────────────────
def build_domain_overview(wb):
    ws = wb.create_sheet("ドメイン概要")
    setup_grid(ws)

    title_band(ws, 1, "ドメイン設計書 — ドメイン概要")
    set_h(ws, 2, 6)

    # メタ 1段目 (row 3)
    set_h(ws, 3, 22)
    MW(ws, 3, 2,  5,  "プロジェクト名", bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 6,  16, "", border=B_all())
    MW(ws, 3, 17, 20, "ドメインID",    bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 21, 24, "", border=B_all())
    MW(ws, 3, 25, 27, "作成者",        bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 28, 29, "", border=B_all())
    MW(ws, 3, 30, 30, "版数",          bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 31, 31, "", border=B_all())

    # メタ 2段目 (row 4)
    set_h(ws, 4, 22)
    MW(ws, 4, 2,  5,  "ドメイン名",    bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 4, 6,  24, "", border=B_all())
    MW(ws, 4, 25, 27, "作成日",        bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 4, 28, 31, "", border=B_all())

    # row 5: スペーサー
    set_h(ws, 5, 10)

    # セクション1: ドメイン概要
    section_band(ws, 6, "■ 1. ドメイン概要")
    label_row(ws, 7,  "業務目的",     height=50)
    label_row(ws, 8,  "対象ユーザー")
    label_row(ws, 9,  "業務概要",     height=50)

    # row 10: スペーサー
    set_h(ws, 10, 10)

    # セクション2: 前提条件・備考
    section_band(ws, 11, "■ 2. 前提条件・備考")
    label_row(ws, 12, "前提条件", height=40)
    label_row(ws, 13, "備考",     height=36)


# ─────────────────────────────────────────────────────────────
# Sheet 3: 業務フロー
# ─────────────────────────────────────────────────────────────
BF_COLS = [
    (2,  3,  "No"),
    (4,  8,  "担当"),
    (9,  22, "操作・処理内容"),
    (23, 31, "関連コンポーネント"),
]

def build_business_flow(wb):
    ws = wb.create_sheet("業務フロー")
    setup_grid(ws)

    title_band(ws, 1, "ドメイン設計書 — 業務フロー")
    set_h(ws, 2, 6)

    # row 3: section_band
    section_band(ws, 3, "■ 業務フロー図")

    # row 4: 説明ラベル行
    set_h(ws, 4, 22)
    MW(ws, 4, 2,  6,  "図", bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 4, 7,  31, "(業務フロー図がここに挿入されます)",
       border=B_all(), fg=C_FONT_GRAY)

    # row 5〜24: 図形エリア (20行)
    diagram_area(ws, 5, 24)

    # row 25: スペーサー
    set_h(ws, 25, 10)

    # row 26: section_band
    section_band(ws, 26, "■ フロー説明")

    # row 27: table_header
    table_header(ws, 27, BF_COLS)

    # row 28〜42: data_rows (15行)
    data_rows(ws, 28, 42, [(cs, ce) for cs, ce, _ in BF_COLS], row_h=24)


# ─────────────────────────────────────────────────────────────
# Sheet 4: 画面構成
# ─────────────────────────────────────────────────────────────
SC_COLS = [
    (2,  3,  "No"),
    (4,  10, "画面名"),
    (11, 18, "コンポーネント"),
    (19, 31, "説明"),
]

def build_screen_layout(wb):
    ws = wb.create_sheet("画面構成")
    setup_grid(ws)

    title_band(ws, 1, "ドメイン設計書 — 画面構成")
    set_h(ws, 2, 6)

    # row 3: section_band
    section_band(ws, 3, "■ 1. 画面一覧")

    # row 4: table_header
    table_header(ws, 4, SC_COLS)

    # row 5〜14: data_rows (10行)
    data_rows(ws, 5, 14, [(cs, ce) for cs, ce, _ in SC_COLS])

    # row 15: スペーサー
    set_h(ws, 15, 10)

    # row 16: section_band
    section_band(ws, 16, "■ 2. 画面遷移図")

    # row 17〜36: 図形エリア (20行)
    diagram_area(ws, 17, 36)

    # row 37: スペーサー
    set_h(ws, 37, 10)

    # row 38: section_band for wireframes
    section_band(ws, 38, "■ 3. 画面ワイヤーフレーム")


# ─────────────────────────────────────────────────────────────
# Sheet 5: コンポーネント構成
# ─────────────────────────────────────────────────────────────
CM_COLS = [
    (2,  5,  "種別"),
    (6,  14, "API名"),
    (15, 31, "役割"),
]
OB_COLS = [
    (2,  7,  "API名"),
    (8,  14, "ラベル"),
    (15, 31, "用途"),
]
EX_COLS = [
    (2,  7,  "種別/名称"),
    (8,  12, "スケジュール"),
    (13, 22, "処理内容"),
    (23, 31, "対象オブジェクト"),
]

def build_component_layout(wb):
    ws = wb.create_sheet("コンポーネント構成")
    setup_grid(ws)

    title_band(ws, 1, "ドメイン設計書 — コンポーネント構成")
    set_h(ws, 2, 6)

    # ── 1. コンポーネント一覧 ──
    section_band(ws, 3, "■ 1. コンポーネント一覧")
    table_header(ws, 4, CM_COLS)
    data_rows(ws, 5, 19, [(cs, ce) for cs, ce, _ in CM_COLS])

    # row 20: スペーサー
    set_h(ws, 20, 10)

    # ── 2. コンポーネント関連図 ──
    section_band(ws, 21, "■ 2. コンポーネント関連図")
    diagram_area(ws, 22, 41)

    # row 42: スペーサー
    set_h(ws, 42, 10)

    # ── 3. 使用オブジェクト ──
    section_band(ws, 43, "■ 3. 使用オブジェクト")
    table_header(ws, 44, OB_COLS)
    data_rows(ws, 45, 54, [(cs, ce) for cs, ce, _ in OB_COLS])

    # row 55: スペーサー
    set_h(ws, 55, 10)

    # ── 4. 外部連携・バッチ定義 ──
    section_band(ws, 56, "■ 4. 外部連携・バッチ定義")
    table_header(ws, 57, EX_COLS)
    data_rows(ws, 58, 63, [(cs, ce) for cs, ce, _ in EX_COLS])


# ─────────────────────────────────────────────────────────────
# エントリポイント
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, help="出力先 xlsx パス")
    args = parser.parse_args()

    wb = Workbook()
    build_revision(wb)
    build_domain_overview(wb)
    build_business_flow(wb)
    build_screen_layout(wb)
    build_component_layout(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"ドメイン設計書テンプレート生成完了: {out}")


if __name__ == "__main__":
    main()
