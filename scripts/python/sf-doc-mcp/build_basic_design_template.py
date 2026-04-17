"""
基本設計書テンプレート Excel を生成する。

4シート構成:
  1. 改版履歴             : 基本情報 + 改版履歴テーブル
  2. グループ概要         : メタ + 業務目的 / 対象ユーザー / 利用シーン / 前提条件
  3. 業務フロー           : No / 担当 / 操作・処理内容 / 関連コンポーネント
  4. コンポーネント・オブジェクト: 構成コンポーネント表 + 関連オブジェクト表 + 外部連携表

方針:
  - 設計書テンプレートと同じデザインシステム（狭幅グリッド + セル結合 + 罫線統一）
  - グリッド: col A=2.0、col 2〜31=4.2（30列）
  - 色: 1F3864（タイトル濃紺）/ 2E75B6（ヘッダ青）/ 0070C0（セクション帯）/ D9E1F2（ラベル薄青）

Usage:
  python build_basic_design_template.py --output "C:/.../基本設計書テンプレート.xlsx"
"""
from __future__ import annotations
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── カラー ──────────────────────────────────────────────────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

# ── ヘルパー ───────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_med():
    return Border(left=MED, right=MED, top=MED, bottom=MED)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h): ws.row_dimensions[row].height = h
def set_w(ws, letter, w): ws.column_dimensions[letter].width = w


# ── 共通グリッド ───────────────────────────────────────────────
GRID_LEFT  = 2
GRID_RIGHT = 31

def setup_grid(ws):
    set_w(ws, "A", 2.0)
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        set_w(ws, get_column_letter(i), 4.2)
    ws.sheet_view.showGridLines = False

def title_band(ws, row, text, height=36):
    set_h(ws, row, height)
    MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_TITLE_DARK, h="center", size=14)

def section_band(ws, row, text, height=26, cs=None, ce=None):
    cs = cs if cs is not None else GRID_LEFT
    ce = ce if ce is not None else GRID_RIGHT
    set_h(ws, row, height)
    MW(ws, row, cs, ce, text,
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11)

def table_header(ws, row, cols: list[tuple], height=26):
    """cols: [(cs, ce, label), ...]"""
    set_h(ws, row, height)
    for cs, ce, label in cols:
        MW(ws, row, cs, ce, label,
           bold=True, fg=C_FONT_W, bg=C_HDR_BLUE, h="center", border=B_all())

def data_rows(ws, row_start, row_end, col_groups: list[tuple], row_h=22):
    """col_groups: [(cs, ce), ...] — 各行に空セルを罫線付きで配置"""
    for r in range(row_start, row_end + 1):
        set_h(ws, r, row_h)
        for cs, ce in col_groups:
            MW(ws, r, cs, ce, "", border=B_all())

def label_row(ws, row, label, label_cs=2, label_ce=6, val_cs=7, val_ce=31, height=22):
    set_h(ws, row, height)
    MW(ws, row, label_cs, label_ce, label,
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, row, val_cs, val_ce, "", border=B_all())


# ─────────────────────────────────────────────────────────────
# Sheet 1: 改版履歴
# ─────────────────────────────────────────────────────────────
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6,  11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

def build_revision(wb):
    ws = wb.active
    ws.title = "改版履歴"
    setup_grid(ws)

    title_band(ws, 1, "改版履歴")
    set_h(ws, 2, 6)

    # メタ行
    set_h(ws, 3, 22)
    MW(ws, 3, 2, 5,  "プロジェクト名", bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 6, 18, "", border=B_all())
    MW(ws, 3, 19, 22, "作成日",        bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 23, 31, "", border=B_all())

    set_h(ws, 4, 8)

    # テーブルヘッダー
    table_header(ws, 5, [(cs, ce, label) for label, (cs, ce) in REV_COLS.items()])

    # データ行 (20行)
    data_rows(ws, 6, 25, [(cs, ce) for cs, ce in REV_COLS.values()])


# ─────────────────────────────────────────────────────────────
# Sheet 2: グループ概要
# ─────────────────────────────────────────────────────────────
def build_group_overview(wb):
    ws = wb.create_sheet("グループ概要")
    setup_grid(ws)

    title_band(ws, 1, "基本設計書 — グループ概要")
    set_h(ws, 2, 6)

    # メタ 1段目: プロジェクト名 / グループID / 作成者 / 版数
    set_h(ws, 3, 22)
    MW(ws, 3, 2,  5,  "プロジェクト名", bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 6,  16, "", border=B_all())
    MW(ws, 3, 17, 20, "グループID",    bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 21, 24, "", border=B_all())
    MW(ws, 3, 25, 27, "作成者",        bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 28, 29, "", border=B_all())
    MW(ws, 3, 30, 30, "版数",          bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 31, 31, "", border=B_all())

    # メタ 2段目: グループ名 / 作成日
    set_h(ws, 4, 22)
    MW(ws, 4, 2,  5,  "グループ名",    bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 4, 6,  24, "", border=B_all())
    MW(ws, 4, 25, 27, "作成日",        bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 4, 28, 31, "", border=B_all())

    set_h(ws, 5, 10)

    # セクション1: グループ概要
    section_band(ws, 6, "■ 1. グループ概要")
    label_row(ws, 7,  "業務目的",     height=40)
    label_row(ws, 8,  "対象ユーザー")
    label_row(ws, 9,  "利用シーン",   height=36)

    set_h(ws, 10, 10)

    # セクション2: 前提条件・備考
    section_band(ws, 11, "■ 2. 前提条件・備考")
    label_row(ws, 12, "前提条件", height=40)
    label_row(ws, 13, "備考",     height=36)


# ─────────────────────────────────────────────────────────────
# Sheet 3: 業務フロー
# ─────────────────────────────────────────────────────────────
BF_COLS = [
    (2,  3,  "No"),
    (4,  8,  "担当"),
    (9,  22, "操作・処理内容"),
    (23, 31, "関連コンポーネント"),
]

def build_business_flow(wb):
    ws = wb.create_sheet("業務フロー")
    setup_grid(ws)

    title_band(ws, 1, "基本設計書 — 業務の流れ")
    set_h(ws, 2, 6)

    section_band(ws, 3, "■ 業務フロー")
    table_header(ws, 4, BF_COLS)

    # データ行 (15行)
    data_rows(ws, 5, 19, [(cs, ce) for cs, ce, _ in BF_COLS], row_h=24)


# ─────────────────────────────────────────────────────────────
# Sheet 4: コンポーネント・オブジェクト
# ─────────────────────────────────────────────────────────────
CM_COLS = [
    (2,  5,  "種別"),
    (6,  14, "API名"),
    (15, 31, "役割概要"),
]
OB_COLS = [
    (2,  7,  "API名"),
    (8,  14, "ラベル"),
    (15, 31, "用途"),
]
EX_COLS = [
    (2,  7,  "連携先"),
    (8,  10, "方向"),
    (11, 21, "データ内容"),
    (22, 31, "タイミング"),
]

def build_components(wb):
    ws = wb.create_sheet("コンポーネント・オブジェクト")
    setup_grid(ws)

    title_band(ws, 1, "基本設計書 — コンポーネント・オブジェクト定義")
    set_h(ws, 2, 6)

    # ── 構成コンポーネント ──
    section_band(ws, 3, "■ 1. 構成コンポーネント")
    table_header(ws, 4, CM_COLS)
    data_rows(ws, 5, 19, [(cs, ce) for cs, ce, _ in CM_COLS])

    set_h(ws, 20, 10)

    # ── 関連オブジェクト ──
    section_band(ws, 21, "■ 2. 関連オブジェクト")
    table_header(ws, 22, OB_COLS)
    data_rows(ws, 23, 32, [(cs, ce) for cs, ce, _ in OB_COLS])

    set_h(ws, 33, 10)

    # ── 外部連携（データがない場合は空テーブルのまま残す）──
    section_band(ws, 34, "■ 3. 外部連携")
    table_header(ws, 35, EX_COLS)
    data_rows(ws, 36, 41, [(cs, ce) for cs, ce, _ in EX_COLS])


# ─────────────────────────────────────────────────────────────
# エントリポイント
# ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, help="出力先 xlsx パス")
    args = parser.parse_args()

    wb = Workbook()
    build_revision(wb)
    build_group_overview(wb)
    build_business_flow(wb)
    build_components(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"基本設計書テンプレート生成完了: {out}")


if __name__ == "__main__":
    main()
