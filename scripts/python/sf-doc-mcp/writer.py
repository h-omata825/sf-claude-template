# -*- coding: utf-8 -*-
"""Excel 定義書出力"""

from __future__ import annotations
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ #
# カラーパレット
# ------------------------------------------------------------------ #
C = {
    "cover_band_bg":   "243447",
    "cover_band_fg":   "FFFFFF",
    "cover_sub_bg":    "334D63",
    "cover_sub_fg":    "C9D8E8",
    "cover_idx_hd_bg": "3D5A73",
    "cover_idx_hd_fg": "FFFFFF",
    "cover_idx_bg":    "F7F9FB",
    "cover_idx_alt":   "EDF2F6",
    "sheet_title_bg":  "2E5266",
    "sheet_title_fg":  "FFFFFF",
    "section_bg":      "E4EDF3",
    "section_fg":      "1C3448",
    "col_hd_bg":       "3D6B87",
    "col_hd_fg":       "FFFFFF",
    "label_bg":        "F0F4F7",
    "stripe_bg":       "F7FAFC",
    "custom_fg":       "8B4000",   # カスタム項目の強調色
    "white":           "FFFFFF",
    "black":           "1A1A1A",
    "border_inner":    "B8CCDA",
    "border_outer":    "6B93AD",
}

# ------------------------------------------------------------------ #
# 基本列幅（12列固定）
# ------------------------------------------------------------------ #
BASE_COL_WIDTHS = [5, 6, 20, 20, 12, 7, 6, 6, 26, 20, 16, 22]
MAX_COLS = len(BASE_COL_WIDTHS)  # 12

# セクションごとの spans（合計 = MAX_COLS = 12）
SECTION_SPANS = {
    "fields":           [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    "record_types":     [3, 5, 1, 3],
    "page_layouts":     [4, 5, 3],
    "lightning_pages":  [3, 4, 3, 2],
    "compact_layouts":  [3, 5, 4],
    "search_layouts":   [4, 8],
    "field_sets":       [3, 4, 5],
    "validation_rules": [2, 1, 3, 3, 2, 1],
    "lookup_filters":   [3, 3, 4, 2],
}

# ------------------------------------------------------------------ #
# スタイルヘルパー
# ------------------------------------------------------------------ #

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(fg: str = C["black"], bold: bool = False, size: int = 10,
          name: str = "游ゴシック", strike: bool = False) -> Font:
    return Font(color=fg, bold=bold, size=size, name=name, strike=strike)

def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin(color: str) -> Side: return Side(style="thin",   color=color)
def _medium(color: str) -> Side: return Side(style="medium", color=color)

def _bool_str(val) -> str:
    if val is True:  return "○"
    if val is False: return "×"
    return str(val) if val else ""

def _set_row_height(ws, row: int, h: float): ws.row_dimensions[row].height = h
def _set_col_width(ws, col: int, w: float):
    ws.column_dimensions[get_column_letter(col)].width = w
def _freeze(ws, row: int): ws.freeze_panes = f"A{row}"

# ------------------------------------------------------------------ #
# セクション共通描画
# ------------------------------------------------------------------ #

def _section_title(ws, row: int, label: str):
    _set_row_height(ws, row, 22)
    c = ws.cell(row=row, column=1, value=f"  {label}")
    c.fill = _fill(C["section_bg"])
    c.font = _font(C["section_fg"], bold=True, size=10)
    c.alignment = _align("left")
    c.border = Border(top=_medium(C["border_outer"]), bottom=_thin(C["border_inner"]),
                      left=_medium(C["border_outer"]), right=_medium(C["border_outer"]))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)


def _spanned_row(ws, row: int, values: list, spans: list,
                 is_header: bool = False, stripe: bool = False,
                 first_row: bool = False, last_row: bool = False,
                 center_cols: set = None, custom_fg_cols: set = None,
                 modified: bool = False, strike: bool = False):
    """spans に従い論理列を merge して描画。anchor の right border が merge の右辺に対応。"""
    if is_header:
        h = 20
    else:
        max_lines = max((str(v).count('\n') + 1 for v in values if v), default=1)
        h = 18 if max_lines == 1 else min(14 + 13 * max_lines, 60)
    _set_row_height(ws, row, h)
    center_cols    = center_cols    or set()
    custom_fg_cols = custom_fg_cols or set()

    bg = (C["col_hd_bg"] if is_header else
          (C["stripe_bg"] if stripe    else C["white"]))
    default_fg = C["col_hd_fg"] if is_header else C["black"]

    col = 1
    for ci, (val, span) in enumerate(zip(values, spans)):
        col_end  = col + span - 1
        is_first = ci == 0
        is_last  = ci == len(values) - 1
        if modified and not is_header:
            fg = "FF0000"
        elif ci in custom_fg_cols and not is_header:
            fg = C["custom_fg"]
        else:
            fg = default_fg

        anchor = ws.cell(row=row, column=col, value=str(val) if val is not None else "")
        anchor.fill      = _fill(bg)
        anchor.font      = _font(fg, bold=is_header, size=9, strike=strike)
        anchor.alignment = _align("center" if (is_header or ci in center_cols) else "left",
                                  wrap=not is_header)
        anchor.border = Border(
            top    = _medium(C["border_outer"]) if first_row else _thin(C["border_inner"]),
            bottom = _medium(C["border_outer"]) if (is_header or last_row) else _thin(C["border_inner"]),
            left   = _medium(C["border_outer"]) if is_first  else _thin(C["border_inner"]),
            right  = _medium(C["border_outer"]) if is_last   else _thin(C["border_inner"]),
        )
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col_end)
        col += span


def _write_section(ws, row: int, title: str, cols: list[str],
                   data: list[list], spans: list[int],
                   center_cols: set = None,
                   custom_fg_cols: set = None) -> int:
    _section_title(ws, row, title)
    row += 1
    _spanned_row(ws, row, cols, spans, is_header=True, first_row=True, last_row=True)
    row += 1
    if not data:
        _spanned_row(ws, row, ["（データなし）"] + [""] * (len(cols) - 1),
                     spans, last_row=True)
        return row + 1
    for idx, vals in enumerate(data):
        _spanned_row(ws, row, vals, spans,
                     stripe=idx % 2 == 1,
                     last_row=idx == len(data) - 1,
                     center_cols=center_cols,
                     custom_fg_cols=custom_fg_cols)
        row += 1
    return row


# ------------------------------------------------------------------ #
# メインクラス
# ------------------------------------------------------------------ #

class DefinitionWriter:

    def __init__(self, output_path: str,
                 system_name: str = "", author: str = "",
                 original_author: str = "", original_date: str = "",
                 history: list = None, current_version: str = "1.0",
                 diffs: dict = None, field_changes: dict = None,
                 old_objects: dict = None, is_major: bool = False,
                 meta_payload: dict = None):
        self._path            = output_path
        self._system_name     = system_name
        self._author          = author
        self._original_author = original_author or "—"
        self._original_date   = original_date or str(date.today())
        self._history         = history or []
        self._current_version = current_version
        self._current_major   = current_version.split(".")[0]  # "1", "2" ...
        self._diffs           = diffs or {}
        self._field_changes   = field_changes or {}
        self._old_objects     = old_objects or {}
        self._is_major        = is_major
        self._meta_payload    = meta_payload or {}
        self._wb              = Workbook()
        self._wb.remove(self._wb.active)

    def write(self, metadata_list: list[dict]):
        from meta_store import write_meta
        self._add_cover(metadata_list)
        self._add_revision_history()
        for meta in metadata_list:
            self._write_object(meta)
        write_meta(self._wb, self._meta_payload)
        self._wb.save(self._path)
        print(f"[保存] {self._path}")

    # ------------------------------------------------------------------ #
    # 表紙（コンパクト版）
    # ------------------------------------------------------------------ #

    def _add_cover(self, metadata_list: list[dict]):
        ws = self._wb.create_sheet("表紙")
        ws.sheet_view.showGridLines = False

        for ci, w in enumerate([2, 20, 20, 20, 20, 20, 20, 2], start=1):
            _set_col_width(ws, ci, w)

        today = str(date.today())

        # ── タイトルバンド（行1〜12）──
        for r in range(1, 13):
            _set_row_height(ws, r, 18)
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).fill = _fill(C["cover_band_bg"])

        for r in range(5, 9):
            _set_row_height(ws, r, 30)

        tc = ws.cell(row=5, column=2, value="オブジェクト項目定義書")
        tc.fill = _fill(C["cover_band_bg"])
        tc.font = _font(C["cover_band_fg"], bold=True, size=26)
        tc.alignment = _align("left", "bottom")
        ws.merge_cells("B5:G8")

        sc = ws.cell(row=9, column=2, value="Salesforce メタデータ定義書")
        sc.fill = _fill(C["cover_band_bg"])
        sc.font = _font(C["cover_sub_fg"], size=11)
        sc.alignment = _align("left", "top")
        ws.merge_cells("B9:G9")

        # アクセントライン（行13）
        _set_row_height(ws, 13, 5)
        for ci in range(1, 9):
            ws.cell(row=13, column=ci).fill = _fill(C["cover_sub_bg"])

        # ── ドキュメント情報表（行15〜19）──
        is_update = self._current_version != "1.0"
        meta_items = [
            ("システム名称", self._system_name or "—"),
            ("作成日",       self._original_date),
            ("作成者",       self._original_author or "—"),
            ("更新日",       today if is_update else ""),
            ("更新者",       self._author if is_update else ""),
        ]
        for idx, (lbl, val) in enumerate(meta_items):
            ri    = 15 + idx
            first = idx == 0
            last  = idx == len(meta_items) - 1
            _set_row_height(ws, ri, 20)

            lc = ws.cell(row=ri, column=4, value=lbl)
            lc.fill = _fill(C["label_bg"])
            lc.font = _font(C["section_fg"], bold=True, size=10)
            lc.alignment = _align("left")
            lc.border = Border(
                top    = _medium(C["border_outer"]) if first else _thin(C["border_inner"]),
                bottom = _medium(C["border_outer"]) if last  else _thin(C["border_inner"]),
                left   = _medium(C["border_outer"]),
                right  = _medium(C["border_outer"]),
            )
            ws.merge_cells(start_row=ri, start_column=4, end_row=ri, end_column=5)

            vc = ws.cell(row=ri, column=6, value=val)
            vc.fill = _fill(C["white"])
            vc.font = _font(size=10)
            vc.alignment = _align("left")
            vc.border = Border(
                top    = _medium(C["border_outer"]) if first else _thin(C["border_inner"]),
                bottom = _medium(C["border_outer"]) if last  else _thin(C["border_inner"]),
                left   = _thin(C["border_inner"]),
                right  = _medium(C["border_outer"]),
            )
            ws.merge_cells(start_row=ri, start_column=6, end_row=ri, end_column=7)

        # ── 収録オブジェクト一覧（行21〜）──
        IDX_START = 21
        _set_row_height(ws, IDX_START, 20)

        def _idx_cell(row, c1, c2, val, bg, first_col, last_col,
                      first_row=False, last_row=False):
            cell = ws.cell(row=row, column=c1, value=val)
            is_hd = bg == C["cover_idx_hd_bg"]
            cell.fill = _fill(bg)
            cell.font = _font(C["cover_idx_hd_fg"] if is_hd else C["black"],
                              bold=is_hd, size=9)
            cell.alignment = _align("center" if is_hd else "left")
            cell.border = Border(
                top    = _medium(C["border_outer"]) if first_row else _thin(C["border_inner"]),
                bottom = _medium(C["border_outer"]) if last_row  else _thin(C["border_inner"]),
                left   = _medium(C["border_outer"]) if first_col else _thin(C["border_inner"]),
                right  = _medium(C["border_outer"]) if last_col  else _thin(C["border_inner"]),
            )
            if c2 > c1:
                ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

        _idx_cell(IDX_START, 2, 4, "オブジェクト名", C["cover_idx_hd_bg"],
                  True, False, first_row=True, last_row=True)
        _idx_cell(IDX_START, 5, 7, "API参照名", C["cover_idx_hd_bg"],
                  False, True, first_row=True, last_row=True)

        for idx, meta in enumerate(metadata_list):
            obj   = meta["object_api_name"]
            label = meta.get("object_info", {}).get("label", obj)
            dr    = IDX_START + 1 + idx
            last  = idx == len(metadata_list) - 1
            bg    = C["cover_idx_alt"] if idx % 2 == 1 else C["cover_idx_bg"]
            _set_row_height(ws, dr, 18)
            _idx_cell(dr, 2, 4, label, bg, True,  False, last_row=last)
            _idx_cell(dr, 5, 7, obj,   bg, False, True,  last_row=last)

    # ------------------------------------------------------------------ #
    # 改版履歴シート
    # ------------------------------------------------------------------ #

    def _add_revision_history(self):
        ws = self._wb.create_sheet("改版履歴")
        ws.sheet_view.showGridLines = False

        for ci, w in enumerate(BASE_COL_WIDTHS, start=1):
            _set_col_width(ws, ci, w)

        _set_row_height(ws, 1, 26)
        tc = ws.cell(row=1, column=1, value="  改版履歴")
        tc.fill = _fill(C["sheet_title_bg"])
        tc.font = _font(C["sheet_title_fg"], bold=True, size=12)
        tc.alignment = _align("left", "center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAX_COLS)

        row = 3
        _section_title(ws, row, "改版履歴")
        row += 1

        # No.(1) | バージョン(1) | 改定日(2) | 改定箇所（シート）(2) | 改定内容概要(4) | 改定者(2)
        spans = [1, 1, 2, 2, 4, 2]
        cols  = ["No.", "バージョン", "改定日", "改定箇所（シート）",
                 "改定内容概要", "改定者"]
        _spanned_row(ws, row, cols, spans, is_header=True, first_row=True, last_row=True,
                     center_cols={0, 1})
        row += 1

        if not self._history:
            _spanned_row(ws, row, ["（データなし）"] + [""] * 5, spans, last_row=True)
            return

        for idx, entry in enumerate(self._history):
            vals = [
                entry.get("no", ""),
                entry.get("version", ""),
                entry.get("date", ""),
                entry.get("sheet", ""),
                entry.get("content", ""),
                entry.get("author", ""),
            ]
            _spanned_row(ws, row, vals, spans,
                         stripe=idx % 2 == 1,
                         last_row=idx == len(self._history) - 1,
                         center_cols={0, 1})
            row += 1

    # ------------------------------------------------------------------ #
    # オブジェクトシート
    # ------------------------------------------------------------------ #

    def _write_object(self, meta: dict):
        obj   = meta["object_api_name"]
        label = meta.get("object_info", {}).get("label", obj)
        ws    = self._wb.create_sheet(f"{label}"[:31])
        ws.sheet_view.showGridLines = False

        # 新規追加オブジェクトはタブ色を赤にする
        if self._diffs.get(obj, {}).get("new_object"):
            ws.sheet_properties.tabColor = "FF0000"

        for ci, w in enumerate(BASE_COL_WIDTHS, start=1):
            _set_col_width(ws, ci, w)

        _set_row_height(ws, 1, 26)
        tc = ws.cell(row=1, column=1, value=f"  {label}  /  {obj}")
        tc.fill = _fill(C["sheet_title_bg"])
        tc.font = _font(C["sheet_title_fg"], bold=True, size=12)
        tc.alignment = _align("left", "center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAX_COLS)
        _freeze(ws, 2)

        row = 3
        obj_diff   = self._diffs.get(obj, {})
        is_new_obj = obj_diff.get("new_object", False)
        # 新規追加オブジェクトはタブ色のみ赤。セル内容は差分なしとして黒字で描画する
        field_diffs = None if is_new_obj else obj_diff.get("fields")

        if "object_info" in meta:
            row = self._write_object_info(ws, meta["object_info"], row)
            row += 2

        if "fields" in meta:
            row = self._write_fields(ws, obj, meta["fields"], row,
                                     field_diffs=field_diffs,
                                     field_usage=meta.get("field_usage", {}))
            row += 2

        if "record_types" in meta:
            row = self._write_record_types(ws, meta["record_types"], row)
            row += 2

        if "page_layouts" in meta:
            row = self._write_page_layouts(ws, meta["page_layouts"], row)
            row += 2

        if "lightning_pages" in meta:
            data = [[p.get("developer_name",""), p.get("label",""),
                     p.get("type",""), p.get("id","")]
                    for p in meta["lightning_pages"]]
            row = _write_section(ws, row, "Lightning ページ（FlexiPage）",
                                 ["開発者名","表示ラベル","タイプ","ID"],
                                 data, SECTION_SPANS["lightning_pages"])
            row += 2

        if "compact_layouts" in meta:
            data = [[cl.get("name",""), cl.get("label",""),
                     ", ".join(cl.get("fields",[]))]
                    for cl in meta["compact_layouts"]]
            row = _write_section(ws, row, "コンパクトレイアウト",
                                 ["名前","ラベル","表示項目"],
                                 data, SECTION_SPANS["compact_layouts"])
            row += 2

        if "search_layouts" in meta:
            row = self._write_search_layouts(ws, meta["search_layouts"], row)
            row += 2

        if "field_sets" in meta:
            data = [[fs.get("developer_name",""), fs.get("label",""),
                     fs.get("description","")]
                    for fs in meta["field_sets"]]
            row = _write_section(ws, row, "項目セット",
                                 ["開発者名","ラベル","説明"],
                                 data, SECTION_SPANS["field_sets"])
            row += 2

        if "validation_rules" in meta:
            row = self._write_validation_rules(ws, meta["validation_rules"], row)
            row += 2

        if "lookup_filters" in meta:
            data = [[lf.get("field_api_name",""), lf.get("field_label",""),
                     ", ".join(lf.get("controlling_fields",[])),
                     _bool_str(lf.get("optional_filter"))]
                    for lf in meta["lookup_filters"]]
            _write_section(ws, row, "参照条件（ルックアップフィルター）",
                           ["項目API名","項目ラベル","制御項目","任意フィルター"],
                           data, SECTION_SPANS["lookup_filters"], center_cols={3})

    # ------------------------------------------------------------------ #
    # オブジェクト基本情報（2列並び）
    # ------------------------------------------------------------------ #

    def _write_object_info(self, ws, info: dict, row: int) -> int:
        _section_title(ws, row, "オブジェクト基本情報")
        row += 1

        ZONES = [(1, 3, True, False), (4, 6, False, False),
                 (7, 9, False, False), (10, 12, False, True)]

        SHARING_LABEL = {
            "ReadWrite":          "読み取り/書き込み",
            "Read":               "読み取り専用",
            "Private":            "非公開",
            "ControlledByParent": "親によって制御",
            "FullAccess":         "フルアクセス",
            "ReadWriteTransfer":  "読み取り/書き込み/転送",
        }
        DEPLOY_LABEL = {
            "Deployed":      "リリース済み",
            "InDevelopment": "開発中",
        }

        # ── 基本情報（ラベル3列 + 値9列）──
        basic = [
            ("表示ラベル（単数形）",     info.get("label", "")),
            ("表示ラベル（複数形）",     info.get("label_plural", "")),
            ("API参照名",              info.get("api_name", "")),
            ("レコードIDプレフィックス", info.get("key_prefix", "")),
        ]
        for lbl, val in basic:
            _set_row_height(ws, row, 18)
            lc = ws.cell(row=row, column=1, value=lbl)
            lc.fill = _fill(C["label_bg"])
            lc.font = _font(C["section_fg"], bold=True, size=9)
            lc.alignment = _align("left")
            lc.border = Border(
                top    = _thin(C["border_inner"]),
                bottom = _thin(C["border_inner"]),
                left   = _medium(C["border_outer"]),
                right  = _thin(C["border_inner"]),  # 内側境界
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            vc = ws.cell(row=row, column=4, value=val)
            vc.fill = _fill(C["white"])
            vc.font = _font(size=9)
            vc.alignment = _align("left", wrap=True)
            vc.border = Border(
                top    = _thin(C["border_inner"]),
                bottom = _thin(C["border_inner"]),
                left   = _thin(C["border_inner"]),
                right  = _medium(C["border_outer"]),
            )
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=MAX_COLS)
            row += 1

        # ── ヘルパー ──
        def sub_header(label: str):
            nonlocal row
            _set_row_height(ws, row, 16)
            c = ws.cell(row=row, column=1, value=f"  {label}")
            c.fill = _fill(C["section_bg"])
            c.font = _font(C["section_fg"], bold=True, size=8)
            c.alignment = _align("left")
            c.border = Border(
                top    = _thin(C["border_inner"]),
                bottom = _thin(C["border_inner"]),
                left   = _medium(C["border_outer"]),
                right  = _medium(C["border_outer"]),
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)
            row += 1

        def pair_row(item1, item2=None, is_last: bool = False):
            """1〜2ペア（ラベル+値）を描画。item2=None の場合は値を col4-12 に展開"""
            nonlocal row
            _set_row_height(ws, row, 18)
            lbl1, val1 = item1
            bot = _medium(C["border_outer"]) if is_last else _thin(C["border_inner"])

            if item2 is None:
                # 単独アイテム: label(1-3) + value(4-12) 中央寄せ
                lc = ws.cell(row=row, column=1, value=lbl1)
                lc.fill = _fill(C["label_bg"])
                lc.font = _font(C["section_fg"], bold=True, size=9)
                lc.alignment = _align("left")
                lc.border = Border(top=_thin(C["border_inner"]), bottom=bot,
                                   left=_medium(C["border_outer"]), right=_thin(C["border_inner"]))
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

                vc = ws.cell(row=row, column=4, value=val1)
                vc.fill = _fill(C["white"])
                vc.font = _font(size=9)
                vc.alignment = _align("center")
                vc.border = Border(top=_thin(C["border_inner"]), bottom=bot,
                                   left=_thin(C["border_inner"]), right=_medium(C["border_outer"]))
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=MAX_COLS)
            else:
                # 2ペア: ZONES で描画
                lbl2, val2 = item2
                row_data  = [lbl1, val1, lbl2, val2]
                is_labels = [True, False, True, False]
                for (c1, c2, is_left, is_right), val, is_lbl in zip(ZONES, row_data, is_labels):
                    c = ws.cell(row=row, column=c1, value=val)
                    c.fill = _fill(C["label_bg"] if is_lbl else C["white"])
                    c.font = _font(C["section_fg"] if is_lbl else C["black"],
                                   bold=is_lbl, size=9)
                    c.alignment = _align("left" if is_lbl else "center")
                    c.border = Border(
                        top    = _thin(C["border_inner"]),
                        bottom = bot,
                        left   = _medium(C["border_outer"]) if is_left  else _thin(C["border_inner"]),
                        right  = _medium(C["border_outer"]) if is_right else _thin(C["border_inner"]),
                    )
                    if c2 > c1:
                        ws.merge_cells(start_row=row, start_column=c1,
                                       end_row=row, end_column=c2)
            row += 1

        def wide_row(label: str, val: str, is_last: bool = False):
            """ラベル(1-3) + 値(4-12) の1行"""
            nonlocal row
            _set_row_height(ws, row, 18)
            lc = ws.cell(row=row, column=1, value=label)
            lc.fill = _fill(C["label_bg"])
            lc.font = _font(C["section_fg"], bold=True, size=9)
            lc.alignment = _align("left")
            lc.border = Border(
                top    = _thin(C["border_inner"]),
                bottom = _medium(C["border_outer"]) if is_last else _thin(C["border_inner"]),
                left   = _medium(C["border_outer"]),
                right  = _thin(C["border_inner"]),
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            vc = ws.cell(row=row, column=4, value=val)
            vc.fill = _fill(C["white"])
            vc.font = _font(size=9)
            vc.alignment = _align("left", wrap=True)
            vc.border = Border(
                top    = _thin(C["border_inner"]),
                bottom = _medium(C["border_outer"]) if is_last else _thin(C["border_inner"]),
                left   = _thin(C["border_inner"]),
                right  = _medium(C["border_outer"]),
            )
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=MAX_COLS)
            row += 1

        # ── 追加の機能（describe API から確実に取得）──
        sub_header("追加の機能")
        pair_row(("Chatterフィード",      _bool_str(info.get("feed_enabled"))),
                 ("項目履歴管理",          _bool_str(info.get("track_history"))),
                 is_last=True)

        # ── オブジェクト分類 ──
        sub_header("オブジェクト分類")
        sm = info.get("sharing_model") or ""
        enable_sharing = (None   if not sm else
                          False  if sm == "ControlledByParent" else
                          True)
        pair_row(("共有を許可",           _bool_str(enable_sharing)),
                 ("BulkAPIアクセスを許可", _bool_str(info.get("bulk_api"))),
                 is_last=True)

        # ── リリース状況 ──
        sub_header("リリース状況")
        ds = info.get("deployment_status") or ""
        is_deployed = (True if ds == "Deployed" else (False if ds == "InDevelopment" else None))
        pair_row(("リリース状況", _bool_str(is_deployed)),
                 is_last=True)

        # ── 検索状況 ──
        sub_header("検索状況")
        pair_row(("検索を許可", _bool_str(info.get("searchable"))),
                 is_last=True)

        return row

    # ------------------------------------------------------------------ #
    # 項目定義（No. + 標準/カスタム付き）
    # ------------------------------------------------------------------ #

    def _write_fields(self, ws, obj_api_name: str, fields: list[dict], row: int,
                      field_diffs: dict = None, field_usage: dict = None) -> int:
        cols  = ["No.", "標準", "表示ラベル", "API参照名", "データ型",
                 "桁数", "必須", "一意", "数式 / デフォルト値",
                 "選択リスト値（上位5件）", "参照先", "ヘルプテキスト"]
        spans = SECTION_SPANS["fields"]
        field_usage = field_usage or {}

        # 利用箇所列（col13固定）— カスタム項目のみ使用
        USAGE_COL = MAX_COLS + 1  # 13
        _set_col_width(ws, USAGE_COL, 30)

        # 現バージョンの差分セット
        added_set    = set((field_diffs or {}).get("added",    []))
        modified_set = set((field_diffs or {}).get("modified", []))
        removed_set  = set((field_diffs or {}).get("removed",  []))

        # 前バージョンのフィールドデータ（削除フィールド表示用）
        old_fields_data: dict[str, dict] = {}
        if self._old_objects:
            for f in self._old_objects.get(obj_api_name, {}).get("fields", []):
                old_fields_data[f["api_name"]] = f

        # 注記列の準備（メジャー更新時はスキップ）
        obj_changes = {} if self._is_major else self._field_changes.get(obj_api_name, {})

        seen: list[str] = []
        for changes_list in obj_changes.values():
            for c in changes_list:
                if c["version"] not in seen:
                    seen.append(c["version"])
        all_versions = sorted(seen, key=lambda v: [int(x) for x in v.split(".")])
        # バージョンごとに1列ずつ右にずらす（14, 15, 16, ...）※13は利用箇所列
        version_col_map = {v: MAX_COLS + 2 + i for i, v in enumerate(all_versions)}

        for col in version_col_map.values():
            _set_col_width(ws, col, 12)

        # セクションタイトル
        _section_title(ws, row, f"項目定義  （全 {len(fields)} 件）")
        row += 1

        # ヘッダー行（注記列のヘッダーは不要）
        _spanned_row(ws, row, cols, spans, is_header=True, first_row=True, last_row=True)
        # 利用箇所列ヘッダー
        uh = ws.cell(row=row, column=USAGE_COL, value="利用箇所")
        uh.fill = _fill(C["col_hd_bg"])
        uh.font = _font(C["col_hd_fg"], bold=True, size=9)
        uh.alignment = _align("center", "center")
        uh.border = Border(
            top=_medium(C["border_outer"]), bottom=_medium(C["border_outer"]),
            left=_medium(C["border_outer"]), right=_medium(C["border_outer"]),
        )
        row += 1

        if not fields and not removed_set:
            _spanned_row(ws, row, ["（データなし）"] + [""] * 11, spans, last_row=True)
            return row + 1

        total_rows = len(fields) + len(removed_set)

        def _annotate(row_idx: int, api_name: str):
            """注記列に 'ver{version} {author}' を書き込む"""
            for change in obj_changes.get(api_name, []):
                col = version_col_map.get(change["version"])
                if col is None:
                    continue
                # 同一メジャーバージョン内の変更は赤、過去メジャーは黒
                color = "FF0000" if change["version"].split(".")[0] == self._current_major else C["black"]
                cc = ws.cell(row=row_idx, column=col,
                             value=f"ver{change['version']} {change['author']}")
                cc.font = _font(color, size=8)
                cc.alignment = _align("center")

        # 現行フィールドを描画
        for idx, f in enumerate(fields):
            pl_vals   = f.get("picklist_values", [])
            pl        = ", ".join(pl_vals[:5]) + (" …" if len(pl_vals) > 5 else "")
            formula   = f.get("formula") or f.get("default_value") or ""
            is_custom = f.get("custom", False)
            api_name  = f.get("api_name", "")
            is_changed = (not self._is_major) and (
                api_name in added_set or api_name in modified_set)
            is_last = (idx == total_rows - 1)
            vals = [
                idx + 1,
                "" if is_custom else "○",
                f.get("label", ""),
                api_name,
                f.get("data_type", ""),
                f.get("length", "") or "",
                _bool_str(f.get("required")),
                _bool_str(f.get("unique")),
                formula,
                pl,
                ", ".join(f.get("reference_to", [])),
                f.get("help_text", ""),
            ]
            _spanned_row(ws, row, vals, spans,
                         stripe=idx % 2 == 1,
                         last_row=is_last,
                         center_cols={0, 1, 6, 7},
                         custom_fg_cols={1, 2} if is_custom else set(),
                         modified=is_changed)
            # 利用箇所列（全行に枠線、カスタム項目のみ値を入れる）
            uc = ws.cell(row=row, column=USAGE_COL,
                         value=field_usage.get(api_name, "") if is_custom else "")
            uc.font = _font(C["black"], size=8)
            uc.alignment = Alignment(wrap_text=True, vertical="top")
            uc.fill = _fill(C["stripe_bg"]) if idx % 2 == 1 else _fill(C["white"])
            uc.border = Border(
                top    = _medium(C["border_outer"]) if idx == 0 else _thin(C["border_inner"]),
                bottom = _medium(C["border_outer"]) if is_last  else _thin(C["border_inner"]),
                left   = _medium(C["border_outer"]),
                right  = _medium(C["border_outer"]),
            )
            _annotate(row, api_name)
            row += 1

        # 削除フィールドを赤字＋取り消し線で追記（メジャー更新時はスキップ）
        if not self._is_major:
            removed_list = sorted(removed_set)
            for ridx, api_name in enumerate(removed_list):
                f       = old_fields_data.get(api_name, {"api_name": api_name})
                pl_vals = f.get("picklist_values", [])
                pl      = ", ".join(pl_vals[:5]) + (" …" if len(pl_vals) > 5 else "")
                formula = f.get("formula") or f.get("default_value") or ""
                is_custom = f.get("custom", False)
                is_last   = ridx == len(removed_list) - 1
                vals = [
                    "",   # No. は空
                    "" if is_custom else "○",
                    f.get("label", ""),
                    api_name,
                    f.get("data_type", ""),
                    f.get("length", "") or "",
                    _bool_str(f.get("required")),
                    _bool_str(f.get("unique")),
                    formula,
                    pl,
                    ", ".join(f.get("reference_to", [])),
                    f.get("help_text", ""),
                ]
                _spanned_row(ws, row, vals, spans,
                             last_row=is_last,
                             center_cols={0, 1, 6, 7},
                             modified=True, strike=True)
                # 削除行の利用箇所列（枠線のみ）
                uc = ws.cell(row=row, column=USAGE_COL, value="")
                uc.border = Border(
                    top    = _thin(C["border_inner"]),
                    bottom = _medium(C["border_outer"]) if is_last else _thin(C["border_inner"]),
                    left   = _medium(C["border_outer"]),
                    right  = _medium(C["border_outer"]),
                )
                _annotate(row, api_name)
                row += 1

        return row

    # ------------------------------------------------------------------ #
    # レコードタイプ
    # ------------------------------------------------------------------ #

    def _write_record_types(self, ws, record_types: list[dict], row: int) -> int:
        data = [[rt.get("developer_name",""), rt.get("name",""),
                 _bool_str(rt.get("is_active")), rt.get("description","")]
                for rt in record_types]
        return _write_section(ws, row, "レコードタイプ",
                              ["開発者名","表示ラベル","有効","説明"],
                              data, SECTION_SPANS["record_types"], center_cols={2})

    # ------------------------------------------------------------------ #
    # ページレイアウト（割り当て情報付き）
    # ------------------------------------------------------------------ #

    def _write_page_layouts(self, ws, layouts: list[dict], row: int) -> int:
        data = [[l.get("name",""), l.get("profiles",""), l.get("record_types","")]
                for l in layouts]
        return _write_section(ws, row, "ページレイアウト",
                              ["レイアウト名","割り当てプロファイル","割り当てレコードタイプ"],
                              data, SECTION_SPANS["page_layouts"])

    # ------------------------------------------------------------------ #
    # 検索レイアウト
    # ------------------------------------------------------------------ #

    def _write_search_layouts(self, ws, layouts: list, row: int) -> int:
        data = []
        for sl in (layouts if isinstance(layouts, list) else [layouts]):
            if isinstance(sl, dict):
                fields_str = ", ".join(
                    c.get("label", c.get("fieldApiName",""))
                    for c in sl.get("columns",[]) if isinstance(c, dict)
                )
                data.append([sl.get("layoutType",""), fields_str])
        return _write_section(ws, row, "検索レイアウト",
                              ["レイアウト種別","表示項目"],
                              data, SECTION_SPANS["search_layouts"])

    # ------------------------------------------------------------------ #
    # 入力規則（条件数式付き）
    # ------------------------------------------------------------------ #

    def _write_validation_rules(self, ws, rules: list[dict], row: int) -> int:
        cols = ["ルール名", "有効", "エラー条件数式", "エラーメッセージ", "エラー表示項目", "説明"]
        data = [[vr.get("name",""), _bool_str(vr.get("active")),
                 vr.get("condition_formula",""), vr.get("error_message",""),
                 vr.get("error_display_field",""), vr.get("description","")]
                for vr in rules]
        return _write_section(ws, row, "入力規則",
                              cols, data, SECTION_SPANS["validation_rules"],
                              center_cols={1})
