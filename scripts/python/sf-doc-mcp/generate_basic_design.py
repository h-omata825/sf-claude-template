# -*- coding: utf-8 -*-
"""
基本設計書.xlsx を1機能グループ分生成する（テンプレート読込方式）。

  基本設計書テンプレート.xlsx（build_basic_design_template.py で生成した「器」）を
  コピーしてセル値を流し込む。

4シート構成:
  1. 改版履歴                   : メタ + 履歴テーブル
  2. グループ概要               : 業務目的 / 対象ユーザー / 利用シーン / 前提条件
  3. 業務フロー                 : No / 担当 / 操作・処理内容 / 関連コンポーネント
  4. コンポーネント・オブジェクト: 構成コンポーネント / 関連オブジェクト / 外部連携

Usage:
  python generate_basic_design.py \\
    --input  basic_design.json \\
    --template "C:/.../基本設計書テンプレート.xlsx" \\
    --output-dir "C:/.../出力ルート"

出力先: {output-dir}/basic/【GRP-001】機能グループ名.xlsx

JSON スキーマ:
{
  "group_id": "GRP-001",
  "name_ja": "見積依頼",
  "name_en": "QuotationRequest",
  "project_name": "...",
  "author": "...",
  "version": "1.0",
  "date": "YYYY-MM-DD",
  "purpose": "業務目的",
  "target_users": "対象ユーザー",
  "usage_scene": "利用シーン",
  "business_flow": [
    {"step": "1", "actor": "担当", "action": "操作・処理内容", "system": "コンポーネント名"}
  ],
  "components": [
    {"api_name": "...", "type": "Apex/LWC/...", "role": "役割概要"}
  ],
  "related_objects": [
    {"api_name": "...", "label": "ラベル", "usage": "用途"}
  ],
  "external_integrations": [
    {"target": "連携先", "direction": "送信/受信", "data": "データ内容", "timing": "タイミング"}
  ],
  "prerequisites": "前提条件",
  "notes": "備考"
}
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date as _date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import design_revision as dr
from meta_store import read_meta, write_meta
from version_manager import increment_version

# ── 色定数 ─────────────────────────────────────────────────────────
C_HDR_BLUE  = "2E75B6"
C_BAND_BLUE = "0070C0"
C_LABEL_BG  = "D9E1F2"
C_FONT_D    = "000000"
C_FONT_W    = "FFFFFF"
C_FONT_GRAY = "595959"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

# ── テンプレートと一致させる定数 ────────────────────────────────────
# 改版履歴（build_basic_design_template.py と同一定義）
REV_META_ROW       = 3
REV_META_PROJECT_V = (6, 18)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6,  11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

# グループ概要（各ラベルの値セル）
GRP_META_ROW_1   = 3
GRP_META_ROW_2   = 4
GRP_META_1_V = {
    "project_name": (6,  16),
    "group_id":     (21, 24),
    "author":       (28, 29),
    "version":      (31, 31),
}
GRP_META_2_V = {
    "name_ja": (6,  24),
    "date":    (28, 31),
}
GRP_LABEL_VAL_CS = 7   # ラベル行の値開始列
GRP_LABEL_VAL_CE = 31
GRP_SECTION_ROW = {
    "purpose":      7,
    "target_users": 8,
    "usage_scene":  9,
    "prerequisites": 12,
    "notes":        13,
}

# 業務フロー（テンプレート定数と同一）
BF_DATA_ROW_START = 5
BF_STEP_CS,  BF_STEP_CE  = 2,  3
BF_ACTOR_CS, BF_ACTOR_CE = 4,  8
BF_ACT_CS,   BF_ACT_CE   = 9,  22
BF_SYS_CS,   BF_SYS_CE   = 23, 31

# コンポーネント・オブジェクト（テンプレート定数と同一）
CM_DATA_ROW_START = 5
CM_TYPE_CS, CM_TYPE_CE = 2,  5
CM_API_CS,  CM_API_CE  = 6,  14
CM_ROLE_CS, CM_ROLE_CE = 15, 31

OB_SEC_ROW        = 21
OB_HEAD_ROW       = 22
OB_DATA_ROW_START = 23
OB_API_CS,   OB_API_CE   = 2,  7
OB_LABEL_CS, OB_LABEL_CE = 8,  14
OB_USE_CS,   OB_USE_CE   = 15, 31

EX_SEC_ROW        = 34
EX_HEAD_ROW       = 35
EX_DATA_ROW_START = 36
EX_TGT_CS,  EX_TGT_CE  = 2,  7
EX_DIR_CS,  EX_DIR_CE  = 8,  10
EX_DATA_CS, EX_DATA_CE = 11, 21
EX_TIME_CS, EX_TIME_CE = 22, 31

GRID_RIGHT = 31

SCALAR_FIELDS  = ["purpose", "target_users", "usage_scene", "prerequisites", "notes"]
SECTION_SHEETS = {
    "business_flow":        "業務フロー",
    "components":           "コンポーネント・オブジェクト",
    "related_objects":      "コンポーネント・オブジェクト",
    "external_integrations": "コンポーネント・オブジェクト",
}


# ── スタイルヘルパー ────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h

def _data_row(ws, row, col_groups: list[tuple], row_h=22):
    """テンプレートに追加データ行が必要な場合に罫線付き空行を挿入する。"""
    set_h(ws, row, row_h)
    for cs, ce in col_groups:
        MW(ws, row, cs, ce, "", border=B_all())


# ── シート埋め込み ─────────────────────────────────────────────────
def fill_revision(ws, data: dict, history: list[dict]):
    vs, _ = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("project_name", ""))
    vs, _ = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("date", ""))
    dr.fill_revision_table(ws, history, REV_COLS, REV_DATA_ROW_START)


def fill_group_overview(ws, data: dict, changed_fields: set):
    # メタ行1
    for key, (cs, _) in GRP_META_1_V.items():
        ws.cell(row=GRP_META_ROW_1, column=cs, value=data.get(key, ""))
    # メタ行2
    for key, (cs, _) in GRP_META_2_V.items():
        ws.cell(row=GRP_META_ROW_2, column=cs, value=data.get(key, ""))
    # セクション本文（ラベル行の値セル）
    for key, row in GRP_SECTION_ROW.items():
        val = data.get(key, "")
        if val:
            cell = ws.cell(row=row, column=GRP_LABEL_VAL_CS, value=val)
            if key in changed_fields:
                dr.apply_red(cell, size=10)


def fill_business_flow(ws, data: dict, changed_step_nos: set):
    flows = data.get("business_flow", [])
    r = BF_DATA_ROW_START
    col_groups = [(BF_STEP_CS, BF_STEP_CE), (BF_ACTOR_CS, BF_ACTOR_CE),
                  (BF_ACT_CS, BF_ACT_CE), (BF_SYS_CS, BF_SYS_CE)]

    for i, flow in enumerate(flows):
        step_no = flow.get("step", str(i + 1))
        is_changed = step_no in changed_step_nos
        set_h(ws, r, 24)
        c1 = MW(ws, r, BF_STEP_CS,  BF_STEP_CE,  step_no,
                border=B_all(), h="center")
        c2 = MW(ws, r, BF_ACTOR_CS, BF_ACTOR_CE, flow.get("actor", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, BF_ACT_CS,   BF_ACT_CE,   flow.get("action", ""),
                border=B_all(), wrap=True, v="top")
        c4 = MW(ws, r, BF_SYS_CS,   BF_SYS_CE,   flow.get("system", ""),
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3, c4):
                dr.apply_red(c)
        r += 1

    # データが15行（テンプレート定数）を超えた場合は追加行を描画
    template_end = BF_DATA_ROW_START + 14
    if r > template_end + 1:
        pass  # 既に書き込み済み（テンプレート行は上書き）


def fill_components(ws, data: dict, changed_comp_keys: set, changed_obj_keys: set):
    # ── 構成コンポーネント ──
    components = data.get("components", [])
    r = CM_DATA_ROW_START
    col_groups_cm = [(CM_TYPE_CS, CM_TYPE_CE), (CM_API_CS, CM_API_CE),
                     (CM_ROLE_CS, CM_ROLE_CE)]
    for comp in components:
        key = comp.get("api_name", "")
        is_changed = key in changed_comp_keys
        set_h(ws, r, 22)
        c1 = MW(ws, r, CM_TYPE_CS, CM_TYPE_CE, comp.get("type", ""),
                border=B_all(), h="center")
        c2 = MW(ws, r, CM_API_CS,  CM_API_CE,  key,
                border=B_all())
        c3 = MW(ws, r, CM_ROLE_CS, CM_ROLE_CE, comp.get("role", ""),
                border=B_all(), wrap=True, v="top")
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # ── 関連オブジェクト ──
    objects = data.get("related_objects", [])
    r = OB_DATA_ROW_START
    for obj in objects:
        key = obj.get("api_name", "")
        is_changed = key in changed_obj_keys
        set_h(ws, r, 22)
        c1 = MW(ws, r, OB_API_CS,   OB_API_CE,   key,
                border=B_all())
        c2 = MW(ws, r, OB_LABEL_CS, OB_LABEL_CE, obj.get("label", ""),
                border=B_all(), h="center")
        c3 = MW(ws, r, OB_USE_CS,   OB_USE_CE,   obj.get("usage", ""),
                border=B_all(), wrap=True)
        if is_changed:
            for c in (c1, c2, c3):
                dr.apply_red(c)
        r += 1

    # ── 外部連携 ──
    integrations = data.get("external_integrations", [])
    r = EX_DATA_ROW_START
    for itg in integrations:
        set_h(ws, r, 22)
        MW(ws, r, EX_TGT_CS,  EX_TGT_CE,  itg.get("target", ""),    border=B_all())
        MW(ws, r, EX_DIR_CS,  EX_DIR_CE,  itg.get("direction", ""),  border=B_all(), h="center")
        MW(ws, r, EX_DATA_CS, EX_DATA_CE, itg.get("data", ""),       border=B_all(), wrap=True)
        MW(ws, r, EX_TIME_CS, EX_TIME_CE, itg.get("timing", ""),     border=B_all(), wrap=True)
        r += 1


# ── 差分計算 ────────────────────────────────────────────────────────
def _compute_diffs(prev_data: dict | None, new_data: dict) -> dict:
    if prev_data is None:
        return {"scalars": [], "lists": {}}
    return {
        "scalars": dr.diff_scalars(prev_data, new_data, SCALAR_FIELDS),
        "lists": {
            "business_flow":        dr.diff_list(
                prev_data.get("business_flow", []),
                new_data.get("business_flow", []), "step"),
            "components":           dr.diff_list(
                prev_data.get("components", []),
                new_data.get("components", []), "api_name"),
            "related_objects":      dr.diff_list(
                prev_data.get("related_objects", []),
                new_data.get("related_objects", []), "api_name"),
            "external_integrations": dr.diff_list(
                prev_data.get("external_integrations", []),
                new_data.get("external_integrations", []), "target"),
        },
    }


# ── メイン ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="基本設計書 Excel 生成")
    parser.add_argument("--input",      required=True, help="基本設計 JSON ファイルパス")
    parser.add_argument("--template",   required=True, help="基本設計書テンプレート.xlsx パス")
    parser.add_argument("--output-dir", required=True, help="出力先ディレクトリ")
    parser.add_argument("--source-file", default="",
                        help="更新時: 既存の基本設計書xlsxパス")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"])
    parser.add_argument("--source-hash", default="",
                        help="ソースファイルの SHA256 ハッシュ（source_hash_checker.py の出力値）")
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    today  = _date.today().strftime("%Y-%m-%d")
    author = data.get("author", "")

    group_id = data.get("group_id", "GRP-000")
    name_ja  = data.get("name_ja", "機能グループ")
    safe_name = re.sub(r'[\\/:*?"<>|]', "_", name_ja)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"【{group_id}】{safe_name}.xlsx"

    # ── バージョン判定 ────────────────────────────────────────────
    source_file = args.source_file.strip()
    if not source_file:
        existing = sorted(out_dir.glob(f"【{group_id}】*.xlsx"),
                          key=lambda f: f.stat().st_mtime, reverse=True)
        if existing:
            source_file = str(existing[0])
            print(f"  [AUTO] 既存ファイルを検出: {existing[0].name}")

    prev_meta   = read_meta(source_file) if source_file else None
    prev_data   = prev_meta.get("data") if prev_meta else None

    if prev_meta:
        current_version = increment_version(
            prev_meta.get("version", "1.0"), args.version_increment)
        history    = prev_meta.get("history", [])
        is_initial = False
        print(f"更新モード: {prev_meta.get('version', '?')} → {current_version}")
    else:
        current_version = data.get("version") or "1.0"
        history    = []
        is_initial = True
        print(f"新規作成モード: v{current_version}")

    data["version"] = current_version
    if not data.get("date"):
        data["date"] = today

    diffs = _compute_diffs(prev_data, data)
    if prev_meta and args.version_increment == "minor" and not dr.has_any_diff(diffs):
        print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    last_no = max((h["項番"] for h in history
                   if isinstance(h.get("項番"), int)), default=0)
    new_entries = dr.build_entries(
        current_version, diffs, author, today,
        start_no=last_no + 1,
        is_major=(args.version_increment == "major"),
        is_initial=is_initial,
        section_sheet_map=SECTION_SHEETS,
        scalar_sheet="グループ概要",
    )
    history = history + new_entries

    changed_scalars   = dr.changed_scalar_fields(diffs)
    changed_flows     = dr.changed_ids(diffs, "business_flow")
    changed_comps     = dr.changed_ids(diffs, "components")
    changed_objs      = dr.changed_ids(diffs, "related_objects")
    is_major          = (args.version_increment == "major")

    # ── テンプレ読込 → セル流し込み ──────────────────────────────
    wb = load_workbook(args.template)
    fill_revision(
        wb["改版履歴"], data, history)
    fill_group_overview(
        wb["グループ概要"], data,
        changed_fields=set() if is_major else changed_scalars)
    fill_business_flow(
        wb["業務フロー"], data,
        changed_step_nos=set() if is_major else changed_flows)
    fill_components(
        wb["コンポーネント・オブジェクト"], data,
        changed_comp_keys=set() if is_major else changed_comps,
        changed_obj_keys=set() if is_major else changed_objs)

    meta_payload = {
        "version": current_version,
        "date":    today,
        "author":  author,
        "data":    data,
        "history": history,
    }
    if args.source_hash:
        meta_payload["source_hash"] = args.source_hash
    write_meta(wb, meta_payload)

    wb.save(str(out_path))
    sys.stdout.buffer.write(
        f"[OK] 基本設計書を生成しました: v{current_version} → {out_path}\n".encode("utf-8"))

    # 同一IDで別名の旧ファイルを削除
    for old_f in out_dir.glob(f"【{group_id}】*.xlsx"):
        if old_f.resolve() != out_path.resolve():
            old_f.unlink()
            sys.stdout.buffer.write(
                f"  [CLEANUP] 旧ファイルを削除: {old_f.name}\n".encode("utf-8"))


if __name__ == "__main__":
    main()
