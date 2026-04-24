"""
画面設計書.xlsx を1画面分生成する（テンプレート読込方式）。

LWC / 画面フロー / Visualforce 共用。JSON の構造で UC・パラメータセクションを
動的に決定するため、画面種別ごとに適した切り口で記述できる。

5シート構成:
  1. 改版履歴       : メタ + 初版自動投入
  2. 画面概要       : メタ2段 + 目的/概要/主要機能(箇条書)/前提/画面遷移
  3. 画面項目定義   : 項目テーブル
  4. 処理詳細       : JSON の usecases[] をUCブロックとして順次描画
  5. パラメーター定義: JSON の param_sections[] をセクションとして順次描画

Usage:
  python generate_screen_design.py \
    --input     screen_design.json \
    --template  "C:/.../画面設計書テンプレート.xlsx" \
    --output-dir "C:/.../出力ルート"
"""
from __future__ import annotations
import argparse
import json
import sys
import tempfile
from datetime import date as _date
from pathlib import Path

# Windows cp932 環境で print 時の UnicodeEncodeError/文字化けを防止する
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

from meta_store import read_meta, write_meta
from tmp_utils import get_project_tmp_dir, set_project_tmp_dir
from version_manager import increment_version
import design_revision as dr

from flowchart_utils import generate_flowchart

# ── 色 ──────────────────────────────────────────────────────────
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_SUB_BAND   = "5B9BD5"
C_LABEL_BG   = "D9E1F2"
C_STEP_BG    = "E2EFDA"
C_SUB_BG     = "F2F2F2"
C_FONT_D     = "000000"
C_FONT_GRAY  = "595959"
C_FONT_W     = "FFFFFF"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

TYPE_FOLDER = {
    "LWC":          "lwc",
    "画面フロー":    "flow",
    "Flow":         "flow",
    "Visualforce":  "visualforce",
    "Aura":         "aura",
    "その他":        "other",
}

# ── テンプレートと一致させる定数 ────────────────────────────────
# 改版履歴
REV_META_ROW       = 3
REV_META_PROJECT_V = (7, 18)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6, 11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

# 画面概要
OV_META_ROW_1 = 3
OV_META_ROW_2 = 4
OV_META_1_V = {
    "project_name": (6, 14),
    "system_name":  (18, 22),
    "name":         (26, 31),
}
OV_META_2_V = {
    "api_name": (6, 14),
    "author":   (18, 20),
    "date":     (24, 26),
    "version":  (29, 31),
}
OV_SECTION_DATA_ROW = {
    "purpose":       7,
    "overview":      10,
    # "features" は箇条書き（bullet list）でデータ行13に書く
    "prerequisites": 16,
    "transition":    19,
}
OV_FEATURES_ROW = 13

# 画面項目定義
IT_DATA_ROW_START = 5
IT_COLS = {
    "no":          (2,  3),
    "label":       (4,  8),
    "api_name":    (9,  14),
    "ui_type":     (15, 17),
    "type":        (18, 20),
    "required":    (21, 22),
    "default":     (23, 25),
    "validation":  (26, 28),
    "note":        (29, 31),
}

# 処理詳細（動的UCブロック）— 機能設計書と同じ「左:表 / 右:フロー図」構成
LG_CONTENT_START_ROW = 3
# 左半分（ステップ詳細テーブル）
LG_LEFT_NO     = (2,  3)
LG_LEFT_TITLE  = (4,  7)
LG_LEFT_DETAIL = (8,  16)
LG_LEFT_END    = 16
# 右半分（フロー図の貼付領域）
LG_FLOW_CS     = 18
LG_FLOW_CE     = 31

# パラメーター定義（動的セクションブロック）
PM_CONTENT_START_ROW = 3
PM_COLS = {
    "no":       (2,  3),
    "key":      (4,  10),
    "type":     (11, 14),
    "required": (15, 16),
    "desc":     (17, 25),
    "default":  (26, 28),
    "note":     (29, 31),
}

GRID_LEFT  = 2
GRID_RIGHT = 31


# ── ヘルパー ───────────────────────────────────────────────────
def _fnt(bold=False, color=C_FONT_D, size=10, italic=False):
    return Font(name="游ゴシック", bold=bold, color=color,
                size=size, italic=italic)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _fill(c): return PatternFill("solid", fgColor=c)
def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, size=10, italic=False, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def _safe_unmerge(ws, row, cs, ce):
    """指定行・列範囲と重なる既存マージを先に解除してから再マージする。"""
    target = {(row, c) for c in range(cs, ce + 1)}
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= row <= mr.max_row and any(
            (row, c) in target
            for c in range(mr.min_col, mr.max_col + 1)
        ):
            ws.unmerge_cells(str(mr))

def _safe_merge_rows(ws, r1, r2, c1, c2):
    """r1〜r2 行・c1〜c2 列を縦結合する（既存 merge との衝突を自動解除）。"""
    if r1 >= r2:
        return
    target = {(r, c) for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)}
    for mr in list(ws.merged_cells.ranges):
        if any((r, c) in target
               for r in range(mr.min_row, mr.max_row + 1)
               for c in range(mr.min_col, mr.max_col + 1)):
            ws.unmerge_cells(str(mr))
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def MW(ws, row, cs, ce, value="", border=None, bg=None,
       bold=False, fg=C_FONT_D, h="left", v="center",
       wrap=True, size=10, italic=False):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    _safe_unmerge(ws, row, cs, ce)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws.cell(row=row, column=cs, value=value)
    cell.font = _fnt(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: cell.fill = _fill(bg)
    if border: cell.border = border
    return cell

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h




# ── 埋め込み ───────────────────────────────────────────────────
def fill_revision(ws, data, history: list[dict]):
    vs, ve = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("project_name", ""))
    vs, ve = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("date", ""))

    dr.fill_revision_table(ws, history, REV_COLS, REV_DATA_ROW_START)


def fill_overview(ws, data, changed_fields: set = None):
    changed_fields = changed_fields or set()
    for key, (cs, ce) in OV_META_1_V.items():
        ws.cell(row=OV_META_ROW_1, column=cs, value=data.get(key, ""))
    for key, (cs, ce) in OV_META_2_V.items():
        ws.cell(row=OV_META_ROW_2, column=cs, value=data.get(key, ""))

    # テキスト系セクション
    for key, r in OV_SECTION_DATA_ROW.items():
        cell = ws.cell(row=r, column=2, value=data.get(key, ""))
        if key in changed_fields:
            dr.apply_red(cell, size=10)

    # 主要機能（箇条書き）
    features = data.get("features", [])
    if features:
        bullet_text = "\n".join(f"・{f}" for f in features)
        cell = ws.cell(row=OV_FEATURES_ROW, column=2, value=bullet_text)
        if "features" in changed_fields:
            dr.apply_red(cell, size=10)

    # ── 追加セクション（transition(row 18-19) 以降に動的追記） ──
    extra_row = 21  # transition row 18-20 の後

    # 業務コンテキスト
    bctx = data.get("business_context", "")
    if bctx:
        ws.row_dimensions[extra_row].height = 18
        MW(ws, extra_row, 2, 31, "6. 業務コンテキスト（この画面が担う業務上の役割）",
           bold=True, bg=C_LABEL_BG, border=B_all())
        extra_row += 1
        ws.row_dimensions[extra_row].height = max(22, len(bctx) // 60 * 14 + 22)
        cell = ws.cell(row=extra_row, column=2, value=bctx)
        cell.alignment = _aln(wrap=True)
        if "business_context" in changed_fields:
            dr.apply_red(cell, size=10)
        extra_row += 1

    # 呼び出しApex一覧
    apex_calls = data.get("apex_calls", [])
    if apex_calls:
        ws.row_dimensions[extra_row].height = 18
        MW(ws, extra_row, 2, 31, "7. 呼び出しApex一覧",
           bold=True, bg=C_LABEL_BG, border=B_all())
        extra_row += 1
        # ヘッダ行
        ws.row_dimensions[extra_row].height = 18
        for (cs, ce), label in [
            ((2, 9), "Apex名"),
            ((10, 18), "呼び出し方式"),
            ((19, 25), "処理契機"),
            ((26, 31), "備考"),
        ]:
            MW(ws, extra_row, cs, ce, label,
               bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, border=B_all(), h="center")
        extra_row += 1
        for apex in apex_calls:
            ws.row_dimensions[extra_row].height = 22
            MW(ws, extra_row, 2,  9,  apex.get("name", ""),      border=B_all())
            MW(ws, extra_row, 10, 18, apex.get("operation", ""), border=B_all())
            MW(ws, extra_row, 19, 25, apex.get("trigger", ""),   border=B_all())
            MW(ws, extra_row, 26, 31, apex.get("note", ""),      border=B_all())
            extra_row += 1


def fill_items(ws, data, changed_item_nos: set = None):
    changed_item_nos = changed_item_nos or set()
    items = data.get("items", [])
    r = IT_DATA_ROW_START
    for i, it in enumerate(items):
        no = it.get("no", str(i + 1))
        is_changed = no in changed_item_nos
        long_text = (it.get("validation", "") or "") + " " + (it.get("note", "") or "")
        set_h(ws, r, max(24, min(80, len(long_text) // 25 * 14 + 24)))

        c1 = MW(ws, r, *IT_COLS["no"],         no,                              border=B_all(), h="center", v="top")
        c2 = MW(ws, r, *IT_COLS["label"],      it.get("label", ""),             border=B_all(), v="top", wrap=True)
        c3 = MW(ws, r, *IT_COLS["api_name"],   it.get("api_name", ""),          border=B_all(), v="top")
        c4 = MW(ws, r, *IT_COLS["ui_type"],    it.get("ui_type", ""),           border=B_all(), h="center", v="top")
        c5 = MW(ws, r, *IT_COLS["type"],       it.get("type", ""),              border=B_all(), h="center", v="top")
        c6 = MW(ws, r, *IT_COLS["required"],   "○" if it.get("required") else "", border=B_all(), h="center", v="top")
        c7 = MW(ws, r, *IT_COLS["default"],    it.get("default", ""),           border=B_all(), v="top", wrap=True)
        c8 = MW(ws, r, *IT_COLS["validation"], it.get("validation", ""),        border=B_all(), v="top", wrap=True)
        c9 = MW(ws, r, *IT_COLS["note"],       it.get("note", ""),              border=B_all(), v="top", wrap=True)

        if is_changed:
            for c in (c1, c2, c3, c4, c5, c6, c7, c8, c9):
                dr.apply_red(c)
        r += 1


def _step_title(step, index):
    t = step.get("title")
    if t:
        return t
    target = step.get("target", "")
    api = step.get("api_call", "")
    if target and api and api != "-":
        return f"{target} → {api}"
    return target or api or f"ステップ{index + 1}"


def _step_detail_text(step):
    parts = []
    detail = step.get("detail", "") or ""
    if detail:
        parts.append(detail)
    api = step.get("api_call", "") or ""
    if api and api != "-":
        parts.append(f"【API/アクション】{api}")
    result = step.get("result", "") or ""
    if result and result != "-":
        parts.append(f"【結果】{result}")
    note = step.get("note", "") or ""
    if note and note != "-":
        parts.append(f"【備考】{note}")
    return "\n".join(parts)


def fill_logic(ws, data, tmp_dir, changed_uc_titles: set = None):
    """usecases[] を「左:表 / 右:フロー図」の UC ブロックとして描画。"""
    changed_uc_titles = changed_uc_titles or set()
    usecases = data.get("usecases", [])
    r = LG_CONTENT_START_ROW
    for uc_idx, uc in enumerate(usecases):
        title = uc.get("title", "")
        is_changed = title in changed_uc_titles

        # UC 帯（全幅）
        set_h(ws, r, 26)
        uc_cell = MW(ws, r, GRID_LEFT, GRID_RIGHT, title,
           bold=True, fg=C_FONT_W, bg=C_SUB_BAND, size=11, border=B_all())
        if is_changed:
            dr.apply_red(uc_cell, bold=True, size=11)
        r += 1

        # トリガー行（全幅）
        trigger = uc.get("trigger", "")
        if trigger:
            set_h(ws, r, 24)
            MW(ws, r, 2, 4, "トリガー",
               bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
            MW(ws, r, 5, 31, trigger, border=B_all())
            r += 1

        # 左側テーブルヘッダ（No | 処理内容 の2列構成）
        table_header_row = r
        set_h(ws, r, 26)
        MW(ws, r, *LG_LEFT_NO,    "No",
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center", border=B_all())
        MW(ws, r, LG_LEFT_TITLE[0], LG_LEFT_END, "処理内容",
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center", border=B_all())
        r += 1

        steps = uc.get("steps", [])
        if not steps:
            set_h(ws, r, 26)
            MW(ws, r, LG_LEFT_NO[0], LG_LEFT_END, "（ステップ未登録）",
               fg=C_FONT_GRAY, italic=True, border=B_all(), h="center")
            r += 1
        else:
            TITLE_CPL     = 40
            DETAIL_CPL    = 40
            SUB_LABEL_CPL = 9
            SUB_DETAIL_CPL = 28
            for i, step in enumerate(steps):
                st_title    = _step_title(step, i)
                detail_text = _step_detail_text(step)
                t_lines = max(1, -(-len(st_title) // TITLE_CPL))
                d_lines = max(1, len(detail_text) // DETAIL_CPL + detail_text.count("\n") + 1) if detail_text else 1

                # タイトル行（緑背景・全幅）
                set_h(ws, r, max(20, t_lines * 16 + 4))
                no_row_start = r
                MW(ws, r, *LG_LEFT_NO, step.get("no", str(i + 1)),
                   bold=True, bg=C_STEP_BG, border=B_all(), h="center", v="center")
                MW(ws, r, LG_LEFT_TITLE[0], LG_LEFT_END, st_title,
                   bold=True, bg=C_STEP_BG, border=B_all(), v="center")
                r += 1

                # 詳細行（白背景・全幅）
                set_h(ws, r, max(22, d_lines * 16 + 4))
                for _col in range(LG_LEFT_NO[0], LG_LEFT_NO[1] + 1):
                    ws.cell(row=r, column=_col).border = B_all()
                    ws.cell(row=r, column=_col).fill = _fill(C_STEP_BG)
                MW(ws, r, LG_LEFT_TITLE[0], LG_LEFT_END, detail_text,
                   border=B_all(), wrap=True, v="top")
                r += 1

                # サブステップ（グレー背景・ラベル列 + 内容列）
                for sub in step.get("sub_steps", []):
                    sub_title  = sub.get("title", "")
                    sub_detail = sub.get("detail", "")
                    sub_no     = sub.get("no", "")
                    sub_label  = f"{sub_no} {sub_title}".strip() if sub_no else sub_title
                    sl_lines = max(1, -(-len(sub_label)  // SUB_LABEL_CPL))
                    sd_lines = max(1, len(sub_detail) // SUB_DETAIL_CPL + sub_detail.count("\n") + 1) if sub_detail else 1
                    set_h(ws, r, max(18, max(sl_lines, sd_lines) * 15 + 4))
                    for _col in range(LG_LEFT_NO[0], LG_LEFT_NO[1] + 1):
                        ws.cell(row=r, column=_col).border = B_all()
                        ws.cell(row=r, column=_col).fill = _fill(C_STEP_BG)
                    MW(ws, r, LG_LEFT_TITLE[0], LG_LEFT_DETAIL[0] - 1, sub_label,
                       bold=True, fg=C_FONT_GRAY, border=B_all(), bg=C_SUB_BG, v="top")
                    MW(ws, r, LG_LEFT_DETAIL[0], LG_LEFT_END, sub_detail,
                       fg=C_FONT_GRAY, border=B_all(), wrap=True, bg=C_SUB_BG, v="top")
                    r += 1

                # No列縦結合（タイトル行 + 詳細行 + サブステップ行）
                _safe_merge_rows(ws, no_row_start, r - 1, LG_LEFT_NO[0], LG_LEFT_NO[1])

        # 右側: フロー図（flow_steps 優先、なければ steps から生成）
        flow_steps = uc.get("flow_steps")
        if flow_steps is None:
            flow_steps = [
                {"no": s.get("no", str(i + 1)),
                 "title": _step_title(s, i),
                 "node_type": s.get("node_type", "process"),
                 "branch": s.get("branch"),
                 "calls": s.get("calls"),
                 "object_ref": s.get("object_ref"),
                 "main_label": s.get("main_label"),
                 "sub_steps": s.get("sub_steps", [])}
                for i, s in enumerate(steps)
            ]
        flow_path = Path(tmp_dir) / f"flow_uc{uc_idx + 1}.png"
        ok = False
        if flow_steps:
            try:
                ok = generate_flowchart(
                    flow_steps, str(flow_path),
                    fig_w=4.6, add_start_end=True, wrap_limit=14,
                )
            except Exception:
                ok = False

        img_h = 0
        if ok and flow_path.exists():
            try:
                img = XLImage(str(flow_path))
                aspect = img.height / img.width if img.width else 1.5
                img_w = 440
                img_h = int(img_w * aspect)   # 高さキャップなし・縦横比維持
                img.width = img_w
                img.height = img_h
                # 気持ち下にオフセット（テーブルヘッダ直上にくっつかないように）
                marker = AnchorMarker(
                    col=LG_FLOW_CS - 1, colOff=0,
                    row=table_header_row - 1, rowOff=pixels_to_EMU(18))
                size = XDRPositiveSize2D(
                    cx=pixels_to_EMU(img_w), cy=pixels_to_EMU(img_h))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
            except Exception as e:
                W(ws, table_header_row, LG_FLOW_CS,
                  f"[フロー図挿入失敗: {e}]", italic=True, fg="888888")
        elif flow_steps:
            # generate_flowchart() が False を返した / PNG が生成されなかった場合も明示する
            # （silent failure 防止。機能設計書側と整合）
            from flowchart_utils import HAS_MPL
            reason = "matplotlib 未インストール" if not HAS_MPL else "フロー図PNG生成失敗"
            W(ws, table_header_row, LG_FLOW_CS,
              f"[フロー図なし: {reason}]", italic=True, fg="BF0000")
            print(f"  [WARNING] UC{uc_idx + 1}: フロー図を挿入できませんでした: {reason}",
                  file=sys.stderr)

        # 画像が次UCへ貫通しないよう、テーブル累積高さが画像高さ未満なら
        # 罫線なしの空行でスペースを確保する
        if img_h > 0:
            need_px = img_h + 24  # 18pxオフセット + 余裕6px
            cum_pt = 0.0
            for rr in range(table_header_row, r):
                h = ws.row_dimensions[rr].height
                cum_pt += (h if h else 15)
            cum_px = cum_pt * 4.0 / 3.0
            if cum_px < need_px:
                deficit_pt = (need_px - cum_px) * 3.0 / 4.0
                spacer_rows = int(deficit_pt // 15) + 1
                for _ in range(spacer_rows):
                    set_h(ws, r, 15)
                    r += 1

        # UC 間スペーサー
        set_h(ws, r, 12)
        r += 1


def fill_params(ws, data, changed_params_map: dict = None,
                changed_section_titles: set = None):
    """param_sections[] を順番に描画。

    changed_params_map: {section_title: set(item_key,...)} 変更のあった項目
    changed_section_titles: セクション自体が追加/変更されたタイトル集合
    """
    changed_params_map    = changed_params_map or {}
    changed_section_titles = changed_section_titles or set()
    sections = data.get("param_sections", [])
    r = PM_CONTENT_START_ROW
    for sec in sections:
        title = sec.get("title", "")
        set_h(ws, r, 26)
        sec_cell = MW(ws, r, GRID_LEFT, GRID_RIGHT, title,
           bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11, border=B_all())
        if title in changed_section_titles:
            dr.apply_red(sec_cell, bold=True, size=11)
        r += 1

        headers = {
            "no": "No", "key": "Key / 名称", "type": "型", "required": "必須",
            "desc": "説明", "default": "デフォルト", "note": "備考",
        }
        set_h(ws, r, 26)
        for k, label in headers.items():
            cs, ce = PM_COLS[k]
            MW(ws, r, cs, ce, label,
               bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center",
               border=B_all())
        r += 1

        items = sec.get("items", [])
        if not items:
            set_h(ws, r, 24)
            MW(ws, r, GRID_LEFT, GRID_RIGHT, "（なし）",
               fg=C_FONT_GRAY, italic=True, border=B_all(),
               h="center")
            r += 1
        else:
            changed_keys = changed_params_map.get(title, set())
            for i, p in enumerate(items):
                desc = p.get("desc", "") or p.get("description", "") or ""
                set_h(ws, r, max(24, min(80, len(desc) // 30 * 14 + 24)))
                item_key = p.get("key", "")
                is_changed = item_key in changed_keys
                vals = {
                    "no":       p.get("no", str(i + 1)),
                    "key":      item_key,
                    "type":     p.get("type", ""),
                    "required": "○" if p.get("required") else "",
                    "desc":     desc,
                    "default":  p.get("default", ""),
                    "note":     p.get("note", ""),
                }
                row_cells = []
                for k, (cs, ce) in PM_COLS.items():
                    row_cells.append(MW(ws, r, cs, ce, vals[k], border=B_all(),
                                        v="top",
                                        h="center" if k in ("no", "required") else "left"))
                if is_changed:
                    for c in row_cells:
                        dr.apply_red(c)
                r += 1

        set_h(ws, r, 12)
        r += 1


SCALAR_FIELDS   = ["purpose", "overview", "features",
                   "prerequisites", "transition"]
SECTION_SHEETS  = {"items":          "画面項目定義",
                   "usecases":       "処理詳細",
                   "param_sections": "パラメーター定義",
                   "param_items":    "パラメーター定義"}


def _flatten_param_items(sections: list) -> list[dict]:
    """param_sections[].items[] を 複合キー "section::key" でフラット化。"""
    out = []
    for sec in sections or []:
        stitle = sec.get("title", "")
        for it in (sec.get("items") or []):
            out.append({
                "_key": f"{stitle}::{it.get('key', '')}",
                "_section": stitle,
                "_item_key": it.get("key", ""),
                **{k: v for k, v in it.items() if k not in ("_key", "_section", "_item_key")},
            })
    return out


def _compute_diffs(prev_data: dict | None, new_data: dict) -> dict:
    if prev_data is None:
        return {"scalars": [], "lists": {}}
    return {
        "scalars": dr.diff_scalars(prev_data, new_data, SCALAR_FIELDS),
        "lists": {
            "items":    dr.diff_list(prev_data.get("items", []),
                                     new_data.get("items", []), "no"),
            "usecases": dr.diff_list(prev_data.get("usecases", []),
                                     new_data.get("usecases", []), "title"),
            "param_sections": dr.diff_list(prev_data.get("param_sections", []),
                                           new_data.get("param_sections", []), "title"),
            "param_items":    dr.diff_list(
                _flatten_param_items(prev_data.get("param_sections", [])),
                _flatten_param_items(new_data.get("param_sections", [])),
                "_key"),
        },
    }


def _build_changed_params_map(diffs: dict) -> dict[str, set]:
    """変更のあった param_items の "section::key" を section→{key,...} に変換。"""
    changed = dr.changed_ids(diffs, "param_items")
    out: dict[str, set] = {}
    for ck in changed:
        if "::" in ck:
            sec, ikey = ck.split("::", 1)
            out.setdefault(sec, set()).add(ikey)
    return out


def _fill_events(wb, data: dict) -> None:
    """イベント定義シートをワークブックに追加する。events[] が空の場合はスキップ。"""
    events = data.get("events", [])
    if not events:
        return

    sheet_name = "イベント定義"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # 列幅設定
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14  # イベント名
    ws.column_dimensions["C"].width = 18  # 対象要素
    ws.column_dimensions["D"].width = 22  # ハンドラ/メソッド
    ws.column_dimensions["E"].width = 36  # 処理内容
    ws.column_dimensions["F"].width = 18  # 備考

    # タイトル行
    ws.row_dimensions[1].height = 26
    MW(ws, 1, 2, 6, "イベント定義",
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=12, h="center")

    # ヘッダ行
    ws.row_dimensions[2].height = 20
    for col, label in [
        (2, "イベント名"),
        (3, "対象要素"),
        (4, "ハンドラ / メソッド"),
        (5, "処理内容"),
        (6, "備考"),
    ]:
        c = ws.cell(row=2, column=col, value=label)
        c.font = _fnt(bold=True, color=C_FONT_W)
        c.alignment = _aln(h="center")
        c.fill = _fill(C_HDR_BLUE)
        c.border = B_all()

    # データ行
    for i, ev in enumerate(events):
        r = i + 3
        desc = ev.get("description", "")
        ws.row_dimensions[r].height = max(22, len(desc) // 36 * 14 + 22)
        for col, val in [
            (2, ev.get("event", "")),
            (3, ev.get("element", "")),
            (4, ev.get("handler", "")),
            (5, desc),
            (6, ev.get("note", "")),
        ]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = _fnt()
            c.alignment = _aln(wrap=True)
            c.border = B_all()
            if i % 2 == 1:
                c.fill = _fill("F2F2F2")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",      required=True)
    parser.add_argument("--template",   required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--source-file", default="",
                        help="更新時: 既存の画面設計書xlsxパス")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"])
    parser.add_argument("--source-hash", default="",
                        help="ソースのSHA256。_meta に保存して次回差分判定に使う")
    parser.add_argument("--author", default="",
                        help="作成者名。JSON の author が空の場合にフォールバックで使用")
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    today  = _date.today().strftime("%Y-%m-%d")
    # --author が指定された場合は JSON 側に反映（改版履歴の変更者と表紙の両方に効く）
    if args.author:
        data["author"] = args.author
    author = data.get("author", "")

    # ── 出力先を先に確定（既存ファイル自動検出に使用） ──────────────
    feat_id   = data.get("id", "F-000")
    name      = data.get("name", "画面")
    type_key  = data.get("type", "その他")
    subfolder = TYPE_FOLDER.get(type_key, "other")
    out_dir   = Path(args.output_dir) / subfolder
    out_dir.mkdir(parents=True, exist_ok=True)
    set_project_tmp_dir(out_dir)
    out_path  = out_dir / f"【{feat_id}】{name}.xlsx"

    # ── バージョン判定 ────────────────────────────────────────
    is_major    = (args.version_increment == "major")
    source_file = args.source_file.strip()

    # --source-file 未指定時: 同一IDの既存ファイルを自動検出
    if not source_file:
        existing = [f for f in out_dir.glob(f"【{feat_id}】*.xlsx")]
        if existing:
            source_file = str(existing[0])
            print(f"  [AUTO] 既存ファイルを検出: {existing[0].name}")

    prev_meta   = read_meta(source_file) if source_file else None
    prev_data   = prev_meta.get("data") if prev_meta else None

    # 第1ゲート: ソースハッシュ一致ならExcel再生成スキップ（詳細設計と同じ方式）
    if (prev_meta and args.source_hash
            and args.version_increment == "minor"
            and prev_meta.get("source_hash") == args.source_hash):
        print("差分なし: ソースハッシュが既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    if prev_meta:
        prev_history_len = len(prev_meta.get("history", []))
        # 改版履歴 20 行制限: 既存履歴が 20 以上なら minor 指定でも major に強制昇格し履歴リセット
        forced_major = False
        if prev_history_len >= 20 and args.version_increment == "minor":
            print(f"  [WARN] 改版履歴が {prev_history_len} 件に達しているため minor → major に強制昇格し、履歴をリセットします")
            args.version_increment = "major"
            is_major = True
            forced_major = True
        current_version = increment_version(
            prev_meta.get("version", "1.0"), args.version_increment)
        # major 時は履歴リセット（手動・強制問わず。メジャーUP 1行だけ残す）
        history    = [] if is_major else prev_meta.get("history", [])
        is_initial = False
        if forced_major:
            print(f"メジャー昇格モード（履歴リセット）: {prev_meta.get('version', '?')} → {current_version}")
        else:
            print(f"更新モード: {prev_meta.get('version', '?')} → {current_version}"
                  + (" (メジャー)" if is_major else ""))
    else:
        current_version = data.get("version") or "1.0"
        history    = []
        is_initial = True
        print(f"新規作成モード: v{current_version}")

    data["version"] = current_version

    diffs = _compute_diffs(prev_data, data)
    if prev_meta and not is_major and not dr.has_any_diff(diffs):
        print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    last_no = max((h["項番"] for h in history
                   if isinstance(h.get("項番"), int)), default=0)
    new_entries = dr.build_entries(
        current_version, diffs, author, today,
        start_no=last_no + 1, is_major=is_major, is_initial=is_initial,
        section_sheet_map=SECTION_SHEETS, scalar_sheet="画面概要",
    )
    history = history + new_entries

    # 変更識別子を抽出（赤字用）
    changed_scalars  = dr.changed_scalar_fields(diffs)
    changed_items    = dr.changed_ids(diffs, "items")
    changed_ucs      = dr.changed_ids(diffs, "usecases")
    changed_sections = dr.changed_ids(diffs, "param_sections")
    changed_params   = _build_changed_params_map(diffs)

    with tempfile.TemporaryDirectory(prefix="screen_design_", dir=get_project_tmp_dir()) as tmp_dir:
        wb = load_workbook(args.template)
        fill_revision(wb["改版履歴"],         data, history)
        fill_overview(wb["画面概要"],         data,
                      changed_fields=set() if is_major else changed_scalars)
        fill_items   (wb["画面項目定義"],     data,
                      changed_item_nos=set() if is_major else changed_items)
        fill_logic   (wb["処理詳細"],         data, tmp_dir,
                      changed_uc_titles=set() if is_major else changed_ucs)
        fill_params  (wb["パラメーター定義"], data,
                      changed_params_map=({} if is_major else changed_params),
                      changed_section_titles=(set() if is_major else changed_sections))
        _fill_events(wb, data)

        # _meta 保存（次回差分判定用）
        meta_payload = {
            "version": current_version,
            "date":    today,
            "author":  author,
            "data":    data,
            "history": history,
        }
        if args.source_hash:
            meta_payload["source_hash"] = args.source_hash
        write_meta(wb, meta_payload)

        wb.save(out_path)
        print(f"画面設計書生成完了: v{current_version} → {out_path}")

        # 同一IDで別名の旧ファイルを削除（名称変更時の二重ファイル防止）
        for old_f in out_dir.glob(f"【{feat_id}】*.xlsx"):
            if old_f.resolve() != out_path.resolve():
                old_f.unlink()
                print(f"  [CLEANUP] 旧ファイルを削除: {old_f.name}")


if __name__ == "__main__":
    main()
