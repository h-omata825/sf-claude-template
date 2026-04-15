"""
機能一覧テンプレート Excel を生成する。

3シート構成:
  1. 改版履歴           : 基本情報 + 改版履歴テーブル
  2. サマリー           : メタ情報 + 種別別件数テーブル
  3. __SHEET_TEMPLATE__ : 種別別シートの雛形（generate_feature_list.py が複製）

方針:
  - 設計書テンプレートと同じデザインシステム（狭幅グリッド + セル結合 + 罫線統一）
  - 結合前に全構成セルに罫線付与（途切れ防止）
  - テーブルヘッダに濃青バンド

Usage:
  python build_feature_list_template.py --output "C:/.../機能一覧テンプレート.xlsx"
"""
from __future__ import annotations
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── カラー ──────────────────────────────────────────────────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

# ── ヘルパー ───────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h
def set_w(ws, letter, w):
    ws.column_dimensions[letter].width = w


# ── 共通グリッド ───────────────────────────────────────────────
GRID_LEFT  = 2
GRID_RIGHT = 31
COL_W      = 4.2

def setup_grid(ws):
    set_w(ws, "A", 2.0)
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        set_w(ws, get_column_letter(i), COL_W)
    ws.sheet_view.showGridLines = False

def title_band(ws, row, text, height=36):
    set_h(ws, row, height)
    MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_TITLE_DARK, h="center", size=14)

def section_band(ws, row, text, height=26,
                 cs=None, ce=None):
    cs = cs if cs is not None else GRID_LEFT
    ce = ce if ce is not None else GRID_RIGHT
    set_h(ws, row, height)
    MW(ws, row, cs, ce, text,
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11)


# ─────────────────────────────────────────────────────────────
# Sheet 1: 改版履歴
# ─────────────────────────────────────────────────────────────
REV_META_ROW       = 3
REV_TABLE_HEAD_ROW = 5
REV_DATA_ROW_START = 6
REV_DATA_ROW_END   = 25

REV_META_PROJECT = (2, 6, 7, 18)
REV_META_DATE    = (19, 22, 23, 31)

REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6, 11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

def build_revision(wb):
    ws = wb.active
    ws.title = "改版履歴"
    setup_grid(ws)

    title_band(ws, 1, "改版履歴")
    set_h(ws, 2, 6)

    set_h(ws, REV_META_ROW, 22)
    ls, le, vs, ve = REV_META_PROJECT
    MW(ws, REV_META_ROW, ls, le, "プロジェクト名",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, REV_META_ROW, vs, ve, "", border=B_all())

    ls, le, vs, ve = REV_META_DATE
    MW(ws, REV_META_ROW, ls, le, "作成日",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, REV_META_ROW, vs, ve, "", border=B_all())

    set_h(ws, 4, 10)

    set_h(ws, REV_TABLE_HEAD_ROW, 24)
    for label, (cs, ce) in REV_COLS.items():
        MW(ws, REV_TABLE_HEAD_ROW, cs, ce, label,
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center",
           border=B_all())

    for r in range(REV_DATA_ROW_START, REV_DATA_ROW_END + 1):
        set_h(ws, r, 22)
        for label, (cs, ce) in REV_COLS.items():
            MW(ws, r, cs, ce, "", border=B_all())


# ─────────────────────────────────────────────────────────────
# Sheet 2: サマリー
# ─────────────────────────────────────────────────────────────
SUM_META_ROW_1 = 3
SUM_META_ROW_2 = 4

# メタ: プロジェクト名 / 作成者 / 作成日 / 合計件数
SUM_META_1 = [
    ("プロジェクト名", 2, 5,   6, 18),
    ("作成日",         19, 22, 23, 31),
]
SUM_META_2 = [
    ("作成者",   2,  5,   6, 12),
    ("バージョン", 13, 16, 17, 22),
    ("合計件数",  23, 25, 26, 31),
]

SUM_SEC_ROW  = 6
SUM_HEAD_ROW = 7
SUM_DATA_ROW_START = 8

# 種別別件数テーブル
SUM_COLS = {
    "No":         (2,  3),
    "種別":       (4,  11),
    "件数":       (12, 15),
    "対応シート": (16, 31),
}

def build_summary(wb):
    ws = wb.create_sheet("サマリー")
    setup_grid(ws)

    title_band(ws, 1, "機能一覧 サマリー")
    set_h(ws, 2, 6)

    set_h(ws, SUM_META_ROW_1, 22)
    for label, ls, le, vs, ve in SUM_META_1:
        MW(ws, SUM_META_ROW_1, ls, le, label,
           bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
        MW(ws, SUM_META_ROW_1, vs, ve, "", border=B_all())
    set_h(ws, SUM_META_ROW_2, 22)
    for label, ls, le, vs, ve in SUM_META_2:
        MW(ws, SUM_META_ROW_2, ls, le, label,
           bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
        MW(ws, SUM_META_ROW_2, vs, ve, "", border=B_all())


    set_h(ws, 5, 12)

    section_band(ws, SUM_SEC_ROW, "■ 種別別の件数")

    set_h(ws, SUM_HEAD_ROW, 26)
    for label, (cs, ce) in SUM_COLS.items():
        MW(ws, SUM_HEAD_ROW, cs, ce, label,
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center",
           border=B_all())
    # データ行は generate 側で動的に追加


# ─────────────────────────────────────────────────────────────
# Sheet 3: __SHEET_TEMPLATE__ (種別別シート雛形)
# ─────────────────────────────────────────────────────────────
# generate_feature_list.py の定数と一致させる
# ST_SEC_ROW=3, ST_HEAD_ROW=4, ST_DATA_ROW_START=5
ST_SEC_ROW  = 3
ST_HEAD_ROW = 4

ST_COLS = {
    "ID":              (2,  4),
    "API名/ファイル名": (5,  11),
    "機能名":          (12, 18),
    "処理概要":        (19, 31),
}

def build_sheet_template(wb):
    """種別別シートの雛形。generate 側で複製して種別名に差し替え。
    メタ行なし（サマリーシートのみに記載）。
    Row1: タイトル / Row2: spacer / Row3: セクション帯 / Row4: ヘッダ / Row5+: データ
    """
    ws = wb.create_sheet("__SHEET_TEMPLATE__")
    setup_grid(ws)

    title_band(ws, 1, "機能一覧 — {TYPE}")
    set_h(ws, 2, 6)

    section_band(ws, ST_SEC_ROW, "■ 機能一覧")

    set_h(ws, ST_HEAD_ROW, 26)
    for label, (cs, ce) in ST_COLS.items():
        MW(ws, ST_HEAD_ROW, cs, ce, label,
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center",
           border=B_all())


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    wb = Workbook()
    build_revision(wb)
    build_summary(wb)
    build_sheet_template(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"機能一覧テンプレート生成完了: {out}")


if __name__ == "__main__":
    main()
