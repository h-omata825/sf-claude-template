"""
詳細設計書テンプレート Excel を生成する。

7シート構成（T-04仕様）:
  1. 改版履歴           : 基本情報 + 改版履歴テーブル
  2. 概要               : 機能名 / 機能概要 / 目的 / 利用者 / 起点画面 / 操作トリガー
  3. 業務フロー         : No / アクター / 処理内容 / 分岐条件
  4. 対象オブジェクト   : オブジェクト名 / 項目API名 / 項目ラベル / 読み書き区分 / 備考
  5. 処理概要           : No / 処理内容 / コンポーネント / 分岐条件
  6. 関連コンポーネント : コンポーネント名 / 種別 / 役割 / 依存方向
  7. 影響範囲           : 5セクション（更新/参照オブジェクト、関連Apex等、外部連携、他機能依存）

方針:
  - 設計書テンプレートと同じデザインシステム（狭幅グリッド + セル結合 + 罫線統一）
  - グリッド: col A=2.0、col 2-31=4.2（30列）

Usage:
  python build_detail_design_template.py --output "C:/.../詳細設計書テンプレート.xlsx"
"""
from __future__ import annotations
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# -- カラー --------------------------------------------------------------
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

# -- ヘルパー -------------------------------------------------------------
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = _fill(bg)
    if border: c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h): ws.row_dimensions[row].height = h
def set_w(ws, letter, w): ws.column_dimensions[letter].width = w

# -- 共通グリッド ---------------------------------------------------------
GRID_LEFT  = 2
GRID_RIGHT = 31

def setup_grid(ws):
    set_w(ws, "A", 2.0)
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        set_w(ws, get_column_letter(i), 4.2)
    ws.sheet_view.showGridLines = False

def title_band(ws, row, text, height=36):
    set_h(ws, row, height)
    MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_TITLE_DARK, h="center", size=14)

def section_band(ws, row, text, height=26, cs=None, ce=None):
    cs = cs if cs is not None else GRID_LEFT
    ce = ce if ce is not None else GRID_RIGHT
    set_h(ws, row, height)
    MW(ws, row, cs, ce, text,
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11)

def table_header(ws, row, cols: list[tuple], height=26):
    """cols: [(cs, ce, label), ...]"""
    set_h(ws, row, height)
    for cs, ce, label in cols:
        MW(ws, row, cs, ce, label,
           bold=True, fg=C_FONT_W, bg=C_HDR_BLUE, h="center", border=B_all())

def data_rows(ws, row_start, row_end, col_groups: list[tuple], row_h=22):
    for r in range(row_start, row_end + 1):
        set_h(ws, r, row_h)
        for cs, ce in col_groups:
            MW(ws, r, cs, ce, "", border=B_all())

def label_row(ws, row, label, label_cs=2, label_ce=6, val_cs=7, val_ce=31, height=22):
    set_h(ws, row, height)
    MW(ws, row, label_cs, label_ce, label,
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, row, val_cs, val_ce, "", border=B_all())


def diagram_area(ws, start_row, title, section_no=2, height_rows=30):
    """図エリアを配置し、使用した行数を返す。

    構造:
      start_row    : セクションバンド「■ {section_no}. {title}」
      +1 ~ +height : 図エリア（col 2-31 マージ、bg=F2F2F2、罫線）
      +height+1    : スペーサー 8pt
    返値: 次のセクションが始まる行番号
    """
    # セクションバンド
    section_band(ws, start_row, f"\u25a0 {section_no}. {title}")
    r = start_row + 1

    # 図エリア: 白背景・罫線なし（画像を配置するため装飾不要）
    for row in range(r, r + height_rows):
        set_h(ws, row, 20)
    ws.merge_cells(start_row=r, start_column=GRID_LEFT,
                   end_row=r + height_rows - 1, end_column=GRID_RIGHT)

    # スペーサー
    spacer_row = r + height_rows
    set_h(ws, spacer_row, 8)

    return spacer_row + 1


# =====================================================================
# Sheet 1: 改版履歴
# =====================================================================
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6,  11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

def build_revision(wb):
    ws = wb.active
    ws.title = "改版履歴"
    setup_grid(ws)

    title_band(ws, 1, "改版履歴")
    set_h(ws, 2, 6)

    set_h(ws, 3, 22)
    MW(ws, 3, 2, 5,  "プロジェクト名", bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 6, 18, "", border=B_all())
    MW(ws, 3, 19, 22, "作成日",        bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, 3, 23, 31, "", border=B_all())

    set_h(ws, 4, 8)
    table_header(ws, 5, [(cs, ce, label) for label, (cs, ce) in REV_COLS.items()])
    data_rows(ws, 6, 25, [(cs, ce) for cs, ce in REV_COLS.values()])


# =====================================================================
# Sheet 2: 概要
# =====================================================================
def build_overview(wb):
    ws = wb.create_sheet("概要")
    setup_grid(ws)

    title_band(ws, 1, "詳細設計書 — 概要")
    set_h(ws, 2, 6)

    r = 3
    for label, h in [
        ("機能名",           22),
        ("機能概要",         48),
        ("目的",             48),
        ("利用者/利用部門",  22),
        ("起点画面",         22),
        ("操作トリガー",     22),
    ]:
        label_row(ws, r, label, label_cs=2, label_ce=6, val_cs=7, val_ce=31, height=h)
        r += 1


# =====================================================================
# Sheet 3: 業務フロー
# =====================================================================
BF_COLS = [
    (2,  3,  "No"),
    (4,  8,  "アクター"),
    (9,  22, "処理内容"),
    (23, 31, "分岐条件"),
]

def build_business_flow(wb):
    ws = wb.create_sheet("業務フロー")
    setup_grid(ws)

    title_band(ws, 1, "詳細設計書 — 業務フロー")
    set_h(ws, 2, 6)

    # テーブル（上）
    section_band(ws, 3, "\u25a0 1. 業務フロー")
    table_header(ws, 4, BF_COLS)
    data_rows(ws, 5, 24, [(cs, ce) for cs, ce, _ in BF_COLS])

    # スペーサー
    set_h(ws, 25, 8)

    # 図エリア（下）
    diagram_area(ws, 26, "業務フロー図（自動生成）", section_no=2)


# =====================================================================
# Sheet 4: 対象オブジェクト
# =====================================================================
OBJ_COLS = [
    (2,  7,  "オブジェクト名"),
    (8,  14, "項目API名"),
    (15, 20, "項目ラベル"),
    (21, 23, "読み書き区分"),
    (24, 31, "備考"),
]

def build_target_objects(wb):
    ws = wb.create_sheet("対象オブジェクト")
    setup_grid(ws)

    title_band(ws, 1, "詳細設計書 — 対象オブジェクト")
    set_h(ws, 2, 6)

    # テーブル（上）
    section_band(ws, 3, "\u25a0 1. 対象オブジェクト・項目一覧")
    table_header(ws, 4, OBJ_COLS)
    data_rows(ws, 5, 34, [(cs, ce) for cs, ce, _ in OBJ_COLS])

    # スペーサー
    set_h(ws, 35, 8)

    # 図エリア（下）
    diagram_area(ws, 36, "オブジェクト関連図（自動生成）", section_no=2)


# =====================================================================
# Sheet 5: 処理概要
# =====================================================================
PROC_COLS = [
    (2,  3,  "No"),
    (4,  15, "処理内容"),
    (16, 22, "コンポーネント"),
    (23, 31, "分岐条件"),
]

def build_process_overview(wb):
    ws = wb.create_sheet("処理概要")
    setup_grid(ws)

    title_band(ws, 1, "詳細設計書 — 処理概要")
    set_h(ws, 2, 6)

    # テーブル（上）
    section_band(ws, 3, "\u25a0 1. 処理概要")
    table_header(ws, 4, PROC_COLS)
    data_rows(ws, 5, 24, [(cs, ce) for cs, ce, _ in PROC_COLS])

    # スペーサー
    set_h(ws, 25, 8)

    # 図エリア（下）
    diagram_area(ws, 26, "処理フロー図（自動生成）", section_no=2)


# =====================================================================
# Sheet 6: 関連コンポーネント
# =====================================================================
COMP_COLS = [
    (2,  9,  "コンポーネント名"),
    (10, 13, "種別"),
    (14, 24, "役割"),
    (25, 31, "依存方向"),
]

def build_related_components(wb):
    ws = wb.create_sheet("関連コンポーネント")
    setup_grid(ws)

    title_band(ws, 1, "詳細設計書 — 関連コンポーネント")
    set_h(ws, 2, 6)

    # テーブル（上）
    section_band(ws, 3, "\u25a0 1. 関連コンポーネント一覧")
    table_header(ws, 4, COMP_COLS)
    data_rows(ws, 5, 19, [(cs, ce) for cs, ce, _ in COMP_COLS])

    # スペーサー
    set_h(ws, 20, 8)

    # 図エリア（下）
    diagram_area(ws, 21, "コンポーネント関連図（自動生成）", section_no=2)


# =====================================================================
# Sheet 7: 影響範囲
# =====================================================================
IMPACT_SECTIONS = [
    {
        "title": "更新オブジェクト",
        "cols": [(2, 9, "オブジェクト名"), (10, 20, "更新項目"), (21, 31, "更新条件")],
        "rows": 8,
    },
    {
        "title": "参照オブジェクト",
        "cols": [(2, 9, "オブジェクト名"), (10, 20, "参照項目"), (21, 31, "参照目的")],
        "rows": 8,
    },
    {
        "title": "関連Apex/Flow/LWC",
        "cols": [(2, 9, "名称"), (10, 13, "種別"), (14, 31, "関連内容")],
        "rows": 8,
    },
    {
        "title": "外部連携影響",
        "cols": [(2, 9, "連携先"), (10, 31, "影響内容")],
        "rows": 5,
    },
    {
        "title": "他機能依存",
        "cols": [(2, 9, "機能名"), (10, 31, "依存内容")],
        "rows": 5,
    },
]

def build_impact_scope(wb):
    ws = wb.create_sheet("影響範囲")
    setup_grid(ws)

    title_band(ws, 1, "詳細設計書 — 影響範囲")
    set_h(ws, 2, 6)

    r = 3
    for sec in IMPACT_SECTIONS:
        section_band(ws, r, f"■ {sec['title']}")
        r += 1
        table_header(ws, r, sec["cols"])
        r += 1
        end_r = r + sec["rows"] - 1
        data_rows(ws, r, end_r, [(cs, ce) for cs, ce, _ in sec["cols"]])
        r = end_r + 1
        # スペーサー行
        set_h(ws, r, 6)
        r += 1


# =====================================================================
# エントリポイント
# =====================================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, help="出力先 xlsx パス")
    args = parser.parse_args()

    wb = Workbook()
    build_revision(wb)
    build_overview(wb)
    build_business_flow(wb)
    build_target_objects(wb)
    build_process_overview(wb)
    build_related_components(wb)
    build_impact_scope(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    print(f"詳細設計書テンプレート生成完了: {out}")


if __name__ == "__main__":
    main()
