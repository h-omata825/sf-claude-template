"""
機能設計書テンプレート Excel を生成する。

4シート構成:
  1. 改版履歴     : 基本情報 + 改版履歴テーブル（20行）
  2. 処理概要     : メタ情報 + 目的/処理概要/前提条件/処理契機
  3. 処理内容     : 左=ステップ詳細 / 右=フローチャート画像
  4. パラメータ定義: 入力/出力パラメータテーブル

方針:
  - 狭幅グリッド（幅3.0の列を約30本）+ セル結合でレイアウト調整
  - ラベル/値を明示的に結合して「プロジェクト名」等が潰れないように
  - 結合セルの全構成セルに罫線を付与（途切れ防止）
  - 冗長なメタ行（処理内容・パラメータ定義）は削除、装飾のみ

Usage:
  python build_template.py --output "C:/.../設計書テンプレート.xlsx"
"""
from __future__ import annotations
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── カラーパレット ──────────────────────────────────────────────
C_TITLE_DARK = "1F3864"   # タイトル濃紺
C_HDR_BLUE   = "2E75B6"   # 中青（テーブルヘッダ）
C_BAND_BLUE  = "0070C0"   # セクションバンド
C_LABEL_BG   = "D9E1F2"   # 薄青（ラベル背景）
C_WHITE      = "FFFFFF"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"
C_FONT_GRAY  = "595959"

THIN  = Side(style="thin",   color="8B9DC3")
MED   = Side(style="medium", color="1F3864")

# ── スタイルヘルパー ────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def B_frame(left=True, right=True, top=True, bottom=True):
    """セクション外枠（太線）。"""
    return Border(
        left=MED if left else THIN,
        right=MED if right else THIN,
        top=MED if top else THIN,
        bottom=MED if bottom else THIN,
    )


def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = border
    return c


def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    """行方向の結合。結合前に全構成セルへ罫線/塗りを付与（途切れ防止）。"""
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)


def MW2D(ws, rs, re, cs, ce, value="", border=None, bg=None, **kwargs):
    """矩形結合（複数行×複数列）。全構成セルへ罫線/塗り付与。"""
    if border:
        for r in range(rs, re + 1):
            for c in range(cs, ce + 1):
                ws.cell(row=r, column=c).border = border
    if bg:
        for r in range(rs, re + 1):
            for c in range(cs, ce + 1):
                ws.cell(row=r, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=rs, start_column=cs, end_row=re, end_column=ce)
    return W(ws, rs, cs, value, border=border, bg=bg, **kwargs)


def set_h(ws, row, h):
    ws.row_dimensions[row].height = h

def set_w(ws, letter, w):
    ws.column_dimensions[letter].width = w


# ── 共通レイアウト ──────────────────────────────────────────────
# グリッド: 列 A=margin, B〜AE（30列）= データ領域
GRID_LEFT  = 2     # B
GRID_RIGHT = 31    # AE
COL_W      = 4.2   # 各データ列の幅（ゆとりを持たせる）

def setup_grid(ws):
    """全シート共通の列幅セットアップ。"""
    set_w(ws, "A", 2.0)
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        set_w(ws, get_column_letter(i), COL_W)
    ws.sheet_view.showGridLines = False


def title_band(ws, row, text, height=36):
    set_h(ws, row, height)
    MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_TITLE_DARK, h="center", size=14)


def section_band(ws, row, text, height=26,
                 cs=None, ce=None):
    cs = cs if cs is not None else GRID_LEFT
    ce = ce if ce is not None else GRID_RIGHT
    set_h(ws, row, height)
    MW(ws, row, cs, ce, text,
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11)


# ─────────────────────────────────────────────────────────────
# Sheet 1: 改版履歴
# ─────────────────────────────────────────────────────────────
REV_META_ROW       = 3
REV_TABLE_HEAD_ROW = 5
REV_DATA_ROW_START = 6
REV_DATA_ROW_END   = 25

# 改版履歴のメタ行の列範囲定義（label_cs, label_ce, value_cs, value_ce）
REV_META_PROJECT = (2, 6, 7, 18)     # プロジェクト名
REV_META_DATE    = (19, 22, 23, 31)  # 作成日

# テーブルヘッダ列範囲
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6, 11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

def build_revision(wb):
    ws = wb.active
    ws.title = "改版履歴"
    setup_grid(ws)

    title_band(ws, 1, "改版履歴")
    set_h(ws, 2, 6)

    # メタ行
    set_h(ws, REV_META_ROW, 22)
    ls, le, vs, ve = REV_META_PROJECT
    MW(ws, REV_META_ROW, ls, le, "プロジェクト名",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center", size=10)
    MW(ws, REV_META_ROW, vs, ve, "", border=B_all())  # ← データ流込

    ls, le, vs, ve = REV_META_DATE
    MW(ws, REV_META_ROW, ls, le, "作成日",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center", size=10)
    MW(ws, REV_META_ROW, vs, ve, "", border=B_all())

    set_h(ws, 4, 10)

    # テーブルヘッダ
    set_h(ws, REV_TABLE_HEAD_ROW, 24)
    for label, (cs, ce) in REV_COLS.items():
        MW(ws, REV_TABLE_HEAD_ROW, cs, ce, label,
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center",
           border=B_frame(top=True, bottom=True, left=True, right=True))

    # データ行
    for r in range(REV_DATA_ROW_START, REV_DATA_ROW_END + 1):
        set_h(ws, r, 22)
        for label, (cs, ce) in REV_COLS.items():
            MW(ws, r, cs, ce, "", border=B_all())


# ─────────────────────────────────────────────────────────────
# Sheet 2: 処理概要
# ─────────────────────────────────────────────────────────────
# メタ2段 + 4セクション
OV_META_ROW_1 = 3
OV_META_ROW_2 = 4

# メタ行1: (label_cs, label_ce, value_cs, value_ce)
OV_META_1 = [
    ("プロジェクト名", 2, 5,  6, 14),
    ("システム名",    15, 17, 18, 22),
    ("機能名",        23, 25, 26, 31),
]
# メタ行2
OV_META_2 = [
    ("API名",    2, 5,  6, 14),
    ("作成者",  15, 17, 18, 20),
    ("作成日",  21, 23, 24, 26),
    ("バージョン", 27, 28, 29, 31),
]

# 各セクション: (section_header_row, data_row, data_row_height)
OV_SECTIONS = [
    ("1. 本書の目的",   6,  7,  80),
    ("2. 処理概要",     9,  10, 110),
    ("3. 前提条件",     12, 13, 100),
    ("4. 処理契機",     15, 16, 60),
]

def build_overview(wb):
    ws = wb.create_sheet("処理概要")
    setup_grid(ws)

    title_band(ws, 1, "処理概要")
    set_h(ws, 2, 6)

    # メタ行1
    set_h(ws, OV_META_ROW_1, 22)
    for label, ls, le, vs, ve in OV_META_1:
        MW(ws, OV_META_ROW_1, ls, le, label,
           bold=True, bg=C_LABEL_BG, border=B_all(), h="center", size=10)
        MW(ws, OV_META_ROW_1, vs, ve, "", border=B_all())

    # メタ行2
    set_h(ws, OV_META_ROW_2, 22)
    for label, ls, le, vs, ve in OV_META_2:
        MW(ws, OV_META_ROW_2, ls, le, label,
           bold=True, bg=C_LABEL_BG, border=B_all(), h="center", size=10)
        MW(ws, OV_META_ROW_2, vs, ve, "", border=B_all())

    set_h(ws, 5, 14)

    # セクション
    prev_data_row = None
    for label, hdr_row, data_row, data_h in OV_SECTIONS:
        # セクション間のスペーサー行
        if prev_data_row is not None:
            for spacer in range(prev_data_row + 1, hdr_row):
                set_h(ws, spacer, 12)
        section_band(ws, hdr_row, label)
        set_h(ws, data_row, data_h)
        # 左端ラベルなしで全幅の入力枠
        MW(ws, data_row, GRID_LEFT, GRID_RIGHT, "",
           border=B_frame(), wrap=True, v="top")
        prev_data_row = data_row


# ─────────────────────────────────────────────────────────────
# Sheet 3: 処理内容
# ─────────────────────────────────────────────────────────────
# 列レイアウト:
#   左半分（B〜N）: ステップ詳細  ← B=No, C〜E=タイトル, F〜N=詳細
#   右半分（O〜AE）: フローチャート画像貼付領域
PROC_LEFT_NO_CS,     PROC_LEFT_NO_CE     = 2, 3     # No
PROC_LEFT_TITLE_CS,  PROC_LEFT_TITLE_CE  = 4, 7     # タイトル
PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE = 8, 16    # 詳細（テキスト本体）
PROC_LEFT_END                            = 16       # 左半分の終端

PROC_FLOW_CS, PROC_FLOW_CE = 18, 31                 # フロー描画エリア

# セクションヘッダ行 = 3、左テーブル列ヘッダ行 = 4、データ開始 = 5
PROC_SEC_ROW      = 3
PROC_HEAD_ROW     = 4
PROC_DATA_ROW_START = 5

def build_process(wb):
    ws = wb.create_sheet("処理内容")
    setup_grid(ws)

    title_band(ws, 1, "処理内容")
    set_h(ws, 2, 6)

    # セクションヘッダ（左:処理内容 / 右:フローチャート）
    set_h(ws, PROC_SEC_ROW, 26)
    MW(ws, PROC_SEC_ROW, PROC_LEFT_NO_CS, PROC_LEFT_END, "■ 処理内容",
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11)
    # セパレータ列（隙間）
    MW(ws, PROC_SEC_ROW, PROC_FLOW_CS, PROC_FLOW_CE, "■ フローチャート",
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, h="center", size=11)

    # 左テーブル列ヘッダ（No / 処理タイトル / 処理詳細）
    set_h(ws, PROC_HEAD_ROW, 24)
    MW(ws, PROC_HEAD_ROW, PROC_LEFT_NO_CS, PROC_LEFT_NO_CE, "No",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, PROC_HEAD_ROW, PROC_LEFT_TITLE_CS, PROC_LEFT_TITLE_CE, "処理タイトル",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, PROC_HEAD_ROW, PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE, "処理詳細",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")

    # 右側フロー領域: 枠だけ用意（画像は generator が貼る）
    for r in range(PROC_DATA_ROW_START, PROC_DATA_ROW_START + 30):
        if not ws.row_dimensions[r].height:
            set_h(ws, r, 22)


# ─────────────────────────────────────────────────────────────
# Sheet 4: パラメータ定義
# ─────────────────────────────────────────────────────────────
# 列レイアウト:
#   B〜F   = パラメータ名
#   G〜J   = 型
#   K〜L   = 必須
#   M〜AE  = 説明
PARAM_NAME_CS, PARAM_NAME_CE   = 2, 6
PARAM_TYPE_CS, PARAM_TYPE_CE   = 7, 10
PARAM_REQ_CS,  PARAM_REQ_CE    = 11, 12
PARAM_DESC_CS, PARAM_DESC_CE   = 13, 31

PARAM_IN_SEC_ROW  = 3
PARAM_IN_HEAD_ROW = 4
# 出力セクション開始行は generator が入力データ末尾から動的に決定

def build_params(wb):
    ws = wb.create_sheet("パラメータ定義")
    setup_grid(ws)

    title_band(ws, 1, "パラメータ定義")
    set_h(ws, 2, 6)

    # 入力パラメータセクション
    section_band(ws, PARAM_IN_SEC_ROW, "■ 入力パラメータ")

    set_h(ws, PARAM_IN_HEAD_ROW, 22)
    MW(ws, PARAM_IN_HEAD_ROW, PARAM_NAME_CS, PARAM_NAME_CE, "パラメータ名",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, PARAM_IN_HEAD_ROW, PARAM_TYPE_CS, PARAM_TYPE_CE, "型",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, PARAM_IN_HEAD_ROW, PARAM_REQ_CS, PARAM_REQ_CE, "必須",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, PARAM_IN_HEAD_ROW, PARAM_DESC_CS, PARAM_DESC_CE, "説明",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    wb = Workbook()
    build_revision(wb)
    build_overview(wb)
    build_process(wb)
    build_params(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"テンプレート生成完了: {out}")


if __name__ == "__main__":
    main()
