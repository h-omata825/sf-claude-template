"""
プロジェクト概要書テンプレート Excel を生成する。

5シート構成:
  1. 表紙         : プロジェクト名・目的・スコープ・期間・体制・改版履歴
  2. システム概要  : 導入背景・解決する課題・システム全体構成（図エリア）・外部連携先一覧
  3. 業務フロー図  : As-Is / To-Be 業務フロー（図貼り付けエリア + 手順表）
  4. ER図         : 主要オブジェクトと関連（図貼り付けエリア + 関連定義表）
  5. 用語集        : 業務用語・Salesforce用語の対照表

方針:
  - 既存設計書と同じデザインシステム（狭幅グリッド + セル結合 + 罫線統一）
  - グリッド: col A=2.0、col 2〜31=4.2（30列）
  - 色: 1F3864（タイトル濃紺）/ 2E75B6（ヘッダ青）/ 0070C0（セクション帯）/ D9E1F2（ラベル薄青）
  - 図エリア: E8E8E8（薄グレー背景）の大型セル結合エリア

Usage:
  python build_basic_doc_template.py --output "C:/.../プロジェクト概要書テンプレート.xlsx"
"""
from __future__ import annotations
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── カラー ──────────────────────────────────────────────────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_DIAGRAM_BG = "F0F0F0"   # 図貼り付けエリア（薄グレー）
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"
C_FONT_GRAY  = "808080"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")
GRAY = Side(style="medium", color="AAAAAA")

GRID_LEFT  = 2
GRID_RIGHT = 31


# ── ヘルパー ───────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_med(): return Border(left=MED,  right=MED,  top=MED,  bottom=MED)
def B_diag(): return Border(left=GRAY, right=GRAY, top=GRAY, bottom=GRAY)

def set_h(ws, row, h): ws.row_dimensions[row].height = h
def set_w(ws, letter, w): ws.column_dimensions[letter].width = w


def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:     c.fill = _fill(bg)
    if border: c.border = border
    return c


def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)


def setup_grid(ws):
    set_w(ws, "A", 2.0)
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        set_w(ws, get_column_letter(i), 4.2)
    ws.sheet_view.showGridLines = False


def title_row(ws, row, text):
    MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_TITLE_DARK,
       h="center", size=14, border=B_med())
    set_h(ws, row, 28)


def section_row(ws, row, text, cs=None, ce=None):
    MW(ws, row, cs or GRID_LEFT, ce or GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, border=B_all())
    set_h(ws, row, 18)


def meta_row(ws, row, label, value="", col_label_end=8, col_val_end=31):
    MW(ws, row, GRID_LEFT, col_label_end, label,
       bold=True, bg=C_LABEL_BG, border=B_all())
    MW(ws, row, col_label_end + 1, col_val_end, value, border=B_all())
    set_h(ws, row, 16)


def hdr_row(ws, row, cols: list[tuple[int, int, str]]):
    for cs, ce, label in cols:
        MW(ws, row, cs, ce, label,
           bold=True, fg=C_FONT_W, bg=C_HDR_BLUE, h="center", border=B_all())
    set_h(ws, row, 18)


def data_rows(ws, row_start, count, cols: list[tuple[int, int]], row_h=16):
    for r in range(row_start, row_start + count):
        for cs, ce in cols:
            for c in range(cs, ce + 1):
                ws.cell(row=r, column=c).border = B_all()
            ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
        set_h(ws, r, row_h)


def diagram_area(ws, row_start, row_end, placeholder="← この領域に図を貼り付けてください"):
    """図貼り付けエリア: 薄グレー背景の大型結合セル"""
    for r in range(row_start, row_end + 1):
        for c in range(GRID_LEFT, GRID_RIGHT + 1):
            ws.cell(row=r, column=c).fill = _fill(C_DIAGRAM_BG)
        set_h(ws, r, 20)
    ws.merge_cells(
        start_row=row_start, start_column=GRID_LEFT,
        end_row=row_end, end_column=GRID_RIGHT
    )
    cell = ws.cell(row=row_start, column=GRID_LEFT, value=placeholder)
    cell.font = _fnt(color=C_FONT_GRAY, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.fill = _fill(C_DIAGRAM_BG)
    cell.border = B_diag()


def text_area(ws, row_start, row_end):
    """自由記述エリア: 白背景の大型結合セル"""
    for r in range(row_start, row_end + 1):
        for c in range(GRID_LEFT, GRID_RIGHT + 1):
            ws.cell(row=r, column=c).border = B_all()
        set_h(ws, r, 18)
    ws.merge_cells(
        start_row=row_start, start_column=GRID_LEFT,
        end_row=row_end, end_column=GRID_RIGHT
    )
    ws.cell(row=row_start, column=GRID_LEFT).border = B_all()
    ws.cell(row=row_start, column=GRID_LEFT).alignment = \
        Alignment(horizontal="left", vertical="top", wrap_text=True)


# ── シート 1: 表紙 ────────────────────────────────────────────
def build_cover_sheet(ws):
    setup_grid(ws)
    r = 1
    set_h(ws, r, 8); r += 1

    title_row(ws, r, "プロジェクト概要書"); r += 1
    set_h(ws, r, 6); r += 1

    # プロジェクト基本情報
    section_row(ws, r, "プロジェクト基本情報"); r += 1
    for label in ["プロジェクト名", "システム名", "目的・背景"]:
        meta_row(ws, r, label); r += 1
    meta_row(ws, r, "スコープ（対象）"); r += 1
    meta_row(ws, r, "スコープ（対象外）"); r += 1
    meta_row(ws, r, "開始日"); r += 1
    meta_row(ws, r, "終了予定日"); r += 1
    meta_row(ws, r, "本番公開日"); r += 1

    set_h(ws, r, 6); r += 1

    # 体制
    section_row(ws, r, "体制"); r += 1
    hdr_row(ws, r, [
        (2,  6,  "役割"),
        (7,  16, "氏名 / 組織"),
        (17, 22, "担当領域"),
        (23, 31, "備考"),
    ]); r += 1
    data_rows(ws, r, 6, [(2, 6), (7, 16), (17, 22), (23, 31)]); r += 6

    set_h(ws, r, 6); r += 1

    # 改版履歴
    section_row(ws, r, "改版履歴"); r += 1
    hdr_row(ws, r, [
        (2,  3,  "版"),
        (4,  7,  "改版日"),
        (8,  12, "改版者"),
        (13, 31, "改版内容"),
    ]); r += 1
    data_rows(ws, r, 10, [(2, 3), (4, 7), (8, 12), (13, 31)])


# ── シート 2: システム概要 ─────────────────────────────────────
def build_system_overview_sheet(ws):
    setup_grid(ws)
    r = 1
    set_h(ws, r, 8); r += 1

    title_row(ws, r, "システム概要"); r += 1
    set_h(ws, r, 6); r += 1

    # 導入背景・課題
    section_row(ws, r, "導入背景・解決する課題"); r += 1
    text_area(ws, r, r + 7); r += 8

    set_h(ws, r, 6); r += 1

    # システム全体構成（図エリア）
    section_row(ws, r, "システム全体構成"); r += 1
    diagram_area(ws, r, r + 18); r += 19

    set_h(ws, r, 6); r += 1

    # 外部連携先一覧
    section_row(ws, r, "外部連携先一覧"); r += 1
    hdr_row(ws, r, [
        (2,  8,  "連携先システム"),
        (9,  12, "方向"),
        (13, 18, "方式"),
        (19, 22, "頻度"),
        (23, 31, "目的・概要"),
    ]); r += 1
    data_rows(ws, r, 8, [(2, 8), (9, 12), (13, 18), (19, 22), (23, 31)])


# ── シート 3: 業務フロー図 ────────────────────────────────────
def build_flow_sheet(ws):
    setup_grid(ws)
    r = 1
    set_h(ws, r, 8); r += 1

    title_row(ws, r, "業務フロー図"); r += 1
    set_h(ws, r, 6); r += 1

    # As-Is
    section_row(ws, r, "As-Is 業務フロー（現状）"); r += 1
    diagram_area(ws, r, r + 22); r += 23

    set_h(ws, r, 8); r += 1

    # As-Is 手順表（サブセクションは薄い帯）
    MW(ws, r, GRID_LEFT, GRID_RIGHT, "As-Is 手順（テキスト補足）",
       bold=True, fg=C_FONT_D, bg=C_LABEL_BG, border=B_all())
    set_h(ws, r, 16); r += 1
    hdr_row(ws, r, [
        (2,  3,  "No"),
        (4,  8,  "担当"),
        (9,  22, "操作・処理内容"),
        (23, 31, "備考"),
    ]); r += 1
    data_rows(ws, r, 8, [(2, 3), (4, 8), (9, 22), (23, 31)], row_h=18); r += 8

    set_h(ws, r, 12); r += 1

    # To-Be
    section_row(ws, r, "To-Be 業務フロー（Salesforce導入後）"); r += 1
    diagram_area(ws, r, r + 22); r += 23

    set_h(ws, r, 8); r += 1

    MW(ws, r, GRID_LEFT, GRID_RIGHT, "To-Be 手順（テキスト補足）",
       bold=True, fg=C_FONT_D, bg=C_LABEL_BG, border=B_all())
    set_h(ws, r, 16); r += 1
    hdr_row(ws, r, [
        (2,  3,  "No"),
        (4,  8,  "担当"),
        (9,  22, "操作・処理内容"),
        (23, 31, "関連コンポーネント"),
    ]); r += 1
    data_rows(ws, r, 10, [(2, 3), (4, 8), (9, 22), (23, 31)], row_h=18)


# ── シート 4: ER図 ────────────────────────────────────────────
def build_er_sheet(ws):
    setup_grid(ws)
    r = 1
    set_h(ws, r, 8); r += 1

    title_row(ws, r, "ER図（オブジェクト関連図）"); r += 1
    set_h(ws, r, 6); r += 1

    # ER図エリア
    section_row(ws, r, "オブジェクト関連図"); r += 1
    diagram_area(ws, r, r + 28); r += 29

    set_h(ws, r, 8); r += 1

    # 関連定義表
    section_row(ws, r, "関連定義表"); r += 1
    hdr_row(ws, r, [
        (2,  8,  "親オブジェクト"),
        (9,  10, "関係"),
        (11, 17, "子オブジェクト"),
        (18, 23, "参照関係項目"),
        (24, 31, "備考"),
    ]); r += 1
    data_rows(ws, r, 15, [(2, 8), (9, 10), (11, 17), (18, 23), (24, 31)])


# ── シート 5: 用語集 ──────────────────────────────────────────
def build_glossary_sheet(ws):
    setup_grid(ws)
    r = 1
    set_h(ws, r, 8); r += 1

    title_row(ws, r, "用語集"); r += 1
    set_h(ws, r, 6); r += 1

    section_row(ws, r, "業務用語・Salesforce用語 対照表"); r += 1
    hdr_row(ws, r, [
        (2,  3,  "No"),
        (4,  10, "業務用語"),
        (11, 18, "Salesforce用語 / オブジェクト名"),
        (19, 31, "説明"),
    ]); r += 1
    data_rows(ws, r, 30, [(2, 3), (4, 10), (11, 18), (19, 31)], row_h=18)


# ── メイン ────────────────────────────────────────────────────
def build(output: Path):
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("表紙",       build_cover_sheet),
        ("システム概要", build_system_overview_sheet),
        ("業務フロー図", build_flow_sheet),
        ("ER図",       build_er_sheet),
        ("用語集",      build_glossary_sheet),
    ]
    for name, builder in sheets:
        ws = wb.create_sheet(name)
        builder(ws)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"saved: {output}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default=str(
        Path(__file__).parent / "プロジェクト概要書テンプレート.xlsx"))
    args = parser.parse_args()
    build(Path(args.output))
