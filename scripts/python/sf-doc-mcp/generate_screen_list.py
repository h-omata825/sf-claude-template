"""
画面一覧.xlsx を生成する（docs/design/ スキャン方式）。

入力 (docs/ 配下):
  docs/design/lwc/     — LWC設計書（*.md）
  docs/design/flow/    — 画面フロー設計書（*.md）
  docs/flow/usecases.md — ユースケース一覧（UC番号マッピング用）

出力:
  画面一覧.xlsx（2シート: 改版履歴 / 画面一覧）

Usage:
  python generate_screen_list.py \\
    --docs-dir <path/to/project/docs> \\
    --output <output/画面一覧.xlsx> \\
    --author "作成者名" \\
    [--project-name "プロジェクト名"]
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 色・スタイル（既存デザインシステムと統一）──────────────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_ALT_ROW    = "F0F4FA"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

GRID_LEFT  = 2
GRID_RIGHT = 31
COL_W      = 4.2


def _fill(c):   return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_med(): return Border(left=MED,  right=MED,  top=MED,  bottom=MED)


def _setup_grid(ws):
    ws.column_dimensions["A"].width = 2.0
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        ws.column_dimensions[get_column_letter(i)].width = COL_W
    ws.sheet_view.showGridLines = False


def _set_h(ws, row, h):
    ws.row_dimensions[row].height = h


def _MW(ws, row, cs, ce, value="", bold=False, fg=C_FONT_D, bg=None,
        h="left", v="center", wrap=True, border=None, size=10):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws.cell(row=row, column=cs, value=value)
    cell.font = _fnt(bold=bold, color=fg, size=size)
    cell.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:     cell.fill = _fill(bg)
    if border: cell.border = border
    return cell


def _title_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_W, bg=C_TITLE_DARK,
        h="center", size=14, border=B_med())
    _set_h(ws, row, 28)


def _section_row(ws, row, text):
    _MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
        bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, border=B_all())
    _set_h(ws, row, 18)


def _meta_row(ws, row, label, value="", col_label_end=8, col_val_end=31):
    _MW(ws, row, GRID_LEFT, col_label_end, label, bold=True, bg=C_LABEL_BG, border=B_all())
    _MW(ws, row, col_label_end + 1, col_val_end, value, border=B_all())
    _set_h(ws, row, 16)


def _hdr_row(ws, row, cols: list[tuple[int, int, str]]):
    for cs, ce, label in cols:
        _MW(ws, row, cs, ce, label, bold=True, fg=C_FONT_W, bg=C_HDR_BLUE,
            h="center", border=B_all())
    _set_h(ws, row, 18)


def _data_row(ws, row, cols_vals: list[tuple[int, int, str]], alt=False):
    bg = C_ALT_ROW if alt else None
    for cs, ce, val in cols_vals:
        if bg:
            for c in range(cs, ce + 1):
                ws.cell(row=row, column=c).fill = _fill(bg)
        ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
        cell = ws.cell(row=row, column=cs, value=val)
        cell.font = _fnt()
        cell.alignment = _aln()
        cell.border = B_all()
        if bg: cell.fill = _fill(bg)
    _set_h(ws, row, 16)


# ── 設計書 MD パーサー ────────────────────────────────────────────

def _table_val(text: str, key: str) -> str:
    m = re.search(rf'\|\s*{re.escape(key)}\s*\|\s*(.+?)\s*\|', text)
    return m.group(1).strip() if m else ""


def parse_design_md(path: Path) -> dict | None:
    try:
        text = path.read_text(encoding="utf-8")
    except Exception:
        return None

    # ファイル名からスクリーン種別を推定
    kind = "LWC" if "/lwc/" in str(path).replace("\\", "/") else "画面フロー"

    # 機能名は H1 タイトルから
    h1_m = re.match(r'^#\s+(.+)$', text, re.MULTILINE)
    name = h1_m.group(1).strip() if h1_m else path.stem

    feature_id = _table_val(text, "機能ID") or "TBD"
    target_obj  = _table_val(text, "担当オブジェクト") or ""
    source_file = _table_val(text, "ソース") or ""

    # 実装種別を上書き（設計書に明記されている場合）
    impl_kind = _table_val(text, "実装種別")
    if impl_kind:
        kind = impl_kind

    # 関連UC（ファイル本文から UC-XX を抽出）
    uc_m = re.findall(r'UC-\d+', text)
    related_uc = ", ".join(sorted(set(uc_m))) if uc_m else ""

    # スコープ（ユーザーストーリーセクションの先頭1文）
    scope_m = re.search(r'##\s*スコープ.+?\n(.*?)(?=\n##|\Z)', text, re.DOTALL)
    scope = scope_m.group(1).strip().split("\n")[0][:60] if scope_m else ""

    return {
        "id":          feature_id,
        "name":        name,
        "type":        kind,
        "related_uc":  related_uc,
        "target_obj":  target_obj,
        "source":      source_file,
        "scope":       scope,
    }


def _scan_screens(docs_dir: Path) -> list[dict]:
    screens = []
    design_root = docs_dir / "design"
    for sub in ("lwc", "flow"):
        folder = design_root / sub
        if not folder.exists():
            continue
        for md in sorted(folder.glob("*.md")):
            info = parse_design_md(md)
            if info:
                screens.append(info)
    return screens


# ── シート構築 ────────────────────────────────────────────────────

def _build_revision_sheet(ws, project_name: str, author: str):
    _setup_grid(ws)
    today = date.today().strftime("%Y-%m-%d")
    r = 1
    _set_h(ws, r, 8); r += 1
    _title_row(ws, r, "画面一覧"); r += 1
    _set_h(ws, r, 6); r += 1
    _meta_row(ws, r, "プロジェクト名",  project_name); r += 1
    _meta_row(ws, r, "作成日",         today); r += 1
    _meta_row(ws, r, "最終更新日",     today); r += 1
    _meta_row(ws, r, "バージョン",     "1.0"); r += 1
    _meta_row(ws, r, "作成者",         author); r += 1
    _set_h(ws, r, 6); r += 1
    _section_row(ws, r, "改版履歴"); r += 1
    _hdr_row(ws, r, [
        (2, 3,  "版"),
        (4, 7,  "改版日"),
        (8, 12, "改版者"),
        (13, 31, "改版内容"),
    ]); r += 1
    for i in range(5):
        for cs, ce in [(2, 3), (4, 7), (8, 12), (13, 31)]:
            for c in range(cs, ce + 1):
                ws.cell(row=r + i, column=c).border = B_all()
            ws.merge_cells(start_row=r + i, start_column=cs, end_row=r + i, end_column=ce)
        _set_h(ws, r + i, 16)


def _build_list_sheet(ws, project_name: str, screens: list[dict]):
    _setup_grid(ws)
    r = 1
    _set_h(ws, r, 8); r += 1
    _title_row(ws, r, f"画面一覧  {project_name}"); r += 1
    _set_h(ws, r, 6); r += 1
    _section_row(ws, r, "画面一覧"); r += 1

    HDR_COLS = [
        (2,  4,  "No"),
        (5,  8,  "機能ID"),
        (9,  16, "画面名"),
        (17, 19, "種別"),
        (20, 23, "対応UC"),
        (24, 27, "主要オブジェクト"),
        (28, 31, "概要"),
    ]
    _hdr_row(ws, r, HDR_COLS); r += 1

    DATA_COLS = [(2, 4), (5, 8), (9, 16), (17, 19), (20, 23), (24, 27), (28, 31)]

    if screens:
        for i, s in enumerate(screens):
            vals = [
                str(i + 1),
                s["id"],
                s["name"],
                s["type"],
                s["related_uc"],
                s["target_obj"],
                s["scope"],
            ]
            _data_row(ws, r, list(zip(
                [cs for cs, _ in DATA_COLS],
                [ce for _, ce in DATA_COLS],
                vals,
            )), alt=(i % 2 == 1))
            r += 1
    else:
        # データなし → 空行10行
        for i in range(10):
            for cs, ce in DATA_COLS:
                for c in range(cs, ce + 1):
                    ws.cell(row=r + i, column=c).border = B_all()
                ws.merge_cells(start_row=r + i, start_column=cs,
                               end_row=r + i, end_column=ce)
            _set_h(ws, r + i, 16)

    # 凡例（末尾）
    r_note = r + max(len(screens), 10) + 2
    _MW(ws, r_note, GRID_LEFT, GRID_RIGHT,
        "※ 種別: LWC / 画面フロー / Visualforce / Aura",
        fg="595959", size=9)
    _set_h(ws, r_note, 14)


# ── メイン ────────────────────────────────────────────────────────

def generate(docs_dir: Path, output: Path, author: str, project_name: str):
    screens = _scan_screens(docs_dir)

    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("改版履歴")
    ws2 = wb.create_sheet("画面一覧")

    _build_revision_sheet(ws1, project_name, author)
    _build_list_sheet(ws2, project_name, screens)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"saved: {output}  ({len(screens)} 件)")


def main():
    ap = argparse.ArgumentParser(description="画面一覧.xlsx 生成")
    ap.add_argument("--docs-dir",     required=True, help="docs/ フォルダのパス")
    ap.add_argument("--output",       required=True, help="出力先 .xlsx パス")
    ap.add_argument("--author",       default="",    help="作成者名")
    ap.add_argument("--project-name", default="",    help="プロジェクト名")
    args = ap.parse_args()
    generate(
        docs_dir=Path(args.docs_dir),
        output=Path(args.output),
        author=args.author,
        project_name=args.project_name,
    )


if __name__ == "__main__":
    main()
