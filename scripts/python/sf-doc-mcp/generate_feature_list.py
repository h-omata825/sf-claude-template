"""
機能一覧.xlsx を生成する（テンプレート駆動・種別別シート）。

テンプレート: 機能一覧テンプレート.xlsx
  - 改版履歴 / サマリー / __SHEET_TEMPLATE__（種別別シートの雛形）

出力:
  - 改版履歴 / サマリー / Apex / Batch / Flow / 画面フロー / LWC / ...（入力JSONに含まれる種別のみ）

Usage:
  python generate_feature_list.py \
    --input features.json \
    --output-dir /path/to/output \
    --author 作成者名 \
    --project-name プロジェクト名 \
    [--system-name システム名] \
    [--template /path/to/機能一覧テンプレート.xlsx]

Input JSON:
[
  { "id":"F-001", "type":"Apex", "name":"機能名", "api_name":"ClassName",
    "overview":"処理概要", "design_file":"【F-001】機能名.xlsx" },
  ...
]
"""
import argparse
import json
import shutil
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from meta_store import read_meta, write_meta
from version_manager import increment_version

DEFAULT_TEMPLATE = Path(r"C:\work\03_tool\新しいフォルダー\機能一覧テンプレート.xlsx")

C_ALT_ROW   = "F2F5FA"
C_FONT_D    = "000000"
C_FONT_LINK = "0563C1"
FONT_NAME   = "游ゴシック"
THIN        = Side(style="thin", color="8B9DC3")

TYPE_ORDER = ["Apex", "Batch", "Schedulable", "Trigger",
              "Flow", "画面フロー", "LWC", "Aura",
              "Visualforce", "その他"]

# ── テンプレートと一致させる定数 ────────────────────────────────
# 改版履歴
REV_META_ROW       = 3
REV_META_PROJECT_V = (7, 18)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6, 11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

# サマリー
SUM_META_ROW_1 = 3
SUM_META_ROW_2 = 4
# 値セル: (cs, ce)
SUM_PROJECT_V = (6, 18)
SUM_DATE_V    = (23, 31)
SUM_AUTHOR_V  = (6, 18)
SUM_TOTAL_V   = (23, 31)
SUM_DATA_ROW_START = 8
SUM_COLS = {
    "No":         (2,  3),
    "種別":       (4,  11),
    "件数":       (12, 15),
    "対応シート": (16, 31),
}

# 種別別シート
ST_META_ROW_1 = 3
ST_META_ROW_2 = 4
ST_COUNT_V   = (23, 31)  # 件数の値セル
ST_DATA_ROW_START = 8
ST_COLS = {
    "ID":              (2,  4),
    "API名/ファイル名": (5,  11),
    "機能名":          (12, 18),
    "処理概要":        (19, 27),
    "設計書ファイル":  (28, 31),
}


def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)
def _fill(c): return PatternFill("solid", fgColor=c)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def MW(ws, row, cs, ce, value="", border=None, bg=None, bold=False,
       fg=C_FONT_D, h="left", v="center", wrap=True, size=10):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    cell = ws.cell(row=row, column=cs, value=value)
    cell.font = _fnt(bold=bold, color=fg, size=size)
    cell.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg: cell.fill = _fill(bg)
    if border: cell.border = border
    return cell

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h


def clone_sheet(wb, source_name: str, new_name: str) -> Worksheet:
    src = wb[source_name]
    new = wb.copy_worksheet(src)
    new.title = new_name
    return new


# ── 差分計算 ──────────────────────────────────────────────────
def _feature_comparable(f: dict) -> tuple:
    """差分検出用に比較対象キーのみ抽出"""
    return (
        f.get("type", ""),
        f.get("name", ""),
        f.get("api_name", ""),
        f.get("overview", "") or "",
        f.get("design_file", "") or "",
    )


def compare_features(old_features: list, new_features: list) -> dict:
    """前回と今回の機能一覧を比較し差分を返す。
    キーは id（feature_ids.yml で安定）を使う。
    """
    old_map = {f.get("id"): f for f in (old_features or []) if f.get("id")}
    new_map = {f.get("id"): f for f in new_features if f.get("id")}

    added    = [new_map[k] for k in new_map if k not in old_map]
    removed  = [old_map[k] for k in old_map if k not in new_map]
    modified = []
    for k in new_map:
        if k in old_map and _feature_comparable(old_map[k]) != _feature_comparable(new_map[k]):
            modified.append({"id": k, "old": old_map[k], "new": new_map[k]})

    return {"added": added, "removed": removed, "modified": modified}


def has_any_diff(diffs: dict) -> bool:
    return bool(diffs["added"] or diffs["removed"] or diffs["modified"])


def build_revision_entries(current_version: str, diffs: dict, author: str,
                           today: str, start_no: int, is_major: bool,
                           is_initial: bool) -> list[dict]:
    """改版履歴エントリを構築する。
    戻り値: [{項番, 版数, 変更箇所, 変更内容, 変更理由, 変更日, 変更者, 備考}, ...]
    """
    if is_initial:
        return [{
            "項番": start_no, "版数": current_version, "変更箇所": "全シート",
            "変更内容": "新規作成", "変更理由": "", "変更日": today,
            "変更者": author, "備考": "",
        }]

    entries = []
    if is_major:
        entries.append({
            "項番": start_no, "版数": current_version, "変更箇所": "全シート",
            "変更内容": "メジャーバージョンアップ", "変更理由": "",
            "変更日": today, "変更者": author, "備考": "",
        })
        start_no_incr = False
    else:
        start_no_incr = True

    rows = []
    for f in diffs["added"]:
        rows.append(("サマリー / " + f.get("type", ""),
                     f"機能追加: {f.get('id')} {f.get('name', '')}"))
    for f in diffs["removed"]:
        rows.append(("サマリー / " + f.get("type", ""),
                     f"機能削除: {f.get('id')} {f.get('name', '')}"))
    for m in diffs["modified"]:
        nv = m["new"]; ov = m["old"]
        changed = []
        for k in ("name", "api_name", "overview", "design_file"):
            if ov.get(k) != nv.get(k):
                changed.append(k)
        rows.append(("サマリー / " + nv.get("type", ""),
                     f"機能変更: {m['id']} {nv.get('name', '')}（{', '.join(changed)}）"))

    for i, (sheet, content) in enumerate(rows):
        is_first_data = (i == 0 and start_no_incr)
        entries.append({
            "項番":     start_no if is_first_data else "",
            "版数":     current_version if is_first_data else "",
            "変更箇所": sheet,
            "変更内容": content,
            "変更理由": "",
            "変更日":   today if is_first_data else "",
            "変更者":   author if is_first_data else "",
            "備考":     "",
        })

    if not entries:
        entries.append({
            "項番": start_no, "版数": current_version, "変更箇所": "—",
            "変更内容": "変更なし", "変更理由": "", "変更日": today,
            "変更者": author, "備考": "",
        })

    return entries


# ── 埋め込み ───────────────────────────────────────────────────
def fill_revision(ws, history: list, project_name: str, today: str):
    """history 配列を改版履歴テーブルに書き込む。"""
    vs, ve = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=project_name)
    vs, ve = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=today)

    r = REV_DATA_ROW_START
    for h in history:
        for label, (cs, ce) in REV_COLS.items():
            ws.cell(row=r, column=cs, value=h.get(label, ""))
        r += 1


def fill_summary(ws, groups: dict, project_name: str, author: str, today: str,
                 sheet_name_map: dict):
    # メタ
    vs, ve = SUM_PROJECT_V
    ws.cell(row=SUM_META_ROW_1, column=vs, value=project_name)
    vs, ve = SUM_DATE_V
    ws.cell(row=SUM_META_ROW_1, column=vs, value=today)
    vs, ve = SUM_AUTHOR_V
    ws.cell(row=SUM_META_ROW_2, column=vs, value=author)
    vs, ve = SUM_TOTAL_V
    total = sum(len(v) for v in groups.values())
    ws.cell(row=SUM_META_ROW_2, column=vs, value=f"{total}件")

    sorted_types = ([t for t in TYPE_ORDER if t in groups]
                    + [t for t in groups.keys() if t not in TYPE_ORDER])
    r = SUM_DATA_ROW_START
    for i, type_key in enumerate(sorted_types):
        count = len(groups[type_key])
        sheet_name = sheet_name_map[type_key]
        set_h(ws, r, 24)
        MW(ws, r, *SUM_COLS["No"], value=str(i + 1),
           border=B_all(), h="center")
        MW(ws, r, *SUM_COLS["種別"], value=type_key, border=B_all())
        MW(ws, r, *SUM_COLS["件数"], value=count,
           border=B_all(), h="center")
        cell = MW(ws, r, *SUM_COLS["対応シート"], value=sheet_name,
                  border=B_all(), fg=C_FONT_LINK)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        r += 1


def fill_type_sheet(ws, type_key: str, features: list):
    # タイトル差し替え
    ws.cell(row=1, column=2, value=f"機能一覧 — {type_key}")
    # 件数
    vs, ve = ST_COUNT_V
    ws.cell(row=ST_META_ROW_2, column=vs, value=f"{len(features)}件")

    r = ST_DATA_ROW_START
    for i, feat in enumerate(features):
        bg = C_ALT_ROW if i % 2 == 1 else None
        overview = feat.get("overview", "") or ""
        # 高さを概要文字数から推定
        set_h(ws, r, max(26, min(120, (len(overview) // 30) * 16 + 28)))
        vals = {
            "ID":              (feat.get("id", ""), "center"),
            "API名/ファイル名": (feat.get("api_name", ""), "left"),
            "機能名":          (feat.get("name", ""), "left"),
            "処理概要":        (overview, "left"),
            "設計書ファイル":  (feat.get("design_file", ""), "left"),
        }
        for label, (cs, ce) in ST_COLS.items():
            val, ha = vals[label]
            MW(ws, r, cs, ce, value=val, border=B_all(), bg=bg,
               h=ha, v="top")
        r += 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",        required=True)
    parser.add_argument("--output-dir",   required=True)
    parser.add_argument("--author",       required=True)
    parser.add_argument("--project-name", default="")
    parser.add_argument("--system-name",  default="")
    parser.add_argument("--template",     default=str(DEFAULT_TEMPLATE))
    parser.add_argument("--source-file",  default="",
                        help="更新時: 既存の機能一覧.xlsx のパス")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"],
                        help="minor: x.1増 / major: 1.0増")
    args = parser.parse_args()

    template = Path(args.template)
    if not template.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {template}")

    features = json.loads(Path(args.input).read_text(encoding="utf-8"))
    today = date.today().strftime("%Y-%m-%d")

    # ── バージョン判定 ──────────────────────────────────────
    is_major    = (args.version_increment == "major")
    source_file = args.source_file.strip()
    prev_meta   = read_meta(source_file) if source_file else None

    if prev_meta:
        current_version = increment_version(prev_meta.get("version", "1.0"),
                                            args.version_increment)
        history         = prev_meta.get("history", [])
        old_features    = prev_meta.get("features", [])
        is_initial      = False
        print(f"更新モード: {prev_meta.get('version', '?')} → {current_version}"
              + (" (メジャー)" if is_major else ""))
    else:
        current_version = "1.0"
        history         = []
        old_features    = []
        is_initial      = True
        print("新規作成モード: v1.0")

    # ── 差分計算 ────────────────────────────────────────────
    diffs = compare_features(old_features, features)
    if prev_meta and not is_major and not has_any_diff(diffs):
        print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    last_no = max((h["項番"] for h in history
                   if isinstance(h.get("項番"), int)), default=0)
    new_entries = build_revision_entries(
        current_version, diffs, args.author, today,
        start_no=last_no + 1, is_major=is_major, is_initial=is_initial,
    )
    history = history + new_entries

    # ── xlsx 生成 ─────────────────────────────────────────────
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "機能一覧.xlsx"
    shutil.copy(template, out_path)

    # 種別ごとにグループ化
    groups: dict = defaultdict(list)
    for feat in features:
        groups[feat.get("type", "その他")].append(feat)

    wb = load_workbook(out_path)

    fill_revision(wb["改版履歴"], history, args.project_name, today)

    sheet_name_map = {}
    sorted_types = ([t for t in TYPE_ORDER if t in groups]
                    + [t for t in groups.keys() if t not in TYPE_ORDER])
    for type_key in sorted_types:
        sheet_name = type_key
        for ch in r'[]:*?/\\':
            sheet_name = sheet_name.replace(ch, "_")
        new_ws = clone_sheet(wb, "__SHEET_TEMPLATE__", sheet_name)
        sheet_name_map[type_key] = sheet_name
        fill_type_sheet(new_ws, type_key, groups[type_key])

    fill_summary(wb["サマリー"], groups, args.project_name, args.author, today,
                 sheet_name_map)

    wb["__SHEET_TEMPLATE__"].sheet_state = "hidden"

    # ── _meta 保存（次回差分判定用）────────────────────────
    write_meta(wb, {
        "version":      current_version,
        "date":         today,
        "project_name": args.project_name,
        "system_name":  args.system_name,
        "author":       args.author,
        "features":     features,
        "history":      history,
    })

    wb.save(out_path)
    print(f"機能一覧生成完了: v{current_version} → {out_path}")
    print(f"  差分: 追加{len(diffs['added'])} / 削除{len(diffs['removed'])} / 変更{len(diffs['modified'])}")
    print(f"  シート: {wb.sheetnames}")


if __name__ == "__main__":
    main()
