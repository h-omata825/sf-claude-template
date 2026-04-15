"""
機能設計書.xlsx を1機能分生成する（テンプレート読込方式）。

  設計書テンプレート.xlsx（build_template.py で生成される「器」）を
  コピーしてセル値を流し込み、右側にフローチャート画像を貼り付ける。

4シート構成:
  1. 改版履歴     : メタ(プロジェクト名/作成日) + 履歴テーブル(初版自動投入)
  2. 処理概要     : メタ2段 + 目的/処理概要/前提条件/処理契機
  3. 処理内容     : 左=ステップ詳細 / 右=フローチャート画像
  4. パラメータ定義: 入力/出力テーブル

Usage:
  python generate_feature_design.py \
    --input     feature_design.json \
    --template  "C:/.../設計書テンプレート.xlsx" \
    --output-dir "C:/.../出力ルート"

出力先: {output-dir}/{type_folder}/【F-001】機能名.xlsx
"""
from __future__ import annotations
import argparse
import json
import sys
import tempfile
from datetime import date as _date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from flowchart_utils import generate_flowchart
from meta_store import read_meta, write_meta
from version_manager import increment_version
import design_revision as dr

# ── 色定数 ──────────────────────────────────────────────────────
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_LABEL_BG   = "D9E1F2"
C_APEX_HDR   = "F4B183"
C_STEP_BG    = "E2EFDA"
C_SUB_BG     = "F2F2F2"
C_FONT_D     = "000000"
C_FONT_GRAY  = "595959"
C_FONT_W     = "FFFFFF"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

TYPE_FOLDER = {
    "Apex":      "apex",
    "Batch":     "batch",
    "Flow":      "flow",
    "Aura":      "aura",
    "その他":    "other",
}

# ── テンプレートと一致させる定数 ────────────────────────────────
# 改版履歴
REV_META_ROW       = 3
REV_META_PROJECT_V = (7, 18)   # 値セル(cs, ce)
REV_META_DATE_V    = (23, 31)
REV_DATA_ROW_START = 6
REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6, 11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

# 処理概要
OV_META_ROW_1 = 3
OV_META_ROW_2 = 4
# (label, value_cs, value_ce)
OV_META_1_V = {
    "project_name": (6, 14),
    "system_name":  (18, 22),
    "name":         (26, 31),
}
OV_META_2_V = {
    "api_name": (6, 14),
    "author":   (18, 20),
    "date":     (24, 26),
    "version":  (29, 31),
}
OV_SECTION_DATA_ROW = {
    "purpose":       7,
    "overview":      10,
    "prerequisites": 13,
    "trigger":       16,
}

# 処理内容
PROC_LEFT_NO_CS,     PROC_LEFT_NO_CE     = 2, 3
PROC_LEFT_TITLE_CS,  PROC_LEFT_TITLE_CE  = 4, 7
PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE = 8, 16
PROC_LEFT_END                            = 16
PROC_FLOW_CS, PROC_FLOW_CE = 18, 31
PROC_DATA_ROW_START = 5

# パラメータ定義
PARAM_NAME_CS, PARAM_NAME_CE = 2, 6
PARAM_TYPE_CS, PARAM_TYPE_CE = 7, 10
PARAM_REQ_CS,  PARAM_REQ_CE  = 11, 12
PARAM_DESC_CS, PARAM_DESC_CE = 13, 31
PARAM_IN_HEAD_ROW = 4
PARAM_IN_DATA_ROW_START = 5  # 入力データ先頭


# ── スタイルヘルパー ────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10, italic=False):
    return Font(name="游ゴシック", bold=bold, color=color, size=size, italic=italic)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_frame(left=True, right=True, top=True, bottom=True):
    return Border(
        left=MED if left else THIN,
        right=MED if right else THIN,
        top=MED if top else THIN,
        bottom=MED if bottom else THIN,
    )

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    """結合前に全構成セルへ罫線/塗り付与（途切れ防止）。"""
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h


# ── APEX呼び出しテーブル ────────────────────────────────────────
def build_apex_table(ws, row, apex):
    """処理内容シートの左半分に APEX 呼び出し表を描画。"""
    set_h(ws, row, 24)
    MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_TITLE_CE, "■APEX呼び出し先",
       bold=True, bg=C_APEX_HDR, border=B_all())
    MW(ws, row, PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE, apex.get("name", ""),
       bold=True, bg=C_APEX_HDR, border=B_all())
    row += 1

    for section_label, key in [("Input", "inputs"), ("Output", "responses")]:
        set_h(ws, row, 20)
        MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_END, section_label,
           bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
        row += 1

        set_h(ws, row, 20)
        MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_TITLE_CE, "Key",
           bold=True, bg=C_LABEL_BG, border=B_all())
        MW(ws, row, PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE, "Value",
           bold=True, bg=C_LABEL_BG, border=B_all())
        row += 1

        items = apex.get(key, [])
        if not items:
            set_h(ws, row, 22)
            MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_TITLE_CE, "（なし）",
               fg=C_FONT_GRAY, italic=True, border=B_all())
            MW(ws, row, PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE, "",
               border=B_all())
            row += 1
        else:
            for it in items:
                key_text = it.get("key", "")
                val_text = it.get("value", "")
                val_lines = max(1, len(val_text) // 28 + val_text.count("\n") + 1)
                set_h(ws, row, max(22, val_lines * 16))
                MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_TITLE_CE, key_text,
                   border=B_all(), v="top")
                MW(ws, row, PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE, val_text,
                   border=B_all(), v="top", wrap=True)
                row += 1

    set_h(ws, row, 8)
    row += 1
    return row


# ── シート埋め込み ─────────────────────────────────────────────
def fill_revision(ws, data, history: list[dict]):
    # メタ: プロジェクト名・作成日
    vs, ve = REV_META_PROJECT_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("project_name", ""))
    vs, ve = REV_META_DATE_V
    ws.cell(row=REV_META_ROW, column=vs, value=data.get("date", ""))

    # 改版履歴行
    dr.fill_revision_table(ws, history, REV_COLS, REV_DATA_ROW_START)


def fill_overview(ws, data, changed_fields: set = None):
    changed_fields = changed_fields or set()
    # メタ行1
    for key, (cs, ce) in OV_META_1_V.items():
        ws.cell(row=OV_META_ROW_1, column=cs, value=data.get(key, ""))
    # メタ行2
    for key, (cs, ce) in OV_META_2_V.items():
        ws.cell(row=OV_META_ROW_2, column=cs, value=data.get(key, ""))
    # セクション本文
    for key, r in OV_SECTION_DATA_ROW.items():
        cell = ws.cell(row=r, column=2, value=data.get(key, ""))
        if key in changed_fields:
            dr.apply_red(cell, size=10)


def _estimate_content_height_pts(steps: list) -> float:
    """処理内容の行高さ合計をポイント単位で概算する（フロー図の縦幅合わせに使用）。

    fill_process() と同じ高さロジックを使用するため、実際の行高さと概ね一致する。
    """
    APPROX_CHARS = 28
    total = 0.0
    for step in steps:
        total += 26  # タイトル行
        detail = step.get("detail", "")
        if detail:
            est = max(2, len(detail) // APPROX_CHARS + detail.count("\n") + 1)
            total += max(48, est * 16)
        for sub in step.get("sub_steps", []):
            total += 20  # サブタイトル行
            sdetail = sub.get("detail", "")
            if sdetail:
                slines = max(2, len(sdetail.splitlines()) + 1)
                total += max(28, slines * 13)
        total += 10  # ステップ間スペーサー
    return max(total, 200)


def fill_process(ws, data, flowchart_path, changed_step_nos: set = None):
    changed_step_nos = changed_step_nos or set()
    row = PROC_DATA_ROW_START
    for step in data.get("steps", []):
        step_no = step.get("no", "")
        is_changed = step_no in changed_step_nos
        # ステップ行（タイトル）
        set_h(ws, row, 26)
        c_no = MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_NO_CE, step_no,
                  bold=True, border=B_all(), h="center", bg=C_STEP_BG)
        c_t  = MW(ws, row, PROC_LEFT_TITLE_CS, PROC_LEFT_TITLE_CE, step.get("title", ""),
                  bold=True, border=B_all(), bg=C_STEP_BG)
        MW(ws, row, PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE, "",
           border=B_all(), bg=C_STEP_BG)
        if is_changed:
            dr.apply_red(c_no, bold=True)
            dr.apply_red(c_t,  bold=True)
        row += 1

        # 詳細
        detail = step.get("detail", "")
        if detail:
            # 文字数から概算行数を出して高さ調整（ゆとり持たせる）
            approx_chars_per_line = 28
            est_lines = max(2, len(detail) // approx_chars_per_line + detail.count("\n") + 1)
            set_h(ws, row, max(48, est_lines * 16))
            MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_NO_CE, "", border=B_all())
            c_d = MW(ws, row, PROC_LEFT_TITLE_CS, PROC_LEFT_DETAIL_CE, detail,
                     border=B_all(), wrap=True, v="top")
            if is_changed:
                dr.apply_red(c_d)
            row += 1

        # サブステップ
        for sub in step.get("sub_steps", []):
            set_h(ws, row, 20)
            MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_NO_CE, sub.get("no", ""),
               fg=C_FONT_GRAY, border=B_all(), h="center", bg=C_SUB_BG)
            MW(ws, row, PROC_LEFT_TITLE_CS, PROC_LEFT_TITLE_CE, sub.get("title", ""),
               bold=True, fg=C_FONT_GRAY, border=B_all(), bg=C_SUB_BG)
            MW(ws, row, PROC_LEFT_DETAIL_CS, PROC_LEFT_DETAIL_CE, "",
               border=B_all(), bg=C_SUB_BG)
            row += 1

            sdetail = sub.get("detail", "")
            if sdetail:
                slines = max(2, len(sdetail.splitlines()) + 1)
                set_h(ws, row, max(28, slines * 13))
                MW(ws, row, PROC_LEFT_NO_CS, PROC_LEFT_NO_CE, "",
                   border=B_all(), bg=C_SUB_BG)
                MW(ws, row, PROC_LEFT_TITLE_CS, PROC_LEFT_DETAIL_CE, sdetail,
                   fg=C_FONT_GRAY, border=B_all(), wrap=True, v="top", bg=C_SUB_BG)
                row += 1

        # APEX呼び出し表
        for apex in step.get("apex_calls", []):
            row = build_apex_table(ws, row, apex)

        # ステップ間のスペーサー
        set_h(ws, row, 10)
        row += 1

    # フロー画像貼付
    if flowchart_path and Path(flowchart_path).exists():
        try:
            img = XLImage(flowchart_path)
            img.anchor = f"{get_column_letter(PROC_FLOW_CS)}4"
            # 画像の自然な縦横比を保って貼り付ける（generate_flowchart側がGAP調整済み）
            ratio = img.height / img.width if img.width else 1.6
            img.width  = 440
            img.height = min(int(440 * ratio), 2400)
            ws.add_image(img)
        except Exception as e:
            W(ws, 4, PROC_FLOW_CS, f"[フロー図挿入失敗: {e}]", italic=True, fg="888888")


def fill_params(ws, data, changed_input_keys: set = None,
                changed_output_keys: set = None):
    changed_input_keys  = changed_input_keys  or set()
    changed_output_keys = changed_output_keys or set()
    # 入力パラメータ
    r = PARAM_IN_DATA_ROW_START
    inputs = data.get("input_params", [])
    if not inputs:
        set_h(ws, r, 20)
        MW(ws, r, PARAM_NAME_CS, PARAM_DESC_CE, "（なし）",
           fg=C_FONT_GRAY, italic=True, border=B_all())
        r += 1
    else:
        for inp in inputs:
            key = inp.get("key", "")
            is_changed = key in changed_input_keys
            set_h(ws, r, 22)
            c1 = MW(ws, r, PARAM_NAME_CS, PARAM_NAME_CE, key,
                    border=B_all())
            c2 = MW(ws, r, PARAM_TYPE_CS, PARAM_TYPE_CE, inp.get("type", ""),
                    border=B_all(), h="center")
            c3 = MW(ws, r, PARAM_REQ_CS, PARAM_REQ_CE,
                    "○" if inp.get("required") else "", border=B_all(), h="center")
            c4 = MW(ws, r, PARAM_DESC_CS, PARAM_DESC_CE, inp.get("description", ""),
                    border=B_all(), wrap=True)
            if is_changed:
                for c in (c1, c2, c3, c4):
                    dr.apply_red(c)
            r += 1

    # 出力パラメータセクション
    r += 1
    set_h(ws, r, 24)
    MW(ws, r, PARAM_NAME_CS, PARAM_DESC_CE, "■ 出力パラメータ",
       bold=True, fg=C_FONT_W, bg=C_BAND_BLUE, size=11)
    r += 1
    set_h(ws, r, 22)
    MW(ws, r, PARAM_NAME_CS, PARAM_NAME_CE, "パラメータ名",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, r, PARAM_TYPE_CS, PARAM_TYPE_CE, "型",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, r, PARAM_DESC_CS, PARAM_DESC_CE, "説明",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    # 「必須」列は出力側では不要 → 空セル
    MW(ws, r, PARAM_REQ_CS, PARAM_REQ_CE, "",
       bold=True, bg=C_LABEL_BG, border=B_all())
    r += 1

    outputs = data.get("output_params", [])
    if not outputs:
        set_h(ws, r, 20)
        MW(ws, r, PARAM_NAME_CS, PARAM_DESC_CE, "（なし）",
           fg=C_FONT_GRAY, italic=True, border=B_all())
        r += 1
    else:
        for out in outputs:
            key = out.get("key", "")
            is_changed = key in changed_output_keys
            set_h(ws, r, 22)
            c1 = MW(ws, r, PARAM_NAME_CS, PARAM_NAME_CE, key,
                    border=B_all())
            c2 = MW(ws, r, PARAM_TYPE_CS, PARAM_TYPE_CE, out.get("type", ""),
                    border=B_all(), h="center")
            c3 = MW(ws, r, PARAM_REQ_CS, PARAM_REQ_CE, "",
                    border=B_all())
            c4 = MW(ws, r, PARAM_DESC_CS, PARAM_DESC_CE, out.get("description", ""),
                    border=B_all(), wrap=True)
            if is_changed:
                for c in (c1, c2, c3, c4):
                    dr.apply_red(c)
            r += 1


SCALAR_FIELDS   = ["purpose", "overview", "prerequisites", "trigger"]
SECTION_SHEETS  = {"steps": "処理内容",
                   "input_params": "パラメータ定義",
                   "output_params": "パラメータ定義"}


def _compute_diffs(prev_data: dict | None, new_data: dict) -> dict:
    if prev_data is None:
        return {"scalars": [], "lists": {}}
    return {
        "scalars": dr.diff_scalars(prev_data, new_data, SCALAR_FIELDS),
        "lists": {
            "steps":         dr.diff_list(prev_data.get("steps", []),
                                          new_data.get("steps", []), "no"),
            "input_params":  dr.diff_list(prev_data.get("input_params", []),
                                          new_data.get("input_params", []), "key"),
            "output_params": dr.diff_list(prev_data.get("output_params", []),
                                          new_data.get("output_params", []), "key"),
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",      required=True)
    parser.add_argument("--template",   required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--source-file", default="",
                        help="更新時: 既存の設計書xlsxパス")
    parser.add_argument("--version-increment", default="minor",
                        choices=["minor", "major"])
    args = parser.parse_args()

    data = json.loads(Path(args.input).read_text(encoding="utf-8"))
    today = _date.today().strftime("%Y-%m-%d")
    author = data.get("author", "")

    # ── バージョン判定 ────────────────────────────────────────
    is_major    = (args.version_increment == "major")
    source_file = args.source_file.strip()
    prev_meta   = read_meta(source_file) if source_file else None
    prev_data   = prev_meta.get("data") if prev_meta else None

    if prev_meta:
        current_version = increment_version(
            prev_meta.get("version", "1.0"), args.version_increment)
        history    = prev_meta.get("history", [])
        is_initial = False
        print(f"更新モード: {prev_meta.get('version', '?')} → {current_version}"
              + (" (メジャー)" if is_major else ""))
    else:
        current_version = data.get("version") or "1.0"
        history    = []
        is_initial = True
        print(f"新規作成モード: v{current_version}")

    data["version"] = current_version

    diffs = _compute_diffs(prev_data, data)
    if prev_meta and not is_major and not dr.has_any_diff(diffs):
        print("差分なし: 既存ファイルと一致しているため更新をスキップしました")
        sys.exit(0)

    last_no = max((h["項番"] for h in history
                   if isinstance(h.get("項番"), int)), default=0)
    new_entries = dr.build_entries(
        current_version, diffs, author, today,
        start_no=last_no + 1, is_major=is_major, is_initial=is_initial,
        section_sheet_map=SECTION_SHEETS, scalar_sheet="処理概要",
    )
    history = history + new_entries

    # 変更識別子を抽出（赤字用）
    changed_scalars = dr.changed_scalar_fields(diffs)
    changed_steps   = dr.changed_ids(diffs, "steps")
    changed_inputs  = dr.changed_ids(diffs, "input_params")
    changed_outputs = dr.changed_ids(diffs, "output_params")

    # フロー図PNG生成（処理内容の行高さに合わせて縦幅を揃える）
    steps = data.get("steps", [])
    content_pts  = _estimate_content_height_pts(steps)
    target_h_in  = content_pts / 72  # pt → inch
    flowchart_path = None
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp_path = tmp.name
    if generate_flowchart(steps, tmp_path, target_h=target_h_in):
        flowchart_path = tmp_path
    else:
        try: Path(tmp_path).unlink()
        except Exception: pass

    # テンプレ読込→セル流し込み
    wb = load_workbook(args.template)
    fill_revision(wb["改版履歴"],  data, history)
    fill_overview(wb["処理概要"],  data,
                  changed_fields=set() if is_major else changed_scalars)
    fill_process (wb["処理内容"],  data, flowchart_path,
                  changed_step_nos=set() if is_major else changed_steps)
    fill_params  (wb["パラメータ定義"], data,
                  changed_input_keys =set() if is_major else changed_inputs,
                  changed_output_keys=set() if is_major else changed_outputs)

    # _meta 保存（次回差分判定用）
    write_meta(wb, {
        "version": current_version,
        "date":    today,
        "author":  author,
        "data":    data,
        "history": history,
    })

    type_key  = data.get("type", "その他")
    subfolder = TYPE_FOLDER.get(type_key, "other")
    out_dir   = Path(args.output_dir) / subfolder
    out_dir.mkdir(parents=True, exist_ok=True)

    feat_id  = data.get("id", "F-000")
    name     = data.get("name", "機能")
    out_path = out_dir / f"【{feat_id}】{name}.xlsx"
    wb.save(out_path)
    print(f"生成完了: v{current_version} → {out_path}")

    if flowchart_path:
        try: Path(flowchart_path).unlink()
        except Exception: pass


if __name__ == "__main__":
    main()
