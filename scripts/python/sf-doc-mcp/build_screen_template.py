"""
画面設計書テンプレート Excel を生成する（LWC / 画面フロー / Visualforce 用）。

5シート構成:
  1. 改版履歴       : 基本情報 + 改版履歴テーブル
  2. 画面概要       : メタ情報 + 目的/概要/前提/画面遷移/画面イメージ
  3. 画面項目定義   : 項目テーブル（No/項目名/API名/UI種別/型/必須/初期値/バリデーション/備考）
  4. 処理詳細       : ユースケース別ステップ（UC1-UC4 固定枠）
  5. パラメーター定義: @api / CustomEvent / Apex呼び出し

方針:
  - 設計書テンプレート（build_template.py）と同じ狭幅グリッド（列 4.2 × 30本）
  - セル結合でラベルが潰れないように
  - 結合前に全構成セルに罫線付与（途切れ防止）
  - セクション間スペーサー行でゆとりを持たせる

Usage:
  python build_screen_template.py --output "C:/.../画面設計書テンプレート.xlsx"
"""
from __future__ import annotations
import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── カラーパレット ──────────────────────────────────────────────
C_TITLE_DARK = "1F3864"
C_HDR_BLUE   = "2E75B6"
C_BAND_BLUE  = "0070C0"
C_SUB_BAND   = "5B9BD5"    # ユースケース帯
C_LABEL_BG   = "D9E1F2"
C_FONT_W     = "FFFFFF"
C_FONT_D     = "000000"
C_FONT_GRAY  = "595959"

THIN = Side(style="thin",   color="8B9DC3")
MED  = Side(style="medium", color="1F3864")

# ── 共通ヘルパー ───────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _fnt(bold=False, color=C_FONT_D, size=10):
    return Font(name="游ゴシック", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def B_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def B_frame(left=True, right=True, top=True, bottom=True):
    return Border(
        left=MED if left else THIN,
        right=MED if right else THIN,
        top=MED if top else THIN,
        bottom=MED if bottom else THIN,
    )

def W(ws, row, col, value="", bold=False, fg=C_FONT_D, bg=None,
      h="left", v="center", wrap=True, border=None, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _fnt(bold=bold, color=fg, size=size)
    c.alignment = _aln(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if border:
        c.border = border
    return c

def MW(ws, row, cs, ce, value="", border=None, bg=None, **kwargs):
    if border:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = border
    if bg:
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    return W(ws, row, cs, value, border=border, bg=bg, **kwargs)

def set_h(ws, row, h):
    ws.row_dimensions[row].height = h

def set_w(ws, letter, w):
    ws.column_dimensions[letter].width = w


# ── 共通グリッド ───────────────────────────────────────────────
GRID_LEFT  = 2
GRID_RIGHT = 31
COL_W      = 4.2

def setup_grid(ws):
    set_w(ws, "A", 2.0)
    for i in range(GRID_LEFT, GRID_RIGHT + 1):
        set_w(ws, get_column_letter(i), COL_W)
    ws.sheet_view.showGridLines = False

def title_band(ws, row, text, height=36):
    set_h(ws, row, height)
    MW(ws, row, GRID_LEFT, GRID_RIGHT, text,
       bold=True, fg=C_FONT_W, bg=C_TITLE_DARK, h="center", size=14)

def section_band(ws, row, text, height=26, bg=C_BAND_BLUE,
                 cs=None, ce=None):
    cs = cs if cs is not None else GRID_LEFT
    ce = ce if ce is not None else GRID_RIGHT
    set_h(ws, row, height)
    MW(ws, row, cs, ce, text,
       bold=True, fg=C_FONT_W, bg=bg, size=11)


# ─────────────────────────────────────────────────────────────
# Sheet 1: 改版履歴
# ─────────────────────────────────────────────────────────────
REV_META_ROW       = 3
REV_TABLE_HEAD_ROW = 5
REV_DATA_ROW_START = 6
REV_DATA_ROW_END   = 25

REV_META_PROJECT = (2, 6, 7, 18)     # label cs,ce / value cs,ce
REV_META_DATE    = (19, 22, 23, 31)

REV_COLS = {
    "項番":     (2,  3),
    "版数":     (4,  5),
    "変更箇所": (6, 11),
    "変更内容": (12, 17),
    "変更理由": (18, 23),
    "変更日":   (24, 26),
    "変更者":   (27, 29),
    "備考":     (30, 31),
}

def build_revision(wb):
    ws = wb.active
    ws.title = "改版履歴"
    setup_grid(ws)

    title_band(ws, 1, "改版履歴")
    set_h(ws, 2, 6)

    set_h(ws, REV_META_ROW, 22)
    ls, le, vs, ve = REV_META_PROJECT
    MW(ws, REV_META_ROW, ls, le, "プロジェクト名",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, REV_META_ROW, vs, ve, "", border=B_all())

    ls, le, vs, ve = REV_META_DATE
    MW(ws, REV_META_ROW, ls, le, "作成日",
       bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
    MW(ws, REV_META_ROW, vs, ve, "", border=B_all())

    set_h(ws, 4, 10)

    set_h(ws, REV_TABLE_HEAD_ROW, 24)
    for label, (cs, ce) in REV_COLS.items():
        MW(ws, REV_TABLE_HEAD_ROW, cs, ce, label,
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center",
           border=B_all())

    for r in range(REV_DATA_ROW_START, REV_DATA_ROW_END + 1):
        set_h(ws, r, 22)
        for label, (cs, ce) in REV_COLS.items():
            MW(ws, r, cs, ce, "", border=B_all())


# ─────────────────────────────────────────────────────────────
# Sheet 2: 画面概要
# ─────────────────────────────────────────────────────────────
OV_META_ROW_1 = 3
OV_META_ROW_2 = 4

OV_META_1 = [
    ("プロジェクト名", 2, 5,  6, 14),
    ("システム名",    15, 17, 18, 22),
    ("画面名",        23, 25, 26, 31),
]
OV_META_2 = [
    ("API名",      2, 5,  6, 14),
    ("作成者",    15, 17, 18, 20),
    ("作成日",    21, 23, 24, 26),
    ("バージョン", 27, 28, 29, 31),
]

# セクション: (タイトル, ヘッダ行, データ行, データ行高)
OV_SECTIONS = [
    ("1. 本書の目的",                     6,  7,  70),
    ("2. 画面概要（どんな画面か・対象ユーザー）", 9,  10, 100),
    ("3. 主要機能（この画面でできること）",   12, 13, 140),
    ("4. 前提条件",                       15, 16, 90),
    ("5. 画面遷移（遷移元 / 遷移先）",      18, 19, 80),
]

# 画面イメージ画像貼付エリア（セクションタイトル + 大枠）
OV_IMG_SEC_ROW    = 21
OV_IMG_DATA_START = 22
OV_IMG_DATA_ROWS  = 20

def build_overview(wb):
    ws = wb.create_sheet("画面概要")
    setup_grid(ws)

    title_band(ws, 1, "画面概要")
    set_h(ws, 2, 6)

    set_h(ws, OV_META_ROW_1, 22)
    for label, ls, le, vs, ve in OV_META_1:
        MW(ws, OV_META_ROW_1, ls, le, label,
           bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
        MW(ws, OV_META_ROW_1, vs, ve, "", border=B_all())

    set_h(ws, OV_META_ROW_2, 22)
    for label, ls, le, vs, ve in OV_META_2:
        MW(ws, OV_META_ROW_2, ls, le, label,
           bold=True, bg=C_LABEL_BG, border=B_all(), h="center")
        MW(ws, OV_META_ROW_2, vs, ve, "", border=B_all())

    set_h(ws, 5, 14)

    prev_data_row = None
    for label, hdr_row, data_row, data_h in OV_SECTIONS:
        if prev_data_row is not None:
            for spacer in range(prev_data_row + 1, hdr_row):
                set_h(ws, spacer, 12)
        section_band(ws, hdr_row, label)
        set_h(ws, data_row, data_h)
        MW(ws, data_row, GRID_LEFT, GRID_RIGHT, "",
           border=B_frame(), wrap=True, v="top")
        prev_data_row = data_row

    # 画面イメージ（画像貼付領域）
    for spacer in range(prev_data_row + 1, OV_IMG_SEC_ROW):
        set_h(ws, spacer, 12)
    section_band(ws, OV_IMG_SEC_ROW, "6. 画面イメージ（ワイヤーフレーム / キャプチャ）")
    # 画像貼付枠: 複数行の矩形枠
    for r in range(OV_IMG_DATA_START, OV_IMG_DATA_START + OV_IMG_DATA_ROWS):
        set_h(ws, r, 22)
        for c in range(GRID_LEFT, GRID_RIGHT + 1):
            ws.cell(row=r, column=c).border = B_all()
    # 外枠（太線）
    top_r = OV_IMG_DATA_START
    bot_r = OV_IMG_DATA_START + OV_IMG_DATA_ROWS - 1
    for c in range(GRID_LEFT, GRID_RIGHT + 1):
        ws.cell(row=top_r, column=c).border = Border(
            top=MED,
            bottom=THIN,
            left=MED if c == GRID_LEFT else THIN,
            right=MED if c == GRID_RIGHT else THIN)
        ws.cell(row=bot_r, column=c).border = Border(
            top=THIN,
            bottom=MED,
            left=MED if c == GRID_LEFT else THIN,
            right=MED if c == GRID_RIGHT else THIN)


# ─────────────────────────────────────────────────────────────
# Sheet 3: 画面項目定義
# ─────────────────────────────────────────────────────────────
IT_META_ROW_1 = 3
IT_META_ROW_2 = 4

IT_META_1 = [
    ("プロジェクト名", 2, 5,  6, 14),
    ("画面名",        15, 17, 18, 22),
    ("API名",         23, 25, 26, 31),
]
IT_META_2 = [
    ("作成者",    2, 5,  6, 14),
    ("作成日",   15, 17, 18, 22),
    ("バージョン", 23, 25, 26, 31),
]

IT_SEC_ROW  = 3
IT_HEAD_ROW = 4
IT_DATA_ROW_START = 5
IT_DATA_ROWS      = 30

IT_COLS = {
    "No":              (2,  3),
    "項目名":          (4,  8),
    "API名/プロパティ":(9,  14),
    "UI種別":          (15, 17),
    "型":              (18, 20),
    "必須":            (21, 22),
    "初期値":          (23, 25),
    "バリデーション":   (26, 28),
    "備考":            (29, 31),
}

def build_items(wb):
    ws = wb.create_sheet("画面項目定義")
    setup_grid(ws)

    title_band(ws, 1, "画面項目定義")
    set_h(ws, 2, 6)

    section_band(ws, IT_SEC_ROW, "■ 項目一覧")

    set_h(ws, IT_HEAD_ROW, 26)
    for label, (cs, ce) in IT_COLS.items():
        MW(ws, IT_HEAD_ROW, cs, ce, label,
           bold=True, bg=C_HDR_BLUE, fg=C_FONT_W, h="center",
           border=B_all())

    for r in range(IT_DATA_ROW_START, IT_DATA_ROW_START + IT_DATA_ROWS):
        set_h(ws, r, 24)
        for label, (cs, ce) in IT_COLS.items():
            MW(ws, r, cs, ce, "", border=B_all())


# ─────────────────────────────────────────────────────────────
# Sheet 4: 処理詳細（ユースケース別）
# ─────────────────────────────────────────────────────────────
LG_META_ROW_1 = 3
LG_META_ROW_2 = 4

LG_META_1 = [
    ("プロジェクト名", 2, 5,  6, 14),
    ("画面名",        15, 17, 18, 22),
    ("API名",         23, 25, 26, 31),
]
LG_META_2 = [
    ("作成者",    2, 5,  6, 14),
    ("作成日",   15, 17, 18, 22),
    ("バージョン", 23, 25, 26, 31),
]

LG_COLS = {
    "No":                (2,  3),
    "対象項目/要素":     (4,  9),
    "処理内容":          (10, 18),
    "呼び出しAPI/アクション": (19, 23),
    "成功時/失敗時の挙動":   (24, 28),
    "備考":              (29, 31),
}

def build_logic(wb):
    """処理詳細シート。タイトルのみ配置し、本文は generator が動的に書き込む。"""
    ws = wb.create_sheet("処理詳細")
    setup_grid(ws)

    title_band(ws, 1, "処理詳細")
    set_h(ws, 2, 6)
    # 本文（UCブロック）は generate_screen_design.py が書き込む


# ─────────────────────────────────────────────────────────────
# Sheet 5: パラメーター定義
# ─────────────────────────────────────────────────────────────
PM_META_ROW_1 = 3
PM_META_ROW_2 = 4

PM_META_1 = [
    ("プロジェクト名", 2, 5,  6, 14),
    ("画面名",        15, 17, 18, 22),
    ("API名",         23, 25, 26, 31),
]
PM_META_2 = [
    ("作成者",    2, 5,  6, 14),
    ("作成日",   15, 17, 18, 22),
    ("バージョン", 23, 25, 26, 31),
]

PM_COLS = {
    "No":            (2,  3),
    "Key / 名称":    (4,  10),
    "型":            (11, 14),
    "必須":          (15, 16),
    "説明":          (17, 25),
    "デフォルト":    (26, 28),
    "備考":          (29, 31),
}

def build_params(wb):
    """パラメーター定義シート。タイトルのみ配置し、本文は generator が動的に書き込む。"""
    ws = wb.create_sheet("パラメーター定義")
    setup_grid(ws)

    title_band(ws, 1, "パラメーター定義")
    set_h(ws, 2, 6)
    # 本文（セクションブロック）は generate_screen_design.py が書き込む


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    wb = Workbook()
    build_revision(wb)
    build_overview(wb)
    build_items(wb)
    build_logic(wb)
    build_params(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"画面設計書テンプレート生成完了: {out}")


if __name__ == "__main__":
    main()
